import os, sqlite3, html, urllib.parse, subprocess, platform, mimetypes
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import base64
import json
import re
from io import BytesIO
from datetime import datetime
from functools import wraps
from flask import Flask, request, redirect, url_for, session, render_template_string, flash, jsonify, send_from_directory, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
try:
    import extract_msg
except Exception:
    extract_msg = None

try:
    import psycopg
    from psycopg.rows import dict_row
except Exception:
    psycopg = None
    dict_row = None


APP_NAME = "PMW Ticket + Fabrication"
APP_VERSION = "v51.2 Auto Backup Disabled Fix"
APP_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(APP_DIR, "pmw_schedule.db")
UPLOAD_FOLDER = os.path.join(APP_DIR, "uploads")
EXPORT_FOLDER = os.path.join(APP_DIR, "exports")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXPORT_FOLDER, exist_ok=True)
CLOUD_TICKET_FOLDER = os.path.join(APP_DIR, 'cloud_ticket_files')
os.makedirs(CLOUD_TICKET_FOLDER, exist_ok=True)
CLOUD_ATTACHMENT_FOLDER = os.path.join(APP_DIR, 'cloud_ticket_attachments')
os.makedirs(CLOUD_ATTACHMENT_FOLDER, exist_ok=True)

DISPLAY_COLS = [1,2,3,4,5,6]
ROLE_LEVEL = {"viewer":1,"editor":2,"admin":3}


def auto_upgrade_sqlite_schema():
    """Safe startup upgrade for existing SQLite DBs on Render/free hosting.
    Adds newer columns that older DB files may be missing.
    """
    try:
        con = sqlite3.connect(DB_PATH)
        cur = con.cursor()

        cur.execute("""CREATE TABLE IF NOT EXISTS users(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password_hash TEXT,
            role TEXT,
            active INTEGER DEFAULT 1,
            created_at TEXT,
            link_path TEXT DEFAULT '',
            link_label TEXT DEFAULT ''
        )""")

        cur.execute("""CREATE TABLE IF NOT EXISTS workbook_cells(
            sheet_name TEXT,
            row_num INTEGER,
            col_num INTEGER,
            value TEXT DEFAULT '',
            updated_by TEXT DEFAULT '',
            updated_at TEXT DEFAULT '',
            PRIMARY KEY(sheet_name,row_num,col_num)
        )""")

        for coldef in [
            "bg_color TEXT DEFAULT ''",
            "text_color TEXT DEFAULT ''",
            "link_path TEXT DEFAULT ''",
            "link_label TEXT DEFAULT ''",
            "font_size TEXT DEFAULT ''",
            "bold TEXT DEFAULT ''",
            "rich_html TEXT DEFAULT ''"
        ]:
            try:
                cur.execute("ALTER TABLE workbook_cells ADD COLUMN " + coldef)
            except sqlite3.OperationalError:
                pass

        cur.execute("""CREATE TABLE IF NOT EXISTS ticket_links(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_number TEXT,
            subject TEXT,
            sender TEXT,
            received TEXT,
            file_path TEXT UNIQUE,
            created_at TEXT
        )""")

        cur.execute("""CREATE TABLE IF NOT EXISTS audit_log(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT,
            action TEXT,
            details TEXT,
            created_at TEXT
        )""")

        con.commit()
        con.close()
        print("SQLite schema auto-upgrade complete:", DB_PATH)
    except Exception as e:
        print("SQLite schema auto-upgrade failed:", repr(e))

auto_upgrade_sqlite_schema()


app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "pmw-local-dev-secret")
PMW_UPLOAD_KEY = os.environ.get("PMW_UPLOAD_KEY", "pmw-upload-dev-key")

DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = "postgresql://" + DATABASE_URL[len("postgres://"):]
USE_POSTGRES = bool(DATABASE_URL and psycopg)

class PgCursor:
    def __init__(self, cur):
        self.cur = cur

    def _convert_query(self, query):
        q = query.strip()
        upper = q.upper()

        # SQLite upsert compatibility for workbook cells
        if upper.startswith("INSERT OR REPLACE INTO WORKBOOK_CELLS"):
            q = q.replace("INSERT OR REPLACE INTO workbook_cells", "INSERT INTO workbook_cells")
            if "ON CONFLICT" not in q:
                q += """ ON CONFLICT(sheet_name,row_num,col_num) DO UPDATE SET
                    value=EXCLUDED.value,
                    bg_color=EXCLUDED.bg_color,
                    text_color=EXCLUDED.text_color,
                    link_path=EXCLUDED.link_path,
                    link_label=EXCLUDED.link_label,
                    font_size=EXCLUDED.font_size,
                    bold=EXCLUDED.bold,
                    rich_html=EXCLUDED.rich_html,
                    updated_by=EXCLUDED.updated_by,
                    updated_at=EXCLUDED.updated_at"""
        elif upper.startswith("INSERT OR REPLACE INTO"):
            q = q.replace("INSERT OR REPLACE INTO", "INSERT INTO")

        return q.replace("?", "%s")

    def execute(self, query, params=()):
        self.cur.execute(self._convert_query(query), params or ())
        return self

    def fetchone(self):
        return self.cur.fetchone()

    def fetchall(self):
        return self.cur.fetchall()

    @property
    def lastrowid(self):
        try:
            self.cur.execute("SELECT LASTVAL() AS id")
            row = self.cur.fetchone()
            return row["id"] if row else None
        except Exception:
            return None

class PgConnection:
    def __init__(self, con):
        self.con = con

    def cursor(self):
        return PgCursor(self.con.cursor(row_factory=dict_row))

    def execute(self, query, params=()):
        cur = self.cursor()
        return cur.execute(query, params)

    def commit(self):
        self.con.commit()

    def close(self):
        self.con.close()


def db():
    if USE_POSTGRES:
        con = psycopg.connect(DATABASE_URL)
        return PgConnection(con)
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con

def init_db():
    con=db(); cur=con.cursor()

    if USE_POSTGRES:
        cur.execute("""CREATE TABLE IF NOT EXISTS users(
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE,
            password_hash TEXT,
            role TEXT,
            active INTEGER DEFAULT 1,
            created_at TEXT
        )""")
        cur.execute("""CREATE TABLE IF NOT EXISTS workbook_cells(
            sheet_name TEXT,
            row_num INTEGER,
            col_num INTEGER,
            value TEXT DEFAULT '',
            updated_by TEXT DEFAULT '',
            updated_at TEXT DEFAULT '',
            PRIMARY KEY(sheet_name,row_num,col_num)
        )""")
        for coldef in [
            "bg_color TEXT DEFAULT ''",
            "text_color TEXT DEFAULT ''",
            "link_path TEXT DEFAULT ''",
            "link_label TEXT DEFAULT ''",
            "font_size TEXT DEFAULT ''",
            "bold TEXT DEFAULT ''",
            "rich_html TEXT DEFAULT ''"
        ]:
            try:
                cur.execute("ALTER TABLE workbook_cells ADD COLUMN " + coldef)
            except Exception:
                try:
                    con.con.rollback()
                except Exception:
                    pass

        cur.execute("""CREATE TABLE IF NOT EXISTS ticket_links(
            id SERIAL PRIMARY KEY,
            job_number TEXT,
            subject TEXT,
            sender TEXT,
            received TEXT,
            file_path TEXT UNIQUE,
            created_at TEXT
        )""")
        cur.execute("""CREATE TABLE IF NOT EXISTS audit_log(
            id SERIAL PRIMARY KEY,
            username TEXT,
            action TEXT,
            details TEXT,
            created_at TEXT
        )""")
    else:
        cur.execute("""CREATE TABLE IF NOT EXISTS users(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password_hash TEXT,
            role TEXT,
            active INTEGER DEFAULT 1,
            created_at TEXT
        )""")
        cur.execute("""CREATE TABLE IF NOT EXISTS workbook_cells(
            sheet_name TEXT,
            row_num INTEGER,
            col_num INTEGER,
            value TEXT DEFAULT '',
            updated_by TEXT DEFAULT '',
            updated_at TEXT DEFAULT '',
            PRIMARY KEY(sheet_name,row_num,col_num)
        )""")
        for coldef in [
            "bg_color TEXT DEFAULT ''",
            "text_color TEXT DEFAULT ''",
            "link_path TEXT DEFAULT ''",
            "link_label TEXT DEFAULT ''",
            "font_size TEXT DEFAULT ''",
            "bold TEXT DEFAULT ''",
            "rich_html TEXT DEFAULT ''"
        ]:
            try:
                cur.execute("ALTER TABLE workbook_cells ADD COLUMN " + coldef)
            except sqlite3.OperationalError:
                pass

        cur.execute("""CREATE TABLE IF NOT EXISTS ticket_links(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_number TEXT,
            subject TEXT,
            sender TEXT,
            received TEXT,
            file_path TEXT UNIQUE,
            created_at TEXT
        )""")
        cur.execute("""CREATE TABLE IF NOT EXISTS audit_log(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT,
            action TEXT,
            details TEXT,
            created_at TEXT
        )""")

    for u,p,r in [("admin","admin123","admin"),("shop","shop123","editor"),("viewer","view123","viewer")]:
        if not cur.execute("SELECT id FROM users WHERE username=?",(u,)).fetchone():
            cur.execute(
                "INSERT INTO users(username,password_hash,role,active,created_at) VALUES(?,?,?,?,?)",
                (u,generate_password_hash(p),r,1,datetime.now().isoformat(timespec='seconds'))
            )

    con.commit(); con.close()

def log(action, details=""):

    con=db(); con.execute("INSERT INTO audit_log(username,action,details,created_at) VALUES(?,?,?,?)",(session.get('username','system'),action,details,datetime.now().isoformat(timespec='seconds'))); con.commit(); con.close()

def can(role): return ROLE_LEVEL.get(session.get('role',''),0) >= ROLE_LEVEL[role]

def login_required(fn):
    @wraps(fn)
    def w(*a,**k):
        if not session.get('user_id'): return redirect(url_for('login'))
        return fn(*a,**k)
    return w

def role_required(role):
    def deco(fn):
        @wraps(fn)
        def w(*a,**k):
            if not can(role):
                flash("You do not have permission for that.")
                return redirect(url_for('index'))
            return fn(*a,**k)
        return w
    return deco

def clean(v):
    if v is None: return ""
    if hasattr(v, "strftime"): return v.strftime("%m/%d/%Y")
    return str(v)

def import_workbook(path):
    wb = load_workbook(path, data_only=True, keep_vba=True)
    con=db(); cur=con.cursor(); cur.execute("DELETE FROM workbook_cells")
    count=0
    for ws in wb.worksheets:
        if ws.title.lower().startswith('chart'): continue
        sheet = 'Fabrication Schedule' if ws.title == 'Sheet1' else ws.title
        for r in range(1, max(50, min(ws.max_row or 50, 90))+1):
            for c in DISPLAY_COLS:
                v = clean(ws.cell(r,c).value)
                cur.execute("INSERT OR REPLACE INTO workbook_cells(sheet_name,row_num,col_num,value,updated_at) VALUES(?,?,?,?,?)",(sheet,r,c,v,datetime.now().isoformat(timespec='seconds')))
                if v: count += 1
    con.commit(); con.close(); log("IMPORT_WORKBOOK", os.path.basename(path)); return count

def sheet_names():
    con=db()
    if USE_POSTGRES:
        rows=con.execute("""
            SELECT sheet_name
            FROM workbook_cells
            GROUP BY sheet_name
            ORDER BY MIN(CASE WHEN sheet_name='Fabrication Schedule' THEN 0 ELSE 1 END), sheet_name
        """).fetchall()
    else:
        rows=con.execute("SELECT DISTINCT sheet_name FROM workbook_cells ORDER BY CASE WHEN sheet_name='Fabrication Schedule' THEN 0 ELSE 1 END, sheet_name").fetchall()
    con.close()
    names=[r['sheet_name'] for r in rows]
    return names or ['Fabrication Schedule']

def cells_for(sheet):
    con=db(); rows=con.execute("SELECT row_num,col_num,value FROM workbook_cells WHERE sheet_name=?",(sheet,)).fetchall(); con.close()
    d={(r['row_num'],r['col_num']):r['value'] or '' for r in rows}
    for r in range(1,51):
        for c in DISPLAY_COLS: d.setdefault((r,c),"")
    return d

def cell_meta_for(sheet):
    con=db()
    rows=con.execute("SELECT row_num,col_num,value,bg_color,text_color,link_path,link_label,font_size,bold,rich_html FROM workbook_cells WHERE sheet_name=?",(sheet,)).fetchall()
    con.close()
    d={}
    for r in rows:
        d[(r['row_num'],r['col_num'])]={
            "value": r["value"] or "",
            "bg_color": r["bg_color"] or "",
            "text_color": r["text_color"] or "",
            "link_path": r["link_path"] or "",
            "link_label": r["link_label"] or "",
            "font_size": r["font_size"] or "",
            "bold": r["bold"] or "",
            "rich_html": r["rich_html"] or ""
        }
    for rr in range(1,51):
        for cc in DISPLAY_COLS:
            d.setdefault((rr,cc), {"value":"","bg_color":"","text_color":"","link_path":"","link_label":"","font_size":"","bold":"","rich_html":""})
    return d

def ticket_rows_from_drop(path):
    """Reads PMW Ticket emails drop.xlsx.
    Columns:
    B received, C sender, D subject, E job #, G saved .msg path
    """
    wb = load_workbook(path, data_only=False)
    ws = wb.active
    rows=[]
    for r in range(2, (ws.max_row or 1)+1):
        received = clean(ws.cell(r,2).value)
        sender = clean(ws.cell(r,3).value)
        subject = clean(ws.cell(r,4).value)
        job = clean(ws.cell(r,5).value).strip()
        file_path = clean(ws.cell(r,7).value).strip()
        if not file_path:
            h = ws.cell(r,6).hyperlink
            if h:
                file_path = h.target or ""
        if job or subject or file_path:
            rows.append({
                "source_row": r,
                "received": received,
                "sender": sender,
                "subject": subject,
                "job_number": job,
                "file_path": file_path
            })
    return rows

def import_one_ticket_from_drop(path, mode="newest"):
    rows = ticket_rows_from_drop(path)
    if not rows:
        return None, "No ticket rows found."

    con=db(); cur=con.cursor()
    imported = None

    # newest means last non-duplicate row in the workbook.
    candidates = list(reversed(rows)) if mode == "newest" else rows
    for t in candidates:
        fp = (t.get("file_path") or "").strip()
        if not fp:
            continue
        exists = cur.execute("SELECT id FROM ticket_links WHERE file_path=?",(fp,)).fetchone()
        if exists:
            continue
        now=datetime.now().isoformat(timespec='seconds')
        cur.execute("""INSERT INTO ticket_links(job_number,subject,sender,received,file_path,created_at)
                       VALUES(?,?,?,?,?,?)""",
                    (t.get("job_number",""), t.get("subject",""), t.get("sender",""), t.get("received",""), fp, now))
        tid = cur.lastrowid
        con.commit()
        imported = cur.execute("SELECT * FROM ticket_links WHERE id=?",(tid,)).fetchone()
        break

    con.close()
    if imported:
        log("IMPORT_ONE_TICKET", imported["file_path"] or "")
        return imported, ""
    return None, "No new ticket found. The newest rows may already be imported."


def import_last_n_tickets_from_drop(path, count=1):
    """Import the newest N non-duplicate ticket rows from the ticket drop workbook."""
    try:
        count = int(count)
    except Exception:
        count = 1
    count = max(1, min(count, 25))

    rows = ticket_rows_from_drop(path)
    if not rows:
        return [], "No ticket rows found."

    con=db(); cur=con.cursor()
    imported=[]
    now=datetime.now().isoformat(timespec='seconds')

    for t in reversed(rows):
        if len(imported) >= count:
            break
        fp = (t.get("file_path") or "").strip()
        if not fp:
            continue
        exists = cur.execute("SELECT id FROM ticket_links WHERE file_path=?",(fp,)).fetchone()
        if exists:
            continue
        cur.execute("""INSERT INTO ticket_links(job_number,subject,sender,received,file_path,created_at)
                       VALUES(?,?,?,?,?,?)""",
                    (t.get("job_number",""), t.get("subject",""), t.get("sender",""), t.get("received",""), fp, now))
        tid = cur.lastrowid
        imported.append(cur.execute("SELECT * FROM ticket_links WHERE id=?",(tid,)).fetchone())

    con.commit(); con.close()
    if imported:
        log("IMPORT_LAST_TICKETS", f"{len(imported)} ticket(s)")
        return imported, ""
    return [], "No new tickets found. The newest rows may already be imported."

def save_selected_ticket_rows_from_drop(path, selected_rows):
    """Import only checked row numbers from the ticket drop workbook."""
    wanted=set()
    for x in selected_rows:
        try:
            wanted.add(int(x))
        except Exception:
            pass

    rows=ticket_rows_from_drop(path)
    if not wanted:
        return [], "No ticket rows were selected."

    con=db(); cur=con.cursor()
    imported=[]
    skipped=0
    now=datetime.now().isoformat(timespec='seconds')

    for t in rows:
        if int(t.get("source_row", 0)) not in wanted:
            continue
        fp=(t.get("file_path") or "").strip()
        if not fp:
            skipped += 1
            continue
        exists=cur.execute("SELECT id FROM ticket_links WHERE file_path=?",(fp,)).fetchone()
        if exists:
            skipped += 1
            continue
        cur.execute("""INSERT INTO ticket_links(job_number,subject,sender,received,file_path,created_at)
                       VALUES(?,?,?,?,?,?)""",
                    (t.get("job_number",""), t.get("subject",""), t.get("sender",""), t.get("received",""), fp, now))
        tid=cur.lastrowid
        imported.append(cur.execute("SELECT * FROM ticket_links WHERE id=?",(tid,)).fetchone())

    con.commit(); con.close()
    if imported:
        log("IMPORT_SELECTED_TICKETS", f"{len(imported)} ticket(s), skipped {skipped}")
        return imported, f"Imported {len(imported)} selected ticket(s)."
    return [], f"No new tickets imported. {skipped} selected row(s) may already be imported or missing a saved .msg path."

def first_empty_schedule_row(sheet, col_num):
    con=db()
    for r in range(3,51):
        row=con.execute("SELECT value FROM workbook_cells WHERE sheet_name=? AND row_num=? AND col_num=?",(sheet,r,col_num)).fetchone()
        if not row or not (row["value"] or "").strip():
            con.close()
            return r
    con.close()
    return 50

def add_ticket_to_schedule(ticket_id, side, sheet="Fabrication Schedule"):
    con=db(); cur=con.cursor()
    t=cur.execute("SELECT * FROM ticket_links WHERE id=?",(ticket_id,)).fetchone()
    if not t:
        con.close()
        return False, "Ticket not found."

    if side == "numbering":
        sort_col, job_col, note_col = 1,2,3
    else:
        sort_col, job_col, note_col = 4,5,6

    r = first_empty_schedule_row(sheet, job_col)
    text = (t["subject"] or t["job_number"] or "Ticket").strip()
    now=datetime.now().isoformat(timespec='seconds')
    cloud_file = ''
    try:
        cloud_file = t['cloud_file'] or ''
    except Exception:
        cloud_file = ''
    path = ("/ticket_view_email/" + str(ticket_id)) if cloud_file.strip() else (t["file_path"] or "")
    label = (t["subject"] or t["job_number"] or "Ticket")[:80]

    for c, val in [(job_col, text), (note_col, "TICKET")]:
        cur.execute("""INSERT OR REPLACE INTO workbook_cells(
            sheet_name,row_num,col_num,value,bg_color,text_color,link_path,link_label,font_size,bold,rich_html,updated_by,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (sheet,r,c,val,"#fff066" if c==job_col else "", "", path if c==job_col else "", label if c==job_col else "", "", "", "", session.get("username",""), now))
    con.commit(); con.close()
    log("ADD_TICKET_TO_SCHEDULE", f"{side} row {r}: {text}")
    return True, f"Ticket added to {side} row {r}."

def open_path_on_this_pc(path):
    path=(path or "").strip()
    if not path:
        return False, "No path saved."
    try:
        if platform.system().lower().startswith("win"):
            os.startfile(path)
        else:
            reveal_file(path)
        return True, "Opened."
    except Exception as e:
        return False, str(e)



def upsert_workbook_cell_cur(cur, sheet, r, c, val='', bg='', txt='', link='', label='', fsize='', bold='', rich='', user=''):
    now = datetime.now().isoformat(timespec='seconds')
    if USE_POSTGRES:
        cur.execute("""
            INSERT INTO workbook_cells(
                sheet_name,row_num,col_num,value,bg_color,text_color,link_path,link_label,font_size,bold,rich_html,updated_by,updated_at
            )
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(sheet_name,row_num,col_num) DO UPDATE SET
                value=EXCLUDED.value,
                bg_color=EXCLUDED.bg_color,
                text_color=EXCLUDED.text_color,
                link_path=EXCLUDED.link_path,
                link_label=EXCLUDED.link_label,
                font_size=EXCLUDED.font_size,
                bold=EXCLUDED.bold,
                rich_html=EXCLUDED.rich_html,
                updated_by=EXCLUDED.updated_by,
                updated_at=EXCLUDED.updated_at
        """, (sheet,r,c,val,bg,txt,link,label,fsize,bold,rich,user,now))
    else:
        cur.execute("""INSERT OR REPLACE INTO workbook_cells(
            sheet_name,row_num,col_num,value,bg_color,text_color,link_path,link_label,font_size,bold,rich_html,updated_by,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",(sheet,r,c,val,bg,txt,link,label,fsize,bold,rich,user,now))
    return now


def upsert_workbook_cell(sheet, r, c, val='', bg='', txt='', link='', label='', fsize='', bold='', rich='', user=''):
    con = db()
    cur = con.cursor()
    now = upsert_workbook_cell_cur(cur, sheet, r, c, val, bg, txt, link, label, fsize, bold, rich, user)
    con.commit()
    con.close()
    return now


def save_posted_cells(sheet):
    con = db()
    cur = con.cursor()
    user = session.get('username','')
    try:
        for r in range(3,51):
            for c in DISPLAY_COLS:
                key=f"cell_{r}_{c}"
                if key in request.form:
                    val=(request.form.get(key) or '').replace('\r',' ').replace('\n',' ').strip()

                    # Blank visible cell clears everything including ticket links/icons.
                    if val == '':
                        upsert_workbook_cell_cur(cur, sheet,r,c,'','','','','','','','',user)
                        continue

                    bg=(request.form.get(f"bg_{r}_{c}") or '').strip()
                    txt=(request.form.get(f"txt_{r}_{c}") or '').strip()
                    link=(request.form.get(f"link_{r}_{c}") or '').strip()
                    label=(request.form.get(f"label_{r}_{c}") or '').strip()
                    fsize=(request.form.get(f"fsize_{r}_{c}") or '').strip()
                    bold=(request.form.get(f"bold_{r}_{c}") or '').strip()
                    rich=(request.form.get(f"rich_{r}_{c}") or '').strip()
                    upsert_workbook_cell_cur(cur, sheet,r,c,val,bg,txt,link,label,fsize,bold,rich,user)
        con.commit()
    finally:
        con.close()


def sort_side(sheet, key_col, job_col, note_col):
    """V44 SIMPLE LOCKED ROW SORT

    This intentionally does only one simple thing:
    - Lock 3 cells together from each row:
        key/order number + job/description/email-link + note/done
    - Sort those locked rows by the key/order number.
    - Delete the old side rows.
    - Reinsert the locked rows.

    It does NOT renumber.
    It does NOT sort by job name.
    It does NOT reuse destination-row numbers.
    """
    side_cols = [key_col, job_col, note_col]
    user = session.get('username','')
    con = db()
    cur = con.cursor()

    def cell_dict(row):
        if not row:
            return {
                "value": "",
                "bg_color": "",
                "text_color": "",
                "link_path": "",
                "link_label": "",
                "font_size": "",
                "bold": "",
                "rich_html": ""
            }
        return {
            "value": row["value"] or "",
            "bg_color": row["bg_color"] or "",
            "text_color": row["text_color"] or "",
            "link_path": row["link_path"] or "",
            "link_label": row["link_label"] or "",
            "font_size": row["font_size"] or "",
            "bold": row["bold"] or "",
            "rich_html": row["rich_html"] or ""
        }

    def get_cell(r, c):
        row = cur.execute("""SELECT value,bg_color,text_color,link_path,link_label,font_size,bold,rich_html
                             FROM workbook_cells
                             WHERE sheet_name=? AND row_num=? AND col_num=?""",
                          (sheet, r, c)).fetchone()
        return cell_dict(row)

    def has_data(group):
        for cell in group:
            if (
                cell["value"].strip()
                or cell["bg_color"].strip()
                or cell["text_color"].strip()
                or cell["link_path"].strip()
                or cell["link_label"].strip()
                or cell["font_size"].strip()
                or cell["bold"].strip()
                or cell["rich_html"].strip()
            ):
                return True
        return False

    def numeric_key(v):
        raw = str(v or "").strip().replace("\xa0", " ")
        # Extract normal numeric characters only. This handles accidental spaces or formatting.
        cleaned = "".join(ch for ch in raw if ch.isdigit() or ch in ".-")
        if cleaned in ("", ".", "-", "-."):
            return (1, 999999, raw.lower())
        try:
            return (0, float(cleaned), raw.lower())
        except Exception:
            return (1, 999999, raw.lower())

    try:
        locked_rows = []

        # 1) Read complete row groups exactly as they exist now.
        for r in range(3, 51):
            group = [
                get_cell(r, key_col),
                get_cell(r, job_col),
                get_cell(r, note_col)
            ]
            if has_data(group):
                # Stable tiebreaker is original row number, not job text.
                locked_rows.append((numeric_key(group[0]["value"]), r, group))

        # 2) Sort only by the number cell; same numbers keep original order.
        locked_rows.sort(key=lambda x: (x[0], x[1]))

        # 3) Delete the old cells on this side entirely to prevent stuck destination-row numbers.
        placeholders = ",".join(["?"] * len(side_cols))
        params = [sheet] + side_cols
        try:
            cur.execute(
                f"DELETE FROM workbook_cells WHERE sheet_name=? AND row_num>=3 AND row_num<=50 AND col_num IN ({placeholders})",
                tuple(params)
            )
        except Exception:
            # Fallback clear if delete is blocked for any reason.
            for r in range(3, 51):
                for c in side_cols:
                    upsert_workbook_cell_cur(cur, sheet, r, c, '', '', '', '', '', '', '', '', user)

        # 4) Reinsert sorted locked groups. Number stays attached to the job.
        target_r = 3
        for _, original_r, group in locked_rows:
            for idx, c in enumerate(side_cols):
                cell = group[idx]
                upsert_workbook_cell_cur(
                    cur, sheet, target_r, c,
                    cell["value"],
                    cell["bg_color"],
                    cell["text_color"],
                    cell["link_path"],
                    cell["link_label"],
                    cell["font_size"],
                    cell["bold"],
                    cell["rich_html"],
                    user
                )
            target_r += 1

        con.commit()
    finally:
        con.close()


def email_body(sheet):
    d=cells_for(sheet); lines=[]
    for r in range(3,51):
        n=d.get((r,1),'').strip(); j=d.get((r,2),'').strip(); f=d.get((r,4),'').strip(); fj=d.get((r,5),'').strip()
        if n or j: lines.append(f"Numbering {n}: {j}".strip())
        if f or fj: lines.append(f"Fabrication {f}: {fj}".strip())
    return "Please see today's schedule below.\r\n\r\n" + "\r\n".join(lines) + "\r\n\r\nOpen the PMW app for the live version."


def make_schedule_pdf(sheet):
    """Create a PDF copy of the active schedule and return the full file path.
    V16 keeps PMW cell colors and marks linked ticket/email cells with an envelope.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    meta = cell_meta_for(sheet)
    d = cells_for(sheet)
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_sheet = ''.join(ch if ch.isalnum() or ch in ('-', '_') else '_' for ch in sheet)[:40]
    filename = f"PMW_Schedule_{safe_sheet}_{stamp}.pdf"
    path = os.path.join(EXPORT_FOLDER, filename)

    doc = SimpleDocTemplate(path, pagesize=landscape(letter), rightMargin=0.35*inch, leftMargin=0.35*inch, topMargin=0.30*inch, bottomMargin=0.30*inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('PMWTitle', parent=styles['Heading1'], alignment=1, fontName='Helvetica-Bold', fontSize=17, leading=20)
    cell_style = ParagraphStyle('PMWCell', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=9)
    head_style = ParagraphStyle('PMWHead', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, leading=11, alignment=1)

    title = d.get((1,2),'FABRICATION SCHEDULE') or 'FABRICATION SCHEDULE'
    datev = get_schedule_date_settings(sheet)['display_date']
    story = [Paragraph(f"{html.escape(title)} &nbsp;&nbsp;&nbsp; {html.escape(datev)}", title_style), Spacer(1, 6)]

    data = [[Paragraph('NUMBER', head_style), Paragraph('NUMBERING', head_style), Paragraph('STATUS/NOTES', head_style), Paragraph('NUMBER', head_style), Paragraph('FABRICATION', head_style), Paragraph('STATUS/NOTES', head_style)]]
    row_source=[]
    for r in range(3, 51):
        vals = []
        has_content = False
        for c in DISPLAY_COLS:
            m = meta.get((r,c), {})
            txt = str(m.get("value",""))
            if m.get("link_path"):
                txt = "✉ " + txt
            if txt.strip() or m.get("bg_color") or m.get("text_color"):
                has_content = True
            vals.append(Paragraph(html.escape(txt), cell_style))
        if has_content:
            data.append(vals)
            row_source.append(r)
    if len(data) == 1:
        data.append(['','','','','',''])

    tbl = Table(data, colWidths=[0.55*inch, 3.30*inch, 0.85*inch, 0.55*inch, 3.30*inch, 0.85*inch], repeatRows=1)
    style_cmds=[
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#d9ead3')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('ALIGN', (2,0), (2,-1), 'CENTER'),
        ('ALIGN', (3,0), (3,-1), 'CENTER'),
        ('ALIGN', (5,0), (5,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
    ]

    for pdf_row, sheet_row in enumerate(row_source, start=1):
        for ci, c in enumerate(DISPLAY_COLS):
            m = meta.get((sheet_row,c), {})
            bg = (m.get("bg_color") or "").strip()
            txt = (m.get("text_color") or "").strip()
            fsize = (m.get("font_size") or "").strip()
            bold = (m.get("bold") or "").strip()
            if bg:
                try:
                    style_cmds.append(('BACKGROUND', (ci,pdf_row), (ci,pdf_row), colors.HexColor(bg)))
                except Exception:
                    pass
            if txt:
                try:
                    style_cmds.append(('TEXTCOLOR', (ci,pdf_row), (ci,pdf_row), colors.HexColor(txt)))
                except Exception:
                    pass
            if fsize:
                try:
                    style_cmds.append(('FONTSIZE', (ci,pdf_row), (ci,pdf_row), float(fsize)))
                except Exception:
                    pass
            if bold:
                style_cmds.append(('FONTNAME', (ci,pdf_row), (ci,pdf_row), 'Helvetica-Bold'))

    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)
    doc.build(story)
    return path


def schedule_numbers_to_rows(sheet, side, start_num, end_num):
    """Translate visible schedule numbers into real rows.
    Example: user enters 1 to 5, app finds rows where the Number column has 1-5.
    It ignores blank number rows.
    """
    meta = cell_meta_for(sheet)
    if side == "numbering":
        key_cols = [1]
    elif side == "fabrication":
        key_cols = [4]
    else:
        key_cols = [1,4]

    try:
        a = float(str(start_num).strip())
        b = float(str(end_num).strip())
    except Exception:
        return []

    low, high = min(a,b), max(a,b)
    rows = []
    for r in range(3,51):
        for c in key_cols:
            val = str(meta.get((r,c),{}).get("value","")).strip()
            if not val:
                continue
            try:
                n = float(val)
            except Exception:
                continue
            if low <= n <= high:
                rows.append(r)
                break
    return sorted(set(rows))

def make_snip_pdf(sheet, start_row, end_row, side):
    """v45.5: Portrait snip that auto-fills the whole page.

    Only affects Snip / Print / Email PDF generation.
    It scales row height and text based on how many rows are included.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    meta = cell_meta_for(sheet)
    d = cells_for(sheet)

    selected_rows = schedule_numbers_to_rows(sheet, side, start_row, end_row)
    if not selected_rows:
        try:
            sr = max(3, int(float(str(start_row).strip())))
            er = min(50, int(float(str(end_row).strip())))
            if er < sr:
                sr, er = er, sr
            selected_rows = list(range(sr, er + 1))
        except Exception:
            selected_rows = []

    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_sheet = ''.join(ch if ch.isalnum() or ch in ('-', '_') else '_' for ch in sheet)[:40]
    filename = f"PMW_Snip_{safe_sheet}_{side}_{start_row}-{end_row}_{stamp}.pdf"
    path = os.path.join(EXPORT_FOLDER, filename)

    page_w, page_h = letter  # portrait
    margin = 0.18 * inch
    content_w = page_w - (2 * margin)

    if side == "numbering":
        cols = [1, 2, 3]
        headers = ['#', 'NUMBERING', 'DONE']
        widths = [0.45 * inch, content_w - (0.45 * inch + 0.60 * inch), 0.60 * inch]
    elif side == "fabrication":
        cols = [4, 5, 6]
        headers = ['#', 'FABRICATION', 'DONE']
        widths = [0.45 * inch, content_w - (0.45 * inch + 0.60 * inch), 0.60 * inch]
    else:
        cols = DISPLAY_COLS
        headers = ['#', 'NUMBERING', 'DONE', '#', 'FABRICATION', 'DONE']
        widths = [0.32 * inch, 3.08 * inch, 0.42 * inch, 0.32 * inch, 3.08 * inch, 0.42 * inch]

    # First collect rows so we know how many actually print.
    row_source = []
    raw_values = []
    for r in selected_rows:
        vals = []
        has = False
        for c in cols:
            m = meta.get((r, c), {})
            txt = str(m.get("value", ""))
            if m.get("link_path"):
                txt = "✉ " + txt
            if txt.strip() or m.get("bg_color") or m.get("text_color"):
                has = True
            vals.append(txt)
        if has:
            raw_values.append(vals)
            row_source.append(r)

    if not raw_values:
        raw_values = [['' for _ in headers]]

    printed_rows = len(raw_values)

    # Calculate row height so the table fills nearly the full portrait page.
    top_reserved = 0.72 * inch  # title/date/spacer area
    available_table_h = page_h - (2 * margin) - top_reserved

    # Header gets a little less height than data rows.
    data_row_h = available_table_h / max(1, printed_rows + 0.85)
    data_row_h = max(0.38 * inch, min(data_row_h, 0.92 * inch))
    header_h = max(0.34 * inch, min(data_row_h * 0.85, 0.55 * inch))
    row_heights = [header_h] + [data_row_h for _ in range(printed_rows)]

    # Larger font for fewer rows, automatically scaled by row height.
    # This keeps 1-12 row snips big for the shop, while larger snips still fit.
    cell_font = max(8.5, min(18.5, data_row_h * 0.28))
    leading = cell_font + 2.2
    header_font = max(9, min(18, header_h * 0.36))
    title_font = 19
    date_font = 13

    doc = SimpleDocTemplate(
        path,
        pagesize=letter,
        rightMargin=margin,
        leftMargin=margin,
        topMargin=0.16 * inch,
        bottomMargin=0.16 * inch
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'SnipTitle',
        parent=styles['Heading1'],
        alignment=1,
        fontName='Helvetica-Bold',
        fontSize=title_font,
        leading=title_font + 1,
        spaceAfter=0
    )
    date_style = ParagraphStyle(
        'SnipDate',
        parent=styles['Normal'],
        alignment=1,
        fontName='Helvetica-Bold',
        fontSize=date_font,
        leading=date_font + 1,
        spaceAfter=2
    )
    cell_style = ParagraphStyle(
        'SnipCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=cell_font,
        leading=leading,
        wordWrap='CJK'
    )
    head_style = ParagraphStyle(
        'SnipHead',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=header_font,
        leading=header_font + 1,
        alignment=1
    )

    title = d.get((1, 2), 'FABRICATION SCHEDULE') or 'FABRICATION SCHEDULE'
    datev = get_schedule_date_settings(sheet)['display_date']

    story = [
        Paragraph(html.escape(title), title_style),
        Paragraph(html.escape(datev), date_style),
        Spacer(1, 1)
    ]

    data = [[Paragraph(h, head_style) for h in headers]]
    for vals in raw_values:
        data.append([Paragraph(html.escape(v), cell_style) for v in vals])

    tbl = Table(data, colWidths=widths, rowHeights=row_heights, repeatRows=1, hAlign='CENTER')

    style_cmds = [
        ('GRID', (0, 0), (-1, -1), 0.85, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d9ead3')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (-1, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]

    if len(cols) == 6:
        for center_col in [0, 2, 3, 5]:
            style_cmds.append(('ALIGN', (center_col, 0), (center_col, -1), 'CENTER'))

    for pdf_row, sheet_row in enumerate(row_source, start=1):
        for ci, c in enumerate(cols):
            m = meta.get((sheet_row, c), {})
            bg = (m.get("bg_color") or "").strip()
            txt = (m.get("text_color") or "").strip()
            bold = (m.get("bold") or "").strip()
            # Do not let small manual font sizes shrink the snip; this is for shop visibility.
            if bg:
                try:
                    style_cmds.append(('BACKGROUND', (ci, pdf_row), (ci, pdf_row), colors.HexColor(bg)))
                except Exception:
                    pass
            if txt:
                try:
                    style_cmds.append(('TEXTCOLOR', (ci, pdf_row), (ci, pdf_row), colors.HexColor(txt)))
                except Exception:
                    pass
            if bold:
                style_cmds.append(('FONTNAME', (ci, pdf_row), (ci, pdf_row), 'Helvetica-Bold'))

    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)
    doc.build(story)
    return path


def open_outlook_draft_with_attachment(pdf_path, sheet):
    """Try to open an Outlook draft with the PDF already attached. Returns (ok, message)."""
    try:
        import win32com.client
        outlook = win32com.client.Dispatch('Outlook.Application')
        mail = outlook.CreateItem(0)
        mail.Subject = f"PMW Schedule {datetime.now().strftime('%m-%d-%y')}"
        mail.Body = "Please see attached PMW schedule PDF.\n\n"
        mail.Attachments.Add(os.path.abspath(pdf_path))
        mail.Display(True)
        return True, 'Outlook draft opened with the PDF attached.'
    except Exception as e:
        return False, str(e)


def reveal_file(path):
    try:
        if platform.system().lower().startswith('win'):
            subprocess.Popen(['explorer', '/select,', os.path.normpath(path)])
        elif platform.system().lower() == 'darwin':
            subprocess.Popen(['open', '-R', path])
        else:
            subprocess.Popen(['xdg-open', os.path.dirname(path)])
    except Exception:
        pass


def cloud_notice_banner():
    if os.environ.get("RENDER"):
        return "<div style='background:#fff3cd;border:1px solid #d6b656;padding:7px;margin:6px;font-weight:bold'></div>"
    return ""


def db_mode_banner():
    if USE_POSTGRES:
        return "<div style='background:#d4edda;border:2px solid #28a745;padding:7px;margin:6px;font-weight:bold'></div>"
    if os.environ.get("RENDER"):
        return "<div style='background:#f8d7da;border:2px solid #dc3545;padding:7px;margin:6px;font-weight:bold'>Database Mode: SQLite TEMP — data can reset. Add DATABASE_URL to Render Environment.</div>"
    return ""

BASE = """
<!doctype html><html><head><meta name='viewport' content='width=device-width, initial-scale=1, maximum-scale=5, user-scalable=yes, viewport-fit=cover'><title>{{app_name}}</title>
<style>
*{box-sizing:border-box} body{margin:0;background:#e7e6e6;font-family:Calibri,Arial,sans-serif;font-size:11pt;color:#111}.top{height:38px;background:#107c41;color:white;display:flex;align-items:center;justify-content:space-between;padding:0 14px}.brand{font-weight:bold;font-size:17px}.nav a,.nav span{color:white;margin-left:14px;text-decoration:none;font-weight:bold}.toolbar{padding:7px 9px;background:#f3f3f3;border-bottom:1px solid #aaa;display:flex;gap:8px;align-items:center}.grow{flex:1}.small{font-size:12px;color:#555}.flash{background:#fff3cd;padding:7px;border-bottom:1px solid #d6b656}.tabs{display:flex;gap:3px;padding:5px 8px 0}.tab{background:#d9ead3;border:1px solid #888;border-bottom:0;padding:6px 16px;color:#111;text-decoration:none}.active{background:white;font-weight:bold}.workspace{height:calc(100vh - 88px);overflow:auto;padding:0 8px 16px}.sheetline{display:flex;gap:12px;align-items:flex-start;min-width:1320px}.sheetwrap{background:white;border:1px solid #777;box-shadow:0 1px 5px #aaa}.sheet{border-collapse:collapse;table-layout:fixed;width:1160px}.sheet col.num{width:52px}.sheet col.job{width:450px}.sheet col.note{width:78px}.sheet td{border:1px solid black;height:34px;padding:2px 5px;vertical-align:middle;overflow:hidden}.title{text-align:center;font-size:18pt;font-weight:bold}.section{text-align:center;font-size:16pt;font-weight:bold}.date{text-align:center;font-size:14pt;font-weight:bold}.center{text-align:center}.cellinput{width:100%;height:30px;border:0;background:transparent;font:inherit;outline:none;padding:3px 4px}.cellinput:focus{outline:3px solid #107c41}.numinput,.noteinput{text-align:center}.colorbar{display:flex;gap:5px;align-items:center}.swatch{width:26px;height:24px;border:1px solid #555;border-radius:3px;cursor:pointer}.swatch.white{background:white}.swatch.red{background:#ff6666}.swatch.yellow{background:#fff066}.swatch.green{background:#93d050}.swatch.blue{background:#9dc3e6}.swatch.clear{background:linear-gradient(135deg,#fff 45%,#c00 48%,#c00 52%,#fff 55%)}.linkbtn{display:inline-block;background:#ffd966;border:1px solid #a67c00;border-radius:3px;padding:1px 4px;margin-left:3px;text-decoration:none;color:#111;font-size:12px}.cellbox{display:flex;align-items:center}.cellbox .cellinput{flex:1}.selectedCell{outline:3px solid #1d4ed8!important;box-shadow:inset 0 0 0 2px white}.richEditor{min-height:120px;border:1px solid #777;padding:8px;background:white;color:#111;overflow:auto}.richEditor:focus{outline:2px solid #107c41}.richCell{width:100%;min-height:30px;border:0;background:transparent;outline:none;padding:5px 4px;white-space:pre-wrap;overflow:hidden}.richCell:focus{outline:3px solid #107c41;background:#fffde7}.plainHidden{display:none!important}.jobinput{text-align:left}.buttons{width:175px;padding-top:32px}.buttons button,.btn,button{background:#e9e9e9;border:1px solid #777;border-radius:2px;padding:6px 12px;font:11pt Calibri,Arial;cursor:pointer}.buttons button{width:165px;height:36px;margin-bottom:7px}.green{background:#107c41!important;color:white!important;border-color:#0b5f31!important;font-weight:bold}.login{max-width:430px;margin:80px auto;background:white;padding:22px;border:1px solid #888;box-shadow:0 2px 10px #999}.login input{width:100%;margin:5px 0 12px;padding:7px}.admin{background:white;border-collapse:collapse;width:100%;margin:10px}.admin th,.admin td{border:1px solid #aaa;padding:7px;text-align:left}@media print{
.top,.toolbar,.tabs,.buttons,#snipBox{display:none!important}
.workspace{height:auto!important;overflow:visible!important;padding:0!important}
.sheetline{min-width:0!important}
.sheetwrap{border:0!important;box-shadow:none!important}
*{-webkit-print-color-adjust:exact!important;print-color-adjust:exact!important;color-adjust:exact!important}
.sheet td,.cellinput{background-color:inherit!important}
.cellinput{color:inherit!important}
}

/* ===== PMW MOBILE LAYOUT ===== */
.mobileTop,.mobileFab{display:none!important}

@media (max-width: 800px){
  body{font-size:16px;background:#f1f1f1;overflow:hidden}
  .top{
    height:auto;min-height:44px;padding:6px 8px;display:block;position:sticky;top:0;z-index:1000;
  }
  .brand{font-size:15px;line-height:18px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
  .nav{
    display:flex;gap:10px;overflow-x:auto;padding-top:4px;white-space:nowrap;
  }
  .nav a,.nav span{font-size:13px;margin-left:0}
  .toolbar{
    display:none !important;
  }
  .tabs{
    position:sticky;top:54px;z-index:900;background:#f1f1f1;
    padding:4px 6px;overflow-x:auto;white-space:nowrap;
  }
  .tab{padding:7px 12px;font-size:15px}
  .workspace{
    height:calc(100vh - 105px);
    overflow:auto;
    padding:0 4px 75px 4px;
    -webkit-overflow-scrolling:touch;
  }
  .sheetline{min-width:0;display:block}
  .sheetwrap{
    width:max-content;
    min-width:100%;
    box-shadow:none;
    border:0;
    background:white;
  }
  .sheet{
    width:auto;
    min-width:980px;
    font-size:17px;
  }
  .sheet col.num{width:70px}
  .sheet col.job{width:420px}
  .sheet col.note{width:95px}
  .sheet td{
    height:48px;
    padding:4px 8px;
  }
  .title{font-size:21px;white-space:nowrap}
  .section{font-size:20px}
  .date{font-size:16px}
  .cellinput{
    height:42px;
    font-size:17px;
    padding:6px 5px;
  }
  .richCell{min-height:42px;font-size:17px;padding:8px 5px}
  .buttons{
    position:fixed;
    left:0;right:0;bottom:0;
    width:100%;
    background:#107c41;
    padding:7px;
    display:flex;
    overflow-x:auto;
    gap:6px;
    z-index:1200;
    box-shadow:0 -2px 8px rgba(0,0,0,.25);
  }
  .buttons button{
    min-width:115px;
    width:auto;
    height:44px;
    margin:0;
    font-size:14px;
    background:white;
  }
  .buttons button[name='cmd'][value='clear_schedule'],
  .buttons button[name='cmd'][value='delete_comments']{
    display:none;
  }
  .mobileFab{display:none!important}
  .mobileFab button{display:none!important}
  .mobileTop{
    display:flex;
    position:sticky;
    top:92px;
    z-index:850;
    background:#fff8dc;
    border-bottom:1px solid #d6b656;
    padding:5px 6px;
    gap:5px;
    overflow-x:auto;
    white-space:nowrap;
  }
  .mobileTop button{
    height:36px;
    min-width:58px;
    font-size:13px;
    padding:5px 8px;
    background:white;
  }
  .mobileTop .red{background:#ff6666}
  .mobileTop .yellow{background:#fff066}
  .mobileTop .green{background:#93d050;color:#111}
  .mobileTop .blue{background:#9dc3e6;color:#111}
  .mobileTop .white{background:white;color:#111}
  #snipBox,#richBox{
    left:10px!important;
    right:10px!important;
    top:95px!important;
    transform:none!important;
    width:auto!important;
    max-height:70vh;
    overflow:auto;
  }
  .login{
    margin:25px 10px;
    max-width:none;
  }
  .admin{
    font-size:14px;
    margin:0;
    width:max-content;
    min-width:100%;
  }
  .flash{
    font-size:14px;
    padding:6px 8px;
  }
}


/* ===== V24 MOBILE ZOOM + BUTTON FIX ===== */
@media (max-width: 800px){
  .sheetwrap{
    transform-origin: top left;
    transition: transform .12s ease;
  }
  .mobileZoomLabel{
    font-weight:bold;
    padding:8px 10px;
    background:white;
    border:1px solid #777;
    border-radius:4px;
    color:#111;
  }
  .mobileTop button{
    color:#111 !important;
    font-weight:bold;
    min-width:70px;
  }
  .buttons{
    gap:8px !important;
    padding:8px !important;
    min-height:64px;
  }
  .buttons button{
    color:#111 !important;
    background:#fff !important;
    border:2px solid #d9d9d9 !important;
    min-width:145px !important;
    height:48px !important;
    font-size:15px !important;
    font-weight:bold !important;
    line-height:18px !important;
    padding:6px 10px !important;
    opacity:1 !important;
    visibility:visible !important;
    text-indent:0 !important;
  }
  .buttons button[name='cmd'][value='clear_schedule'],
  .buttons button[name='cmd'][value='delete_comments'],
  .buttons button[name='cmd'][value='done']{
    display:none !important;
  }
  .mobileFab{display:none!important}
  .mobileFab button{display:none!important}
}


.userform{background:white;border:1px solid #bbb;padding:14px;margin:10px 0;box-shadow:0 1px 3px #ccc}
.userform input,.userform select{padding:8px;margin:4px;min-width:160px}
.rolebadge{display:inline-block;padding:3px 8px;border-radius:12px;font-weight:bold}
.role-admin{background:#d9ead3}
.role-editor{background:#fff2cc}
.role-viewer{background:#d9eaf7}


/* ===== V35 MOBILE TICKET / ATTACHMENT PREVIEW ===== */
.ticketPreviewWrap{max-width:1180px;margin:0 auto;padding:10px}
.ticketHeaderCard,.ticketAttachmentCard,.ticketBodyCard{
  background:white;
  border:1px solid #bbb;
  padding:14px;
  margin:10px 0;
  box-sizing:border-box;
}
.ticketAttachmentArea{
  background:#f7f7f7;
  border:1px solid #bbb;
  padding:12px;
  margin:10px 0;
}
.ticketAttachmentFrame{
  width:100%;
  height:650px;
  border:1px solid #ccc;
  background:white;
}
.ticketAttachmentImage{
  max-width:100%;
  height:auto;
  border:1px solid #ccc;
  margin-top:8px;
}
.mobileOpenHint{display:none}

@media (max-width: 800px){
  body{
    overflow-x:hidden;
  }
  .ticketPreviewWrap{
    padding:6px;
    width:100%;
    max-width:100%;
  }
  .ticketHeaderCard,.ticketAttachmentCard,.ticketBodyCard,.ticketAttachmentArea{
    padding:10px;
    margin:8px 0;
    width:100%;
    max-width:100%;
  }
  .ticketHeaderCard p{
    font-size:14px;
    line-height:1.35;
    word-break:break-word;
  }
  .ticketAttachmentCard b{
    display:block;
    font-size:16px;
    word-break:break-word;
    margin-bottom:8px;
  }
  .ticketAttachmentCard .btn,
  .ticketHeaderCard .btn{
    display:block;
    width:100%;
    box-sizing:border-box;
    text-align:center;
    margin:6px 0;
    padding:12px 8px;
    font-size:16px;
  }
  .ticketAttachmentFrame{
    height:72vh;
    min-height:420px;
    width:100%;
  }
  .ticketAttachmentImage{
    width:100%;
    max-width:100%;
  }
  .ticketBodyCard{
    font-size:15px !important;
    line-height:1.45 !important;
    word-break:break-word;
    overflow-wrap:anywhere;
  }
  .mobileOpenHint{
    display:block;
    background:#fff3cd;
    border:1px solid #d6b656;
    padding:8px;
    margin:8px 0;
    font-size:14px;
  }
  h2,h3{
    margin:10px 0;
  }
}


/* ===== V36 IPHONE PORTRAIT SCROLL FIX ===== */
.mobilePdfPortraitNotice{display:none}

@media (max-width: 800px){
  html, body{
    height:auto !important;
    min-height:100% !important;
    overflow-y:auto !important;
    overflow-x:hidden !important;
    -webkit-overflow-scrolling:touch !important;
    position:static !important;
  }

  .ticketPreviewWrap,
  .ticketHeaderCard,
  .ticketAttachmentArea,
  .ticketAttachmentCard,
  .ticketBodyCard{
    overflow:visible !important;
    max-height:none !important;
  }

  .ticketAttachmentFrame{
    display:none !important;
  }

  .mobilePdfPortraitNotice{
    display:block !important;
    background:#fff3cd;
    border:1px solid #d6b656;
    padding:10px;
    margin:8px 0;
    font-size:14px;
    line-height:1.35;
  }
}


/* ===== v45.1 Schedule Date Control ===== */
.scheduleDateBox{
  background:#ffffff;
  border:1px solid #999;
  padding:8px 10px;
  margin:6px 0 10px 0;
  display:flex;
  align-items:center;
  gap:8px;
  flex-wrap:wrap;
}
.scheduleDateTitle{
  font-size:18px;
  font-weight:bold;
  color:#073763;
  margin-right:8px;
}
.scheduleDateBox input[type=date]{padding:6px}
.scheduleDateBox button{padding:7px 10px}
@media(max-width:800px){
  .scheduleDateBox{display:block}
  .scheduleDateTitle{display:block;margin-bottom:8px}
  .scheduleDateBox input,.scheduleDateBox button{width:100%;box-sizing:border-box;margin:4px 0}
}

.btn.red, button.red{background:#d9534f!important;color:white!important;border:1px solid #842029!important}
.buttons button.green,.mobileFab button.green,button.green{background:#93d050!important;border:1px solid #38761d!important;font-weight:bold}
.mobileFab{display:none!important}.mobileFab button{display:none!important}
.btn.green,button.green{background:#93d050!important;border:1px solid #38761d!important;font-weight:bold}.btn.red,button.red{background:#d9534f!important;color:white!important;border:1px solid #842029!important}

/* ===== v49.5 iPhone mobile bottom navigation repair ===== */
@media (max-width: 800px){
  html, body{
    max-width:100vw;
    overflow-x:auto!important;
    -webkit-text-size-adjust:100%;
    padding-top: env(safe-area-inset-top);
    padding-bottom: calc(70px + env(safe-area-inset-bottom));
  }

  .tabs, .toolbar, .colorbar{
    max-width:100vw;
    overflow-x:auto!important;
    white-space:nowrap!important;
    -webkit-overflow-scrolling:touch;
    padding-top:8px!important;
    padding-bottom:8px!important;
  }

  .tabs a, .toolbar a, .toolbar button, .colorbar button{
    display:inline-block!important;
    min-height:42px;
    margin:4px;
    font-size:16px!important;
  }

  .workspace{
    width:100vw!important;
    max-width:100vw!important;
    overflow-x:auto!important;
    -webkit-overflow-scrolling:touch;
  }

  .sheetline{
    width:100vw!important;
    max-width:100vw!important;
    overflow-x:auto!important;
    overflow-y:visible!important;
    -webkit-overflow-scrolling:touch;
  }

  .sheetwrap{
    width:max-content!important;
    min-width:1000px!important;
    transform-origin:top left!important;
  }

  table.sheet{
    min-width:1000px!important;
  }

  .buttons{
    width:100vw!important;
    max-width:100vw!important;
    overflow-x:auto!important;
    white-space:nowrap!important;
    -webkit-overflow-scrolling:touch;
    padding:8px 4px 12px!important;
  }

  .buttons button{
    display:inline-block!important;
    width:auto!important;
    min-width:125px;
    min-height:42px;
    margin:4px;
    font-size:16px!important;
  }

  .mobileFab{display:none!important;}

  .mobileActionBar{
    display:block!important;
    position:sticky;
    bottom:54px;
    z-index:999;
    width:100vw!important;
    max-width:100vw!important;
    overflow-x:auto!important;
    overflow-y:hidden!important;
    white-space:nowrap!important;
    -webkit-overflow-scrolling:touch;
    background:#0f7b2f;
    padding:8px 4px calc(8px + env(safe-area-inset-bottom)) 4px;
    box-sizing:border-box;
  }

  .mobileActionBar button{
    display:inline-block!important;
    width:auto!important;
    min-width:125px;
    min-height:44px;
    margin:3px;
    border-radius:6px;
    border:1px solid #38761d;
    background:#fff;
    color:#111;
    font-size:16px!important;
    font-weight:bold;
  }

  .mobileBottomNav{
    display:block!important;
    position:fixed;
    left:0;
    right:0;
    bottom:0;
    z-index:1000;
    background:#111;
    border-top:2px solid #0f7b2f;
    overflow-x:auto;
    white-space:nowrap;
    -webkit-overflow-scrolling:touch;
    padding:7px 4px calc(7px + env(safe-area-inset-bottom)) 4px;
    box-sizing:border-box;
  }

  .mobileBottomNav a{
    display:inline-block;
    color:white;
    background:#0f7b2f;
    padding:10px 12px;
    margin:0 3px;
    border-radius:6px;
    text-decoration:none;
    font-size:15px;
    font-weight:bold;
  }

  table.admin{
    display:block;
    width:max-content;
    min-width:1000px;
    max-width:none;
    overflow-x:auto;
    -webkit-overflow-scrolling:touch;
  }

  .admin th,.admin td{
    white-space:nowrap;
    padding:8px;
    font-size:14px;
  }

  .userform{
    max-width:100vw;
    overflow-x:auto;
    -webkit-overflow-scrolling:touch;
  }

  input, select, textarea, button{
    font-size:16px!important;
  }

  #snipBox{
    left:8px!important;
    right:8px!important;
    top:75px!important;
    width:auto!important;
    max-width:calc(100vw - 16px)!important;
  }
}

@media (min-width: 801px){
  .mobileBottomNav,.mobileActionBar{display:none!important;}
}


/* ===== v51 Plain Cell Edit Arrows Fix ===== */
@media (max-width: 800px){
  html, body{
    max-width:100vw;
    overflow-x:auto!important;
    -webkit-text-size-adjust:100%;
    padding-bottom:20px!important;
  }

  .tabs,.toolbar,.colorbar{
    max-width:100vw;
    overflow-x:auto!important;
    white-space:nowrap!important;
    -webkit-overflow-scrolling:touch;
    padding-top:8px!important;
    padding-bottom:8px!important;
  }

  .tabs a,.toolbar a,.toolbar button,.colorbar button{
    display:inline-block!important;
    min-height:42px!important;
    margin:4px!important;
    font-size:16px!important;
  }

  .workspace{
    width:100vw!important;
    max-width:100vw!important;
    overflow-x:auto!important;
    -webkit-overflow-scrolling:touch;
  }

  .sheetline{
    width:100vw!important;
    max-width:100vw!important;
    overflow-x:auto!important;
    overflow-y:visible!important;
    -webkit-overflow-scrolling:touch;
    padding-bottom:12px!important;
  }

  .sheetwrap{
    transform-origin:top left!important;
    width:max-content!important;
    min-width:1000px!important;
  }

  table.sheet{min-width:1000px!important;}

  .scheduleZoomBar{
    display:block!important;
    width:100vw!important;
    max-width:100vw!important;
    overflow-x:auto!important;
    white-space:nowrap!important;
    -webkit-overflow-scrolling:touch;
    background:#d9ead3;
    border-top:1px solid #6aa84f;
    border-bottom:1px solid #6aa84f;
    padding:8px 4px!important;
    box-sizing:border-box;
  }

  .scheduleZoomBar button{
    display:inline-block!important;
    min-width:125px!important;
    min-height:44px!important;
    margin:3px!important;
    background:#fff!important;
    color:#111!important;
    border:1px solid #38761d!important;
    border-radius:6px!important;
    font-size:16px!important;
    font-weight:bold!important;
  }

  .buttons{
    position:static!important;
    width:100vw!important;
    max-width:100vw!important;
    display:block!important;
    overflow-x:auto!important;
    overflow-y:hidden!important;
    white-space:nowrap!important;
    -webkit-overflow-scrolling:touch;
    padding:8px 4px 12px 4px!important;
    margin:0!important;
    background:#f3f3f3!important;
    box-sizing:border-box!important;
  }

  .buttons button{
    display:inline-block!important;
    position:static!important;
    width:auto!important;
    min-width:125px!important;
    min-height:44px!important;
    margin:3px!important;
    font-size:16px!important;
  }

  .mobileActionBar{
    display:block!important;
    position:static!important;
    width:100vw!important;
    max-width:100vw!important;
    overflow-x:auto!important;
    overflow-y:hidden!important;
    white-space:nowrap!important;
    -webkit-overflow-scrolling:touch;
    background:#0f7b2f!important;
    padding:8px 4px 12px 4px!important;
    margin:0!important;
    box-sizing:border-box!important;
    z-index:auto!important;
  }

  .mobileActionBar button{
    display:inline-block!important;
    position:static!important;
    width:auto!important;
    min-width:125px!important;
    min-height:44px!important;
    margin:3px!important;
    border-radius:6px!important;
    background:white!important;
    color:#111!important;
    border:1px solid #38761d!important;
    font-size:16px!important;
    font-weight:bold!important;
  }

  .mobileFab{display:none!important;}

  .mobileBottomNav{
    display:block!important;
    position:static!important;
    width:100vw!important;
    max-width:100vw!important;
    overflow-x:auto!important;
    overflow-y:hidden!important;
    white-space:nowrap!important;
    -webkit-overflow-scrolling:touch;
    background:#111!important;
    border-top:2px solid #0f7b2f!important;
    padding:8px 4px 12px 4px!important;
    margin-top:8px!important;
    box-sizing:border-box!important;
    z-index:auto!important;
  }

  .mobileBottomNav a{
    display:inline-block!important;
    color:white!important;
    background:#0f7b2f!important;
    padding:10px 12px!important;
    margin:0 3px!important;
    border-radius:6px!important;
    text-decoration:none!important;
    font-size:15px!important;
    font-weight:bold!important;
  }

  table.admin{
    display:block!important;
    width:max-content!important;
    min-width:1000px!important;
    max-width:none!important;
    overflow-x:auto!important;
    -webkit-overflow-scrolling:touch;
  }

  .admin th,.admin td{
    white-space:nowrap!important;
    padding:8px!important;
    font-size:14px!important;
  }

  .userform{
    max-width:100vw!important;
    overflow-x:auto!important;
    -webkit-overflow-scrolling:touch;
  }

  input,select,textarea,button{font-size:16px!important;}

  #snipBox{
    left:8px!important;
    right:8px!important;
    top:75px!important;
    width:auto!important;
    max-width:calc(100vw - 16px)!important;
  }
}
@media (min-width:801px){
  .scheduleZoomBar,.mobileActionBar,.mobileBottomNav{display:none!important;}
}


/* ===== v51 Plain Cell Edit Arrows ===== */
@media (max-width: 800px){
  html, body{
    max-width:100vw;
    overflow-x:auto!important;
    -webkit-text-size-adjust:100%;
    padding-bottom: calc(158px + env(safe-area-inset-bottom)) !important;
  }

  .workspace, .sheetline{
    max-width:100vw!important;
    overflow-x:auto!important;
    -webkit-overflow-scrolling:touch;
  }

  .sheetwrap{
    transform-origin:top left!important;
    width:max-content!important;
    min-width:1000px!important;
  }

  table.sheet{
    min-width:1000px!important;
  }

  .scheduleZoomBar{
    display:block!important;
    width:100vw!important;
    max-width:100vw!important;
    overflow-x:auto!important;
    white-space:nowrap!important;
    -webkit-overflow-scrolling:touch;
    background:#d9ead3!important;
    border-top:1px solid #6aa84f!important;
    border-bottom:1px solid #6aa84f!important;
    padding:8px 4px!important;
    box-sizing:border-box!important;
  }

  .scheduleZoomBar button{
    display:inline-block!important;
    min-width:125px!important;
    min-height:44px!important;
    margin:3px!important;
    background:#fff!important;
    color:#111!important;
    border:1px solid #38761d!important;
    border-radius:6px!important;
    font-size:16px!important;
    font-weight:bold!important;
  }

  .mobileActionBar{
    display:block!important;
    position:fixed!important;
    left:0!important;
    right:0!important;
    bottom: calc(58px + env(safe-area-inset-bottom)) !important;
    z-index:2000!important;
    width:100vw!important;
    max-width:100vw!important;
    overflow-x:auto!important;
    overflow-y:hidden!important;
    white-space:nowrap!important;
    -webkit-overflow-scrolling:touch;
    background:#0f7b2f!important;
    border-top:2px solid #0b5d24!important;
    padding:8px 4px!important;
    box-sizing:border-box!important;
    box-shadow:0 -3px 10px rgba(0,0,0,.25)!important;
  }

  .mobileActionBar button{
    display:inline-block!important;
    position:static!important;
    width:auto!important;
    min-width:132px!important;
    min-height:46px!important;
    margin:3px!important;
    padding:8px 10px!important;
    border-radius:6px!important;
    background:white!important;
    color:#111!important;
    border:1px solid #38761d!important;
    font-size:16px!important;
    font-weight:bold!important;
  }

  .mobileBottomNav{
    display:block!important;
    position:fixed!important;
    left:0!important;
    right:0!important;
    bottom:0!important;
    z-index:2001!important;
    width:100vw!important;
    max-width:100vw!important;
    overflow-x:auto!important;
    overflow-y:hidden!important;
    white-space:nowrap!important;
    -webkit-overflow-scrolling:touch;
    background:#111!important;
    border-top:2px solid #0f7b2f!important;
    padding:7px 4px calc(7px + env(safe-area-inset-bottom)) 4px!important;
    box-sizing:border-box!important;
    box-shadow:0 -3px 10px rgba(0,0,0,.25)!important;
  }

  .mobileBottomNav a{
    display:inline-block!important;
    color:white!important;
    background:#0f7b2f!important;
    padding:10px 12px!important;
    margin:0 3px!important;
    border-radius:6px!important;
    text-decoration:none!important;
    font-size:15px!important;
    font-weight:bold!important;
  }

  .buttons{
    display:block!important;
    width:100vw!important;
    max-width:100vw!important;
    overflow-x:auto!important;
    white-space:nowrap!important;
    -webkit-overflow-scrolling:touch;
    padding:8px 4px 110px 4px!important;
    margin:0!important;
    background:#f3f3f3!important;
    box-sizing:border-box!important;
  }

  .buttons button{
    display:inline-block!important;
    width:auto!important;
    min-width:125px!important;
    min-height:44px!important;
    margin:3px!important;
    font-size:16px!important;
  }

  .mobileFab{
    display:none!important;
  }
}
@media (min-width:801px){
  .mobileActionBar,.mobileBottomNav,.scheduleZoomBar{display:none!important;}
}


/* ===== v51 Plain Cell Edit Arrows ===== */
@media (max-width: 800px){
  html, body{
    padding-bottom: calc(220px + env(safe-area-inset-bottom)) !important;
  }

  .mobileTop{
    display:none!important;
  }

  .mobileFormatBar{
    display:block!important;
    position:fixed!important;
    left:0!important;
    right:0!important;
    bottom: calc(112px + env(safe-area-inset-bottom)) !important;
    z-index:2002!important;
    width:100vw!important;
    max-width:100vw!important;
    overflow-x:auto!important;
    overflow-y:hidden!important;
    white-space:nowrap!important;
    -webkit-overflow-scrolling:touch;
    background:#e8f3e8!important;
    border-top:2px solid #38761d!important;
    padding:7px 4px!important;
    box-sizing:border-box!important;
    box-shadow:0 -3px 8px rgba(0,0,0,.18)!important;
  }

  .mobileFormatBar button,.mobileFormatBar select{
    display:inline-block!important;
    position:static!important;
    min-width:88px!important;
    min-height:42px!important;
    margin:3px!important;
    padding:7px 9px!important;
    border-radius:6px!important;
    border:1px solid #777!important;
    color:#111!important;
    font-size:15px!important;
    font-weight:bold!important;
    vertical-align:middle!important;
  }

  .mobileFormatBar .mfLabel{
    display:inline-block!important;
    padding:9px 8px!important;
    margin:3px!important;
    background:#222!important;
    color:white!important;
    border-radius:6px!important;
    font-size:14px!important;
    font-weight:bold!important;
    vertical-align:middle!important;
  }

  .mobileFormatBar .red{background:#ff6666!important;}
  .mobileFormatBar .yellow{background:#fff066!important;}
  .mobileFormatBar .green{background:#93d050!important;}
  .mobileFormatBar .blue{background:#9dc3e6!important;}
  .mobileFormatBar .white{background:#ffffff!important;}
  .mobileFormatBar .clear{background:#ffffff!important;text-decoration:line-through;}
  .mobileFormatBar .tred{background:#fff!important;color:#c00000!important;}
  .mobileFormatBar .tyellow{background:#fff!important;color:#bf9000!important;}
  .mobileFormatBar .tgreen{background:#fff!important;color:#00b050!important;}
  .mobileFormatBar .tblue{background:#fff!important;color:#0070c0!important;}

  .mobileActionBar{
    bottom: calc(56px + env(safe-area-inset-bottom)) !important;
  }

  .mobileBottomNav{
    bottom:0!important;
  }

  .buttons{
    padding-bottom:160px!important;
  }
}
@media (min-width:801px){
  .mobileFormatBar{display:none!important;}
}


/* ===== v49.9 Mobile number keyboard tweaks ===== */
@media (max-width:800px){
  .mobileFormatBar .bigclear{
    min-width:150px!important;
    background:#ffffff!important;
    border:2px solid #b42318!important;
    color:#b42318!important;
    text-decoration:none!important;
  }
  .numinput{
    -webkit-appearance:none;
  }
}


/* ===== v51 Plain Cell Edit Arrows ===== */
.ticketAlert{
  margin:8px 0;
  padding:10px 12px;
  border-radius:6px;
  font-size:17px;
  font-weight:500;
}
.ticketAlert a{
  margin-left:10px;
  color:#073763;
  background:white;
  border:1px solid #999;
  padding:5px 8px;
  border-radius:4px;
  text-decoration:none;
  font-weight:bold;
}
.ticketAlertRed{
  background:#f8d7da;
  border:2px solid #b42318;
  color:#7f1d1d;
}
.ticketAlertGreen{
  background:#d4edda;
  border:2px solid #0f7b2f;
  color:#0f5132;
}
@media(max-width:800px){
  .ticketAlert{
    font-size:16px;
    margin:8px 4px;
  }
  .ticketAlert a{
    display:inline-block;
    margin-top:6px;
    margin-left:0;
  }
}


/* ===== v51 Plain Cell Edit Arrows ===== */
body > div[style*="PostgreSQL"],
body > div[style*="Render v26"],
.databaseBanner,
.dbBanner,
.renderBanner,
.upgradeBanner{
  display:none!important;
}

/* Hide old status banners by color/border pattern if still present */
body > div[style*="background:#d4edda"],
body > div[style*="background:#fff3cd"],
body > div[style*="background: #d4edda"],
body > div[style*="background: #fff3cd"]{
  display:none!important;
}

.header,.topbar{
  margin-bottom:4px!important;
}
.toolbar{
  margin-top:4px!important;
}
@media(max-width:800px){
  .header,.topbar{
    padding-top:6px!important;
    padding-bottom:6px!important;
  }
  h1{
    margin-top:4px!important;
    margin-bottom:4px!important;
  }
}


/* ===== v50.7 Undo visible styling ===== */
.pmwUndoTopButton,.pmwUndoActionButton{
  background:#fff3cd!important;
  border:2px solid #bf9000!important;
  color:#111!important;
  font-weight:bold!important;
}
@media(max-width:800px){
  .pmwUndoMobileButton,.mobileActionBar button.pmwUndoMobileButton{
    background:#fff3cd!important;
    border:2px solid #bf9000!important;
    color:#111!important;
    font-weight:bold!important;
  }
}

</style>
<link rel="apple-touch-icon" sizes="180x180" href="/static/apple-touch-icon.png">
<link rel="icon" type="image/png" sizes="32x32" href="/static/favicon-32.png">
<link rel="manifest" href="/static/site.webmanifest">
<meta name="theme-color" content="#0f7b2f">
<meta name="apple-mobile-web-app-title" content="PMW">

</head><body>
{% if session.get('user_id') %}<div class='top'><div class='brand'>{{app_name}} <span style='font-size:12px'>{{version}}</span></div><div class='nav'><span>{{session.username}} / {{session.role}}</span><a href='/'>Workbook</a><a href='/tickets'>Tickets</a>{% if can_admin %}<a href='/users'>Users</a><a href='/admin/storage'>Storage</a><a href='/admin/ticket_cleanup'>Cleanup</a><a href='/admin/job_history'>Job History</a><a href='/admin/backup'>Backup</a><a href='/audit'>Audit</a>{% endif %}<a href='/logout'>Logout</a></div></div>{% endif %}
{% for m in get_flashed_messages() %}<div class='flash'>{{m}}</div>{% endfor %}
{{body|safe}}
<div class="mobileBottomNav">
  <a href="/">Schedule</a>
  <a href="/tickets">Tickets</a>
  <a href="/admin/storage">Storage</a>
  <a href="/admin/ticket_cleanup">Cleanup</a>
  <a href="/admin/job_history">Job History</a>
  <a href="/admin/backup">Backup</a>
</div>

</body></html>
"""

def page(body): return render_template_string(BASE, body=db_mode_banner()+cloud_notice_banner()+body, app_name=APP_NAME, version=APP_VERSION, can_admin=can('admin'))

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method=='POST':
        u=request.form.get('username','').strip(); p=request.form.get('password','')
        con=db(); row=con.execute("SELECT * FROM users WHERE username=? AND active=1",(u,)).fetchone(); con.close()
        if row and check_password_hash(row['password_hash'],p):
            session.clear(); session['user_id']=row['id']; session['username']=row['username']; session['role']=row['role']; return redirect('/')
        flash('Bad username or password.')
    return page("""<div class='login'><h1>PMW Login</h1><form method='post'><label>Username</label><input name='username' autofocus><label>Password</label><input type='password' name='password'><button class='green'>Login</button></form><p>admin/admin123 &nbsp; shop/shop123 &nbsp; viewer/view123</p></div>""")

@app.route('/logout')
def logout(): session.clear(); return redirect('/login')


# ===== UNASSIGNED TICKET ALERT v50.2 =====
def unassigned_ticket_count():
    try:
        con = db()
        row = con.execute("""SELECT COUNT(*) AS n FROM ticket_links
                             WHERE COALESCE(scheduled_status,'') <> 'scheduled'""").fetchone()
        con.close()
        return int((row["n"] if row else 0) or 0)
    except Exception:
        return 0

def unassigned_ticket_alert_html():
    n = unassigned_ticket_count()
    if n > 0:
        word = "ticket" if n == 1 else "tickets"
        return f"""<div class='ticketAlert ticketAlertRed'>
            🔴 <b>{n} unassigned {word}</b>
            <span>need to be added to Fabrication or Numbering.</span>
            <a href='/tickets?status=unscheduled'>Open Tickets</a>
        </div>"""
    return """<div class='ticketAlert ticketAlertGreen'>
        🟢 <b>All tickets assigned</b>
    </div>"""


@app.route('/')
@login_required
def index():
    sheets=sheet_names(); active=request.args.get('sheet') or (sheets[0] if sheets else 'Fabrication Schedule')
    d=cells_for(active); meta=cell_meta_for(active); editable=can('editor')
    tabs=''.join(f"<a class='tab {'active' if s==active else ''}' href='/?sheet={urllib.parse.quote(s)}'>{html.escape(s)}</a>" for s in sheets)
    import_html = "" if not can('admin') else """<form method='post' action='/import' enctype='multipart/form-data' style='display:inline-flex;gap:5px;align-items:center'><input type='file' name='workbook' accept='.xlsm,.xlsx'><button>Import Excel</button></form>
<form method='post' action='/preview_ticket_drop' enctype='multipart/form-data' style='display:inline-flex;gap:5px;align-items:center'>
<input type='file' name='ticketlog' accept='.xlsm,.xlsx'>
<button>Open Ticket Drop / Pick Emails</button>
</form>"""
    note = "V18: multi-cell colors, font size, bold, and selected-word editor." if editable else "View only."
    color_html = ""
    if editable:
        color_html = """<div class='colorbar'><button type='button' class='pmwUndoTopButton' onclick='undoLastCellEdit()'>Undo</button>
<b>Cell:</b>
<button type='button' class='swatch red' onclick="setCellColor('#ff6666')" title='Red'></button>
<button type='button' class='swatch yellow' onclick="setCellColor('#fff066')" title='Yellow'></button>
<button type='button' class='swatch green' onclick="setCellColor('#93d050')" title='Green'></button>
<button type='button' class='swatch blue' onclick="setCellColor('#9dc3e6')" title='Blue'></button>
<button type='button' class='swatch white' onclick="setCellColor('#ffffff')" title='White'></button>
<button type='button' class='swatch clear' onclick="setCellColor('')" title='Clear'></button>
<b>Text:</b>
<button type='button' class='swatch red' onclick="setTextColor('#c00000')" title='Red text'></button>
<button type='button' class='swatch yellow' onclick="setTextColor('#bf9000')" title='Gold text'></button>
<button type='button' class='swatch green' onclick="setTextColor('#00b050')" title='Green text'></button>
<button type='button' class='swatch blue' onclick="setTextColor('#0070c0')" title='Blue text'></button>
<button type='button' class='swatch clear' onclick="setTextColor('')" title='Clear text'></button>
<b>Size:</b>
<select onchange="setFontSize(this.value); this.selectedIndex=0"><option value=''>Size</option><option value='10'>10</option><option value='11'>11</option><option value='12'>12</option><option value='14'>14</option><option value='16'>16</option><option value='18'>18</option><option value='20'>20</option></select>
<button type='button' onclick="toggleBold()"><b>B</b></button>
<button type='button' onclick="openRichTextEditor()">Edit selected words</button><button type='button' onclick="clearSelectedCells()">Clear Cell + Link</button>
<span class='small'>Drag across cells or Ctrl+click to format many cells.</span>
</div>"""
    body=f"<div class='toolbar'><b>Excel-style workbook</b><span class='small'>{note}</span>{color_html}<span class='grow'></span>{import_html}</div><div class='tabs'>{tabs}</div><div id='saveStatus' style='position:fixed;right:8px;top:48px;z-index:2000;background:#fff3cd;border:1px solid #d6b656;padding:3px 8px;font-size:12px;display:none'>Saved</div>"
    body += unassigned_ticket_alert_html()
    if editable:
        body += "<div class='mobileTop'><button type='button' class='red' onclick=\"setCellColor('#ff6666')\">Red</button><button type='button' class='yellow' onclick=\"setCellColor('#fff066')\">Yellow</button><button type='button' class='green' onclick=\"setCellColor('#93d050')\">Green</button><button type='button' class='blue' onclick=\"setCellColor('#9dc3e6')\">Blue</button><button type='button' class='white' onclick=\"setCellColor('#ffffff')\">White</button><button type='button' onclick=\"setCellColor('')\">Clear</button><button type='button' onclick=\"toggleBold()\"><b>B</b></button><button type='button' onclick=\"openRichTextEditor()\">Words</button><button type='button' onclick=\"clearSelectedCells()\">Clear</button><button type='button' onclick=\"mobileZoomOut()\">Zoom -</button><button type='button' onclick=\"mobileZoomIn()\">Zoom +</button><span class='mobileZoomLabel' id='mobileZoomLabel'>100%</span></div>"
    if editable:
        body += """<div class='mobileFormatBar'>
<button type='button' class='mf clear bigclear' onclick="clearSelectedCells()">Clear Cell + Link</button>
<span class='mfLabel'>Cell</span>
<button type='button' class='mf red' onclick="setCellColor('#ff6666')">Red</button>
<button type='button' class='mf yellow' onclick="setCellColor('#fff066')">Yellow</button>
<button type='button' class='mf green' onclick="setCellColor('#93d050')">Green</button>
<button type='button' class='mf blue' onclick="setCellColor('#9dc3e6')">Blue</button>
<button type='button' class='mf white' onclick="setCellColor('#ffffff')">White</button>
<button type='button' class='mf clear' onclick="setCellColor('')">Clear Cell</button>
<span class='mfLabel'>Text</span>
<button type='button' class='mf tred' onclick="setTextColor('#c00000')">Red Text</button>
<button type='button' class='mf tyellow' onclick="setTextColor('#bf9000')">Gold Text</button>
<button type='button' class='mf tgreen' onclick="setTextColor('#00b050')">Green Text</button>
<button type='button' class='mf tblue' onclick="setTextColor('#0070c0')">Blue Text</button>
<button type='button' class='mf clear' onclick="setTextColor('')">Clear Text</button>
<span class='mfLabel'>Size</span>
<button type='button' onclick="setFontSize('10')">10</button>
<button type='button' onclick="setFontSize('12')">12</button>
<button type='button' onclick="setFontSize('14')">14</button>
<button type='button' onclick="setFontSize('16')">16</button>
<button type='button' onclick="setFontSize('18')">18</button>
<button type='button' onclick="setFontSize('20')">20</button>
<button type='button' onclick="toggleBold()"><b>Bold</b></button>
<button type='button' onclick="openRichTextEditor()">Edit Words</button>
</div>"""
    body += "<div class='workspace'>"
    if editable:
        sched_date = get_schedule_date_settings(active)
        checked_auto = 'checked' if sched_date['auto_today'] else ''
        checked_manual = '' if sched_date['auto_today'] else 'checked'
        body += f'''<div class="scheduleDateBox">
            <span class="scheduleDateTitle">Schedule Date: {html.escape(sched_date['display_date'])}</span>
            <button type="button" onclick="document.getElementById('dateBoxForm').style.display='block'">Change Date</button>
        </div>
        <div id="dateBoxForm" class="scheduleDateBox" style="display:none">
          <form method="post" action="/set_schedule_date">
            <input type="hidden" name="sheet" value="{html.escape(active)}">
            <label><input type="radio" name="date_mode" value="auto" {checked_auto}> Auto Today</label>
            <label><input type="radio" name="date_mode" value="manual" {checked_manual}> Manual Date</label>
            <input type="date" name="schedule_date" value="{html.escape(sched_date['schedule_date'])}">
            <button type="submit" class="green">Save Date</button>
            <button type="button" onclick="document.getElementById('dateBoxForm').style.display='none'">Cancel</button>
          </form>
        </div>'''
        body += f"<form id='sheetForm' method='post' action='/save_command'><input type='hidden' name='sheet' value='{html.escape(active)}'>"
    body += "<div class='scheduleZoomBar'>"
    body += "<button type='button' onclick='pmwScheduleZoomOut()'>Schedule -</button>"
    body += "<button type='button' onclick='pmwScheduleZoomIn()'>Schedule +</button>"
    body += "<button type='button' onclick='pmwScheduleZoomReset()'>Reset Zoom</button>"
    body += "</div>"
    body += "<div class='sheetline'><div class='sheetwrap'><table class='sheet'><col class='num'><col class='job'><col class='note'><col class='num'><col class='job'><col class='note'>"
    sched_date = get_schedule_date_settings(active)
    datev = sched_date['display_date']
    body += f"<tr><td></td><td class='title' colspan='2'>{html.escape(d.get((1,2),'FABRICATION SCHEDULE') or 'FABRICATION SCHEDULE')}</td><td class='date' colspan='3'>{html.escape(datev)}</td></tr>"
    body += "<tr>"
    for c in DISPLAY_COLS:
        if c in (2,5): content=f"<div class='section'>{html.escape(d.get((2,c),'') or ('NUMBERING' if c==2 else 'FABRICATION'))}</div>"
        elif c in (3,6) and editable: content="<button type='submit' name='cmd' value='done' class='donebtn'>Done</button>"
        else: content=html.escape(d.get((2,c),'') or '')
        body += f"<td class='center'>{content}</td>"
    body += "</tr>"
    for r in range(3,51):
        body += "<tr>"
        for c in DISPLAY_COLS:
            v=d.get((r,c),'')
            m=meta.get((r,c),{})
            bg=m.get('bg_color',''); txt=m.get('text_color',''); link=m.get('link_path',''); label=m.get('link_label','')
            fsize=m.get('font_size',''); bold=m.get('bold',''); rich=m.get('rich_html','')
            style=''
            if bg: style += f"background-color:{html.escape(bg)};"
            if txt: style += f"color:{html.escape(txt)};"
            if fsize: style += f"font-size:{html.escape(fsize)}pt;"
            if bold: style += "font-weight:bold;"
            link_html = f"<a class='linkbtn' href='/open_link?sheet={urllib.parse.quote(active)}&row={r}&col={c}' title='{html.escape(label or link)}'>✉</a>" if link else ""
            if editable:
                cls = 'numinput' if c in (1,4) else ('jobinput' if c in (2,5) else 'noteinput')
                if rich:
                    body += f"<td style='{style}' data-row='{r}' data-col='{c}'><div class='cellbox'><div class='cellinput richCell' contenteditable='true' data-row='{r}' data-col='{c}' style='{style}'>{rich}</div><input class='plainHidden' name='cell_{r}_{c}' data-row='{r}' data-col='{c}' value='{html.escape(v, quote=True)}' autocomplete='off'><input type='hidden' name='bg_{r}_{c}' value='{html.escape(bg, quote=True)}'><input type='hidden' name='txt_{r}_{c}' value='{html.escape(txt, quote=True)}'><input type='hidden' name='link_{r}_{c}' value='{html.escape(link, quote=True)}'><input type='hidden' name='label_{r}_{c}' value='{html.escape(label, quote=True)}'><input type='hidden' name='fsize_{r}_{c}' value='{html.escape(fsize, quote=True)}'><input type='hidden' name='bold_{r}_{c}' value='{html.escape(bold, quote=True)}'><input type='hidden' name='rich_{r}_{c}' value='{html.escape(rich, quote=True)}'><input type='hidden' name='loaded_{r}_{c}' value='{html.escape(m.get('updated_at','') or '', quote=True)}'>{link_html}</div></td>"
                else:
                    body += f"<td style='{style}' data-row='{r}' data-col='{c}'><div class='cellbox'><input class='cellinput {cls}' name='cell_{r}_{c}' data-row='{r}' data-col='{c}' style='{style}' value='{html.escape(v, quote=True)}' autocomplete='off' {("inputmode='tel' enterkeyhint='next' tabindex='" + str((r-2) if c==1 else (100+r-2)) + "'" if c in (1,4) else "tabindex='" + str(1000 + (r*10) + c) + "'")}><input type='hidden' name='bg_{r}_{c}' value='{html.escape(bg, quote=True)}'><input type='hidden' name='txt_{r}_{c}' value='{html.escape(txt, quote=True)}'><input type='hidden' name='link_{r}_{c}' value='{html.escape(link, quote=True)}'><input type='hidden' name='label_{r}_{c}' value='{html.escape(label, quote=True)}'><input type='hidden' name='fsize_{r}_{c}' value='{html.escape(fsize, quote=True)}'><input type='hidden' name='bold_{r}_{c}' value='{html.escape(bold, quote=True)}'><input type='hidden' name='rich_{r}_{c}' value='{html.escape(rich, quote=True)}'><input type='hidden' name='loaded_{r}_{c}' value='{html.escape(m.get('updated_at','') or '', quote=True)}'>{link_html}</div></td>"
            else:
                body += f"<td style='{style}'>{rich if rich else html.escape(v)} {link_html}</td>"
        body += "</tr>"
    body += "</table></div>"
    if editable:
        body += """<div class='buttons'>
<button type='submit' name='cmd' value='clear_schedule' onclick="return confirm('Clear the active schedule?')">Clear Schedule</button>
<button type='submit' name='cmd' value='email_schedule'>Email Schedule</button>
<button type='submit' name='cmd' value='print_pdf'>Print PDF</button>
<button type='submit' name='cmd' value='delete_comments'>Delete Comments</button>
<button type='submit' name='cmd' value='sort_numbering'>Sort Numbering</button>
<button type='submit' name='cmd' value='sort_fabrication'>Sort Fabrication</button>
<button type='button' class='green' onclick='markSelectedComplete()'>Mark Complete</button>
<button type='button' onclick='window.print()'>Browser Print</button>
<button type='button' onclick='openSnipBox()'>Snip / Print / Email</button>
</div>
<div class='mobileActionBar'>
<button type='button' class='pmwUndoActionButton' onclick='undoLastCellEdit()'>Undo</button>
<button type='submit' name='cmd' value='email_schedule'>Email PDF</button>
<button type='submit' name='cmd' value='print_pdf'>Print PDF</button>
<button type='submit' name='cmd' value='sort_numbering'>Sort Numbering</button>
<button type='submit' name='cmd' value='sort_fabrication'>Sort Fabrication</button>
<button type='button' onclick='openSnipBox()'>Snip / Print / Email</button>
<button type='button' onclick='markSelectedComplete()'>Mark Complete</button>
<button type='button' onclick='pmwScheduleZoomOut()'>Schedule -</button>
<button type='button' onclick='pmwScheduleZoomIn()'>Schedule +</button>
<button type='button' onclick='pmwScheduleZoomReset()'>Reset Zoom</button>
</div>"""
    body += "</div>"
    if editable:
        body += """
<div id='snipBox' style='display:none;position:fixed;right:25px;top:90px;background:white;border:2px solid #107c41;box-shadow:0 2px 12px #555;padding:12px;z-index:50;width:310px'>
<h3 style='margin:0 0 8px'>Snip Schedule Area</h3>
<p class='small'>Type the schedule numbers you want, not Excel row numbers. Example: 1 to 5.</p>
<label>Start #</label><input name='snip_start' value='1' style='width:60px'>
<label>End #</label><input name='snip_end' value='5' style='width:60px'>
<br><br>
<label>Section</label>
<select name='snip_side'>
<option value='both'>Both sides</option>
<option value='numbering'>Numbering only</option>
<option value='fabrication'>Fabrication only</option>
</select>
<br><br>
<button name='cmd' value='snip_pdf' class='green'>Create Snip PDF</button>
<button type='button' onclick='openSnipBox(false)'>Cancel</button>
</div>
<div id='richBox' style='display:none;position:fixed;left:50%;top:90px;transform:translateX(-50%);background:white;border:2px solid #107c41;box-shadow:0 2px 12px #555;padding:12px;z-index:60;width:520px;color:#111'>
<h3 style='margin:0 0 8px'>Edit Selected Words In Cell</h3>
<p class='small'>Highlight words inside this box, then use the buttons. Click Save when done.</p>
<div style='display:flex;gap:5px;margin-bottom:8px;align-items:center'>
<button type='button' onclick="richCmd('bold')"><b>B</b></button>
<button type='button' onclick="richColor('#c00000')">Red text</button>
<button type='button' onclick="richColor('#00b050')">Green text</button>
<button type='button' onclick="richColor('#0070c0')">Blue text</button>
<button type='button' onclick="richColor('#bf9000')">Gold text</button>
<select onchange="richFontSize(this.value); this.selectedIndex=0"><option value=''>Size</option><option value='10'>10</option><option value='12'>12</option><option value='14'>14</option><option value='16'>16</option><option value='18'>18</option><option value='20'>20</option></select>
</div>
<div id='richEditor' class='richEditor' contenteditable='true'></div>
<br><button type='button' class='green' onclick='saveRichText()'>Save Rich Text</button> <button type='button' onclick='closeRichTextEditor()'>Cancel</button>
</div>
<script>
(function(){
  window.selectedCells = new Set();
  window.isSelectingCells = false;

  function cellKey(el){ return el.dataset.row + '_' + el.dataset.col; }

  window.isRichCell = function(el){
    return el && el.classList && el.classList.contains('richCell');
  }

  window.syncCell = function(el){
    if(!el) return;
    const r=el.dataset.row, c=el.dataset.col;
    const plain=document.querySelector(`input[name="cell_${r}_${c}"]`);
    const rich=document.querySelector(`input[name="rich_${r}_${c}"]`);
    if(window.isRichCell(el)){
      if(plain) plain.value = (el.textContent || '').replace(/\s+/g,' ').trim();
      if(rich) rich.value = el.innerHTML;
    }
  }

  window.updateSelectedVisuals = function(){
    document.querySelectorAll('td.selectedCell').forEach(td=>td.classList.remove('selectedCell'));
    window.selectedCells.forEach(k=>{
      const [r,c]=k.split('_');
      const el=document.querySelector(`.cellinput[data-row="${r}"][data-col="${c}"]`);
      if(el) el.closest('td').classList.add('selectedCell');
    });
  }

  window.selectOnly = function(el){
    window.selectedCells.clear();
    window.selectedCells.add(cellKey(el));
    window.activeCell=el;
    window.updateSelectedVisuals();
  }

  window.addSelect = function(el){
    window.selectedCells.add(cellKey(el));
    window.activeCell=el;
    window.updateSelectedVisuals();
  }

  document.querySelectorAll('.cellinput').forEach(el=>{
    el.addEventListener('keydown', function(e){
      if(e.key === 'Enter'){
        e.preventDefault(); e.stopPropagation();
        window.syncCell(this);
        const row=parseInt(this.dataset.row,10)+1;
        const col=this.dataset.col;
        const next=document.querySelector(`.cellinput[data-row="${row}"][data-col="${col}"]`);
        if(next){ next.focus(); window.selectOnly(next); if(next.select) next.select(); }
        return false;
      }
    });
    el.addEventListener('mousedown', function(e){
      window.isSelectingCells = true;
      if(e.ctrlKey || e.metaKey){ window.addSelect(this); }
      else{ window.selectOnly(this); }
    });
    el.addEventListener('mouseover', function(){ if(window.isSelectingCells){ window.addSelect(this); } });
    el.addEventListener('focus', function(){ if(!window.isSelectingCells) window.selectOnly(this); });
    el.addEventListener('click', function(e){ if(e.ctrlKey || e.metaKey) window.addSelect(this); else if(!window.isSelectingCells) window.selectOnly(this); });
    el.addEventListener('input', function(){ window.syncCell(this); });
    el.addEventListener('blur', function(){ window.syncCell(this); });
  });

  document.addEventListener('mouseup', function(){ window.isSelectingCells=false; });

  const form=document.getElementById('sheetForm');
  if(form){ form.addEventListener('submit', function(){ document.querySelectorAll('.cellinput').forEach(window.syncCell); }); }
})();

function selectedInputs(){
  const arr=[];
  if(!window.selectedCells || window.selectedCells.size===0){
    if(window.activeCell) return [window.activeCell];
    return [];
  }
  window.selectedCells.forEach(k=>{
    const [r,c]=k.split('_');
    const el=document.querySelector(`.cellinput[data-row="${r}"][data-col="${c}"]`);
    if(el) arr.push(el);
  });
  return arr;
}

function setCellColor(color){
  const cells=selectedInputs();
  if(cells.length===0){ alert('Click a cell first.'); return; }
  cells.forEach(el=>{
    const r=el.dataset.row, c=el.dataset.col;
    el.style.backgroundColor=color || '';
    el.closest('td').style.backgroundColor=color || '';
    const h=document.querySelector(`input[name="bg_${r}_${c}"]`);
    if(h) h.value=color || '';
  });
}

function setTextColor(color){
  const cells=selectedInputs();
  if(cells.length===0){ alert('Click a cell first.'); return; }
  cells.forEach(el=>{
    const r=el.dataset.row, c=el.dataset.col;
    el.style.color=color || '';
    el.closest('td').style.color=color || '';
    const h=document.querySelector(`input[name="txt_${r}_${c}"]`);
    if(h) h.value=color || '';
  });
}

function setFontSize(size){
  const cells=selectedInputs();
  if(cells.length===0){ alert('Click a cell first.'); return; }
  cells.forEach(el=>{
    const r=el.dataset.row, c=el.dataset.col;
    el.style.fontSize=size ? size+'pt' : '';
    el.closest('td').style.fontSize=size ? size+'pt' : '';
    const h=document.querySelector(`input[name="fsize_${r}_${c}"]`);
    if(h) h.value=size || '';
  });
}

function toggleBold(){
  const cells=selectedInputs();
  if(cells.length===0){ alert('Click a cell first.'); return; }
  cells.forEach(el=>{
    const r=el.dataset.row, c=el.dataset.col;
    const h=document.querySelector(`input[name="bold_${r}_${c}"]`);
    const on = !(h && h.value === '1');
    el.style.fontWeight = on ? 'bold' : '';
    el.closest('td').style.fontWeight = on ? 'bold' : '';
    if(h) h.value = on ? '1' : '';
  });
}

function openSnipBox(show=true){
  const box=document.getElementById('snipBox');
  if(!box) return;
  box.style.display = show ? 'block' : 'none';
}

function stripTags(html){
  const d=document.createElement('div'); d.innerHTML=html; return d.textContent || d.innerText || '';
}

function htmlEscape(s){
  return (s || '').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}

function openRichTextEditor(){
  if(!window.activeCell){ alert('Click one cell first.'); return; }
  const r=activeCell.dataset.row, c=activeCell.dataset.col;
  const richHidden=document.querySelector(`input[name="rich_${r}_${c}"]`);
  const editor=document.getElementById('richEditor');
  if(richHidden && richHidden.value){
    editor.innerHTML = richHidden.value;
  }else if(window.isRichCell(activeCell)){
    editor.innerHTML = activeCell.innerHTML;
  }else{
    editor.innerHTML = htmlEscape(activeCell.value || '');
  }
  document.getElementById('richBox').style.display='block';
  setTimeout(()=>editor.focus(), 50);
}

function closeRichTextEditor(){
  document.getElementById('richBox').style.display='none';
}

function richCmd(cmd){ document.execCommand(cmd, false, null); }
function richColor(color){ document.execCommand('foreColor', false, color); }
function richFontSize(size){
  const map={'10':'2','12':'3','14':'4','16':'5','18':'6','20':'7'};
  document.execCommand('fontSize', false, map[size] || '3');
}

function attachRichHandlers(div){
  div.addEventListener('keydown', function(e){
    if(e.key === 'Enter'){
      e.preventDefault(); e.stopPropagation();
      window.syncCell(this);
      const next=document.querySelector(`.cellinput[data-row="${parseInt(this.dataset.row,10)+1}"][data-col="${this.dataset.col}"]`);
      if(next){ next.focus(); window.selectOnly(next); }
      return false;
    }
  });
  div.addEventListener('focus', function(){ window.selectOnly(this); });
  div.addEventListener('click', function(){ window.selectOnly(this); });
  div.addEventListener('input', function(){ window.syncCell(this); });
  div.addEventListener('blur', function(){ window.syncCell(this); });
}

function saveRichText(){
  if(!window.activeCell){ closeRichTextEditor(); return; }
  const r=activeCell.dataset.row, c=activeCell.dataset.col;
  const editor=document.getElementById('richEditor');
  const rich=editor.innerHTML;
  const plain=stripTags(rich).replace(/\s+/g,' ').trim();

  const richHidden=document.querySelector(`input[name="rich_${r}_${c}"]`);
  const plainHidden=document.querySelector(`input[name="cell_${r}_${c}"]`);
  if(richHidden) richHidden.value=rich;
  if(plainHidden) plainHidden.value=plain;

  if(window.isRichCell(activeCell)){
    activeCell.innerHTML=rich;
    window.syncCell(activeCell);
  }else{
    const old=activeCell;
    const div=document.createElement('div');
    div.className='cellinput richCell';
    div.contentEditable='true';
    div.dataset.row=r;
    div.dataset.col=c;
    div.style.cssText=old.style.cssText;
    div.innerHTML=rich;
    old.parentNode.insertBefore(div, old);
    old.classList.add('plainHidden');
    attachRichHandlers(div);
    window.activeCell=div;
    window.selectOnly(div);
    window.syncCell(div);
  }
  closeRichTextEditor();
}
window.mobileScheduleZoom = window.mobileScheduleZoom || 1.0;
function applyMobileZoom(){
  const wrap=document.querySelector('.sheetwrap');
  const label=document.getElementById('mobileZoomLabel');
  if(!wrap) return;
  const z=window.mobileScheduleZoom;
  wrap.style.transform='scale('+z+')';
  wrap.style.width=(100/z)+'%';
  wrap.style.marginBottom=((1-z)*900)+'px';
  if(label) label.textContent=Math.round(z*100)+'%';
  try{ localStorage.setItem('pmw_mobile_zoom', String(z)); }catch(e){}
}
function mobileZoomOut(){
  window.mobileScheduleZoom=Math.max(0.45, (window.mobileScheduleZoom||1)-0.1);
  applyMobileZoom();
}
function mobileZoomIn(){
  window.mobileScheduleZoom=Math.min(1.25, (window.mobileScheduleZoom||1)+0.1);
  applyMobileZoom();
}
document.addEventListener('DOMContentLoaded', function(){
  try{
    const saved=parseFloat(localStorage.getItem('pmw_mobile_zoom')||'1');
    if(saved) window.mobileScheduleZoom=saved;
  }catch(e){}
  if(window.innerWidth <= 800) applyMobileZoom();
});

function getHiddenVal(name){ const el=document.querySelector(`input[name="${name}"]`); return el ? el.value : ''; }
function plainTextFromCell(el){
  if(!el) return '';
  if(el.classList && el.classList.contains('richCell')) return (el.textContent || '').replace(/\s+/g,' ').trim();
  return (el.value || '').trim();
}
function richHtmlFromCell(el){
  if(!el) return '';
  if(el.classList && el.classList.contains('richCell')) return el.innerHTML || '';
  const r=el.dataset.row, c=el.dataset.col;
  return getHiddenVal(`rich_${r}_${c}`);
}
let pmwSaveTimers = {};
function showSaveStatus(text){
  const s=document.getElementById('saveStatus');
  if(!s) return;
  s.textContent=text;
  s.style.display='block';
  clearTimeout(window._pmwSaveStatusTimer);
  window._pmwSaveStatusTimer=setTimeout(()=>{ s.style.display='none'; }, 1800);
}
function autosaveCell(el){
  if(!el || !el.dataset) return;
  const r=el.dataset.row, c=el.dataset.col;
  const key=r+'_'+c;
  clearTimeout(pmwSaveTimers[key]);
  pmwSaveTimers[key]=setTimeout(()=>{
    const sheetInput=document.querySelector('input[name="sheet"]');
    const payload={
      sheet: sheetInput ? sheetInput.value : 'Fabrication Schedule',
      row: r,
      col: c,
      value: plainTextFromCell(el),
      bg_color: getHiddenVal(`bg_${r}_${c}`),
      text_color: getHiddenVal(`txt_${r}_${c}`),
      link_path: getHiddenVal(`link_${r}_${c}`),
      link_label: getHiddenVal(`label_${r}_${c}`),
      font_size: getHiddenVal(`fsize_${r}_${c}`),
      bold: getHiddenVal(`bold_${r}_${c}`),
      rich_html: richHtmlFromCell(el),
      loaded_at: getHiddenVal(`loaded_${r}_${c}`)
    };
    showSaveStatus('Saving...');
    fetch('/autosave_cell', {
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify(payload)
    }).then(r=>r.json()).then(j=>{
      if(j.ok){
        showSaveStatus('Saved');
        const loaded=document.querySelector(`input[name="loaded_${payload.row}_${payload.col}"]`);
        if(loaded && j.saved_at){ loaded.value = j.saved_at; }
      }else if(j.conflict){
        showSaveStatus('Conflict - refresh needed');
        alert('This cell was changed by another user after your page loaded. Your change was NOT saved. The page will refresh so you can see the latest schedule.');
        window.location.reload();
      }else{
        showSaveStatus('Save failed');
      }
    }).catch(()=>showSaveStatus('Save failed'));
  }, 450);
}
document.addEventListener('DOMContentLoaded', function(){
  document.querySelectorAll('.cellinput').forEach(el=>{
    el.addEventListener('input', ()=>autosaveCell(el));
    el.addEventListener('change', ()=>autosaveCell(el));
    el.addEventListener('blur', ()=>autosaveCell(el));
  });
  ['setCellColor','setTextColor','setFontSize','toggleBold'].forEach(fn=>{
    const old=window[fn];
    if(typeof old === 'function'){
      window[fn]=function(...args){
        const ret=old.apply(this,args);
        setTimeout(()=>{ (selectedInputs ? selectedInputs() : []).forEach(autosaveCell); }, 50);
        return ret;
      }
    }
  });
  const oldSaveRich=window.saveRichText;
  if(typeof oldSaveRich === 'function'){
    window.saveRichText=function(...args){
      const ret=oldSaveRich.apply(this,args);
      setTimeout(()=>{ if(window.activeCell) autosaveCell(window.activeCell); }, 50);
      return ret;
    }
  }
});

/* ===== V26 DESKTOP-ONLY ARROW NAVIGATION ===== */
function pmwIsDesktopForArrows(){
  return window.matchMedia && window.matchMedia('(min-width: 801px)').matches;
}
function pmwCellAt(row, col){
  return document.querySelector(`.cellinput[data-row="${row}"][data-col="${col}"]`);
}
function pmwMoveCell(current, rowDelta, colDelta){
  if(!current || !pmwIsDesktopForArrows()) return false;
  const r=parseInt(current.dataset.row || '0',10);
  const c=parseInt(current.dataset.col || '0',10);
  const next=pmwCellAt(r + rowDelta, c + colDelta);
  if(next){
    try{ if(window.syncCell) window.syncCell(current); }catch(e){}
    next.focus();
    if(next.select) next.select();
    try{
      if(window.selectOnly) window.selectOnly(next);
    }catch(e){}
    next.scrollIntoView({block:'nearest', inline:'nearest'});
    return true;
  }
  return false;
}
document.addEventListener('keydown', function(e){
  if(!pmwIsDesktopForArrows()) return;
  const el=document.activeElement;
  if(!el || !el.classList || !el.classList.contains('cellinput')) return;

  // Let Alt+Arrow keep normal cursor/navigation behavior if needed.
  if(e.altKey) return;

  // V51.1:
  // Normal mode: arrows move around the grid.
  // Edit mode: left/right stay inside the cell text; up/down still move cells.
  // A plain cell enters edit mode after typing, double-click, F2, or mouse click inside the cell.
  if(!el.dataset.editMode) el.dataset.editMode='0';

  if(e.key === 'F2'){
    el.dataset.editMode='1';
    const len=(el.value || '').length;
    try{ el.setSelectionRange(len,len); }catch(err){}
    e.preventDefault();
    e.stopPropagation();
    e.stopImmediatePropagation();
    return;
  }

  if(e.key === 'Escape'){
    el.dataset.editMode='0';
    e.preventDefault();
    e.stopPropagation();
    e.stopImmediatePropagation();
    return;
  }

  // Printable typing puts the cell into edit mode.
  if(e.key && e.key.length === 1 && !e.ctrlKey && !e.metaKey && !e.altKey){
    el.dataset.editMode='1';
    return;
  }

  // If editing text, left/right should move the caret inside the input.
  if((e.key === 'ArrowLeft' || e.key === 'ArrowRight') && el.dataset.editMode === '1'){
    e.stopPropagation();
    e.stopImmediatePropagation();
    return;
  }

  let moved=false;
  if(e.key === 'ArrowUp') moved=pmwMoveCell(el,-1,0);
  if(e.key === 'ArrowDown') moved=pmwMoveCell(el,1,0);
  if(e.key === 'ArrowLeft') moved=pmwMoveCell(el,0,-1);
  if(e.key === 'ArrowRight') moved=pmwMoveCell(el,0,1);

  if(moved){
    e.preventDefault();
    e.stopPropagation();
    e.stopImmediatePropagation();
  }
}, true);

document.addEventListener('dblclick', function(e){
  const el=e.target;
  if(pmwIsDesktopForArrows() && el && el.classList && el.classList.contains('cellinput')){
    el.dataset.editMode='1';
    const len=(el.value || '').length;
    try{ el.setSelectionRange(len,len); }catch(err){}
  }
}, true);

document.addEventListener('mousedown', function(e){
  const el=e.target;
  if(pmwIsDesktopForArrows() && el && el.classList && el.classList.contains('cellinput')){
    // A direct mouse click means user is trying to work inside this cell.
    // Delay so focus/selection happens first.
    setTimeout(function(){
      if(document.activeElement === el){
        el.dataset.editMode='1';
      }
    }, 30);
  }
}, true);

/* ===== V30 CLEAR CELL + TICKET LINK ===== */
function clearSelectedCells(){
  const cells = (typeof selectedInputs === 'function') ? selectedInputs() : (window.activeCell ? [window.activeCell] : []);
  if(!cells || cells.length===0){ alert('Click a cell first.'); return; }
  cells.forEach(el=>{
    const r=el.dataset.row, c=el.dataset.col;
    if(el.classList && el.classList.contains('richCell')){
      el.innerHTML='';
    }else{
      el.value='';
    }
    const plain=document.querySelector(`input[name="cell_${r}_${c}"]`);
    if(plain) plain.value='';
    ['bg','txt','link','label','fsize','bold','rich'].forEach(prefix=>{
      const h=document.querySelector(`input[name="${prefix}_${r}_${c}"]`);
      if(h) h.value='';
    });
    el.style.backgroundColor='';
    el.style.color='';
    el.style.fontSize='';
    el.style.fontWeight='';
    const td=el.closest('td');
    if(td){
      td.style.backgroundColor='';
      td.style.color='';
      td.style.fontSize='';
      td.style.fontWeight='';
      const linkBtn=td.querySelector('.linkbtn');
      if(linkBtn) linkBtn.remove();
    }
    try{ if(typeof autosaveCell === 'function') autosaveCell(el); }catch(e){}
  });
}

// ===== v51 Plain Cell Edit Arrows =====
(function(){
  const AUTO_REFRESH_MS = 5 * 60 * 1000;
  const RETURN_REFRESH_AFTER_MS = 45 * 1000;
  let lastHiddenAt = null;
  let editing = false;
  let lastInputAt = 0;

  function isEditingNow(){
    const ae = document.activeElement;
    const tag = ae ? (ae.tagName || '').toLowerCase() : '';
    const activeEditable = ae && (ae.isContentEditable || tag === 'input' || tag === 'textarea' || tag === 'select');
    const recentInput = (Date.now() - lastInputAt) < 8000;
    return editing || activeEditable || recentInput;
  }

  function safeReload(reason){
    if(isEditingNow()) return;
    try { sessionStorage.setItem('pmw_auto_refresh_reason', reason || 'refresh'); } catch(e) {}
    window.location.reload();
  }

  document.addEventListener('input', function(){ lastInputAt = Date.now(); }, true);

  document.addEventListener('focusin', function(e){
    const t = e.target;
    if(t && (t.matches('input, textarea, select') || t.isContentEditable)){
      editing = true;
    }
  }, true);

  document.addEventListener('focusout', function(){
    lastInputAt = Date.now();
    setTimeout(function(){ editing = false; }, 2500);
  }, true);

  document.addEventListener('visibilitychange', function(){
    if(document.hidden){
      lastHiddenAt = Date.now();
    } else {
      if(lastHiddenAt && (Date.now() - lastHiddenAt) > RETURN_REFRESH_AFTER_MS){
        safeReload('Returned to PMW tab');
      }
      lastHiddenAt = null;
    }
  });

  window.addEventListener('pageshow', function(e){
    if(e.persisted){
      safeReload('Phone/browser restored PMW page');
    }
  });

  setInterval(function(){
    if(document.hidden) return;
    safeReload('Auto refresh');
  }, AUTO_REFRESH_MS);

  window.addEventListener('load', function(){
    try {
      const reason = sessionStorage.getItem('pmw_auto_refresh_reason');
      if(reason){
        sessionStorage.removeItem('pmw_auto_refresh_reason');
        const note = document.createElement('div');
        note.textContent = 'PMW auto-refreshed: ' + reason;
        note.style.cssText = 'position:fixed;right:12px;bottom:12px;background:#d4edda;border:1px solid #28a745;padding:8px 10px;z-index:9999;font-weight:bold;border-radius:4px';
        document.body.appendChild(note);
        setTimeout(function(){ note.remove(); }, 3500);
      }
    } catch(e) {}
  });
})();


function markSelectedComplete(){
  const cell = window.activeCell || document.querySelector('.selectedCell');
  if(!cell){
    alert('Click any cell in the job row first, then click Mark Complete.');
    return;
  }
  const r = cell.dataset.row;
  const c = cell.dataset.col;
  if(!r || !c){
    alert('Click any cell in the job row first, then click Mark Complete.');
    return;
  }
  if(!confirm('Mark this job row complete and save it to Job History?')) return;
  const f = document.createElement('form');
  f.method = 'POST';
  f.action = '/mark_complete';
  const sheetInput = document.querySelector('input[name="sheet"]');
  const sheet = sheetInput ? sheetInput.value : 'Fabrication Schedule';
  const fields = {sheet: sheet, row: r, col: c};
  for(const k in fields){
    const i = document.createElement('input');
    i.type = 'hidden';
    i.name = k;
    i.value = fields[k];
    f.appendChild(i);
  }
  document.body.appendChild(f);
  f.submit();
}


// ===== v49.5 schedule-only zoom buttons =====
(function(){
  window.pmwScheduleZoom = window.pmwScheduleZoom || 1.0;
  function applyScheduleZoom(){
    const wrap = document.querySelector('.sheetwrap');
    if(!wrap) return;
    let z = window.pmwScheduleZoom || 1;
    z = Math.max(0.55, Math.min(2.5, z));
    window.pmwScheduleZoom = z;
    wrap.style.setProperty('transform', 'scale(' + z + ')', 'important');
    wrap.style.setProperty('transform-origin', 'top left', 'important');
    wrap.style.setProperty('width', (100 / z) + '%', 'important');
    wrap.style.setProperty('margin-bottom', Math.max(0, (z - 1) * 700) + 'px', 'important');
    try{ localStorage.setItem('pmw_schedule_zoom', String(z)); }catch(e){}
  }
  window.pmwScheduleZoomIn = function(){ window.pmwScheduleZoom = (window.pmwScheduleZoom || 1) + 0.15; applyScheduleZoom(); };
  window.pmwScheduleZoomOut = function(){ window.pmwScheduleZoom = (window.pmwScheduleZoom || 1) - 0.15; applyScheduleZoom(); };
  window.pmwScheduleZoomReset = function(){ window.pmwScheduleZoom = 1; applyScheduleZoom(); };
  window.addEventListener('load', function(){
    try{
      const saved = parseFloat(localStorage.getItem('pmw_schedule_zoom') || '1');
      if(saved && !isNaN(saved)) window.pmwScheduleZoom = saved;
    }catch(e){}
    setTimeout(applyScheduleZoom, 150);
  });
})();


// ===== v49.6 schedule-only zoom, no page zoom =====
(function(){
  window.pmwScheduleZoom = window.pmwScheduleZoom || 1.0;

  function applyScheduleZoom(){
    const wrap = document.querySelector('.sheetwrap');
    if(!wrap) return;
    let z = window.pmwScheduleZoom || 1;
    z = Math.max(0.55, Math.min(2.5, z));
    window.pmwScheduleZoom = z;

    wrap.style.setProperty('transform', 'scale(' + z + ')', 'important');
    wrap.style.setProperty('transform-origin', 'top left', 'important');
    wrap.style.setProperty('width', (100 / z) + '%', 'important');
    wrap.style.setProperty('margin-bottom', Math.max(0, (z - 1) * 800) + 'px', 'important');

    const line = document.querySelector('.sheetline');
    if(line){
      line.style.setProperty('overflow-x', 'auto', 'important');
      line.style.setProperty('-webkit-overflow-scrolling', 'touch', 'important');
    }

    try{ localStorage.setItem('pmw_schedule_zoom', String(z)); }catch(e){}
  }

  window.pmwScheduleZoomIn = function(){
    window.pmwScheduleZoom = (window.pmwScheduleZoom || 1) + 0.20;
    applyScheduleZoom();
  };
  window.pmwScheduleZoomOut = function(){
    window.pmwScheduleZoom = (window.pmwScheduleZoom || 1) - 0.20;
    applyScheduleZoom();
  };
  window.pmwScheduleZoomReset = function(){
    window.pmwScheduleZoom = 1;
    applyScheduleZoom();
  };

  window.addEventListener('load', function(){
    try{
      const saved = parseFloat(localStorage.getItem('pmw_schedule_zoom') || '1');
      if(saved && !isNaN(saved)) window.pmwScheduleZoom = saved;
    }catch(e){}
    setTimeout(applyScheduleZoom, 200);
  });
})();




// ===== v50.1 Vertical sort number entry =====
(function(){
  function nextSortCellBelow(el){
    const r = parseInt(el.dataset.row || '0', 10);
    const c = parseInt(el.dataset.col || '0', 10);
    if(!(c === 1 || c === 4)) return null;

    for(let nr = r + 1; nr <= 50; nr++){
      const next = document.querySelector(`.cellinput[data-row="${nr}"][data-col="${c}"]`);
      if(next) return next;
    }
    return null;
  }

  function moveDown(el){
    try{
      if(typeof autosaveCell === 'function') autosaveCell(el);
    }catch(e){}
    const next = nextSortCellBelow(el);
    if(next){
      setTimeout(function(){
        next.focus();
        if(next.select) next.select();
      }, 80);
    }else{
      el.blur();
    }
  }

  document.addEventListener('keydown', function(e){
    const el = e.target;
    if(!el || !el.classList || !el.classList.contains('numinput')) return;

    if(e.key === 'Enter' || e.key === 'Tab'){
      e.preventDefault();
      moveDown(el);
    }
  }, true);

  document.addEventListener('focusin', function(e){
    const el = e.target;
    if(el && el.classList && el.classList.contains('numinput')){
      // Force iPhone-friendly sort input behavior every time it focuses.
      el.setAttribute('inputmode','tel');
      el.setAttribute('enterkeyhint','next');
      setTimeout(function(){
        try{ if(el.select) el.select(); }catch(err){}
      }, 80);
    }
  }, true);

  // iPhone Safari's Next key may move focus by tabindex before keydown is usable.
  // This tab order is set vertically, but this backup also catches left/right jumps and corrects them.
  let lastSortCell = null;
  document.addEventListener('focusin', function(e){
    const el = e.target;
    if(el && el.classList && el.classList.contains('numinput')){
      if(lastSortCell && lastSortCell !== el){
        const oldCol = parseInt(lastSortCell.dataset.col || '0', 10);
        const newCol = parseInt(el.dataset.col || '0', 10);
        const oldRow = parseInt(lastSortCell.dataset.row || '0', 10);
        const newRow = parseInt(el.dataset.row || '0', 10);
        if((oldCol === 1 || oldCol === 4) && newCol !== oldCol && newRow === oldRow){
          const corrected = nextSortCellBelow(lastSortCell);
          if(corrected){
            setTimeout(function(){
              corrected.focus();
              if(corrected.select) corrected.select();
            }, 30);
          }
        }
      }
      lastSortCell = el;
    }
  }, true);
})();


// ===== v50.7 Undo visible + arrow-key edit fix =====
(function(){
  window.pmwUndoStack = window.pmwUndoStack || [];
  const MAX_UNDO = 25;

  function isCell(el){
    return el && el.matches && el.matches('.cellinput, .richcell');
  }
  function getId(el){
    return el && el.dataset ? {row: el.dataset.row, col: el.dataset.col} : {};
  }
  function getVal(el){
    if(!el) return '';
    return (el.classList && el.classList.contains('richcell')) ? (el.innerHTML || '') : (el.value || '');
  }
  function setVal(el, v){
    if(!el) return;
    if(el.classList && el.classList.contains('richcell')) el.innerHTML = v || '';
    else el.value = v || '';
  }
  function findCell(r,c){
    return document.querySelector(`.cellinput[data-row="${r}"][data-col="${c}"], .richcell[data-row="${r}"][data-col="${c}"]`);
  }
  function saveCell(el){
    try{
      if(typeof autosaveCell === 'function') autosaveCell(el);
      else if(typeof scheduleAutosaveCell === 'function') scheduleAutosaveCell(el);
      else el.dispatchEvent(new Event('change', {bubbles:true}));
    }catch(e){
      try{ el.dispatchEvent(new Event('change', {bubbles:true})); }catch(err){}
    }
  }
  function pushUndo(el, before){
    const id = getId(el);
    if(!id.row || !id.col) return;
    window.pmwUndoStack.push({row:id.row, col:id.col, value:before, scrollX:window.scrollX, scrollY:window.scrollY});
    if(window.pmwUndoStack.length > MAX_UNDO) window.pmwUndoStack.shift();
  }

  document.addEventListener('focusin', function(e){
    const el = e.target;
    if(isCell(el)){
      el.dataset.undoStartValue = getVal(el);
      el.dataset.undoChanged = '0';
    }
  }, true);

  document.addEventListener('input', function(e){
    const el = e.target;
    if(isCell(el) && el.dataset.undoChanged !== '1'){
      pushUndo(el, el.dataset.undoStartValue || '');
      el.dataset.undoChanged = '1';
    }
  }, true);

  document.addEventListener('change', function(e){
    const el = e.target;
    if(isCell(el) && el.dataset.undoChanged !== '1' && (el.dataset.undoStartValue || '') !== getVal(el)){
      pushUndo(el, el.dataset.undoStartValue || '');
      el.dataset.undoChanged = '1';
    }
  }, true);

  window.undoLastCellEdit = function(){
    const item = window.pmwUndoStack.pop();
    if(!item){
      alert('Nothing to undo.');
      return;
    }
    const el = findCell(item.row, item.col);
    if(!el){
      alert('Could not find the cell to undo.');
      return;
    }
    setVal(el, item.value);
    el.focus();
    try{ if(el.select) el.select(); }catch(e){}
    saveCell(el);
    try{ window.scrollTo(item.scrollX || 0, item.scrollY || 0); }catch(e){}
  };

  document.addEventListener('keydown', function(e){
    const el = e.target;
    if(!isCell(el)) return;

    // Let left/right stay inside the input text.
    if(e.key === 'ArrowLeft' || e.key === 'ArrowRight'){
      e.stopImmediatePropagation();
      return;
    }

    // Up/down moves vertically in the same column.
    if(e.key === 'ArrowUp' || e.key === 'ArrowDown'){
      const id = getId(el);
      if(!id.row || !id.col) return;
      e.preventDefault();
      e.stopImmediatePropagation();
      saveCell(el);
      const nr = parseInt(id.row, 10) + (e.key === 'ArrowDown' ? 1 : -1);
      const next = findCell(nr, id.col);
      if(next){
        next.focus();
        try{ if(next.select) next.select(); }catch(err){}
      }
    }
  }, true);
})();

</script>
</form>"""
    body += "</div>"
    return page(body)


@app.route('/set_schedule_date', methods=['POST'])
@login_required
@role_required('editor')
def set_schedule_date():
    sheet = request.form.get('sheet') or 'Fabrication Schedule'
    mode = request.form.get('date_mode') or 'auto'
    chosen_date = (request.form.get('schedule_date') or '').strip()

    auto_today = 1 if mode == 'auto' else 0

    if auto_today:
        chosen_date = datetime.now().strftime("%Y-%m-%d")
    else:
        try:
            datetime.strptime(chosen_date, "%Y-%m-%d")
        except Exception:
            flash("Please choose a valid schedule date.")
            return redirect('/?sheet=' + urllib.parse.quote(sheet))

    upgrade_schedule_settings_table()
    con = db()
    if USE_POSTGRES:
        con.execute("""INSERT INTO schedule_settings(sheet_name,schedule_date,auto_today,updated_by,updated_at)
                       VALUES(?,?,?,?,?)
                       ON CONFLICT(sheet_name) DO UPDATE SET
                         schedule_date=EXCLUDED.schedule_date,
                         auto_today=EXCLUDED.auto_today,
                         updated_by=EXCLUDED.updated_by,
                         updated_at=EXCLUDED.updated_at""",
                    (sheet, chosen_date, auto_today, session.get('username',''), datetime.now().isoformat(timespec='seconds')))
    else:
        con.execute("""INSERT OR REPLACE INTO schedule_settings(sheet_name,schedule_date,auto_today,updated_by,updated_at)
                       VALUES(?,?,?,?,?)""",
                    (sheet, chosen_date, auto_today, session.get('username',''), datetime.now().isoformat(timespec='seconds')))
    con.commit()
    con.close()

    log('SET_SCHEDULE_DATE', f'{sheet} {chosen_date} auto={auto_today}')
    flash("Schedule date updated.")
    return redirect('/?sheet=' + urllib.parse.quote(sheet))



# ===== JOB HISTORY / COMPLETION TRACKING v48 FIXED =====
JOB_RE = re.compile(r"\b\d{2}-\d{3}\b")

def upgrade_job_history_table():
    try:
        con = db()
        cur = con.cursor()
        if USE_POSTGRES:
            cur.execute("""CREATE TABLE IF NOT EXISTS job_history(
                id SERIAL PRIMARY KEY,
                job_number TEXT,
                stage TEXT,
                description TEXT,
                source_sheet TEXT,
                source_row INTEGER,
                source_col INTEGER,
                completed_by TEXT,
                completed_at TEXT,
                created_at TEXT
            )""")
        else:
            cur.execute("""CREATE TABLE IF NOT EXISTS job_history(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_number TEXT,
                stage TEXT,
                description TEXT,
                source_sheet TEXT,
                source_row INTEGER,
                source_col INTEGER,
                completed_by TEXT,
                completed_at TEXT,
                created_at TEXT
            )""")

        for coldef in [
            "link_path TEXT DEFAULT ''",
            "link_label TEXT DEFAULT ''"
        ]:
            try:
                cur.execute("ALTER TABLE job_history ADD COLUMN " + coldef)
                try:
                    con.commit()
                except Exception:
                    pass
            except Exception:
                try:
                    if USE_POSTGRES:
                        con.con.rollback()
                except Exception:
                    pass
        con.commit()
        con.close()
    except Exception as e:
        print("Job history upgrade failed:", repr(e))


def extract_job_number(text):
    m = JOB_RE.search(text or "")
    return m.group(0) if m else ""

def record_job_completion(sheet, row, col, done_col, job_text, link_path='', link_label=''):
    upgrade_job_history_table()
    job_number = extract_job_number(job_text)
    stage = "Numbering" if col in (1,2,3) else "Fabrication"
    now = datetime.now().isoformat(timespec='seconds')
    user = session.get('username','')
    con = db()
    con.execute("""INSERT INTO job_history(
        job_number, stage, description, source_sheet, source_row, source_col,
        completed_by, completed_at, created_at, link_path, link_label
    ) VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
        (job_number, stage, job_text, sheet, int(row), int(done_col), user, now, now, link_path or '', link_label or ''))
    con.commit()
    con.close()
    return job_number, stage

@app.route('/mark_complete', methods=['POST'])
@login_required
@role_required('editor')
def mark_complete():
    sheet = request.form.get('sheet','Fabrication Schedule')
    try:
        row = int(request.form.get('row','0') or 0)
        col = int(request.form.get('col','0') or 0)
    except Exception:
        flash('Select a job row first.')
        return redirect('/?sheet='+urllib.parse.quote(sheet))

    if row < 3 or row > 50:
        flash('Select a valid schedule row first.')
        return redirect('/?sheet='+urllib.parse.quote(sheet))

    if col in (1,2,3):
        job_col = 2
        done_col = 3
    elif col in (4,5,6):
        job_col = 5
        done_col = 6
    else:
        flash('Select a job row first.')
        return redirect('/?sheet='+urllib.parse.quote(sheet))

    con = db()
    job = con.execute("SELECT value,link_path,link_label FROM workbook_cells WHERE sheet_name=? AND row_num=? AND col_num=?",
                      (sheet,row,job_col)).fetchone()
    con.close()

    job_text = ''
    link_path = ''
    link_label = ''
    if job:
        job_text = (job['value'] or '').strip()
        try:
            link_path = (job['link_path'] or '').strip()
            link_label = (job['link_label'] or '').strip()
        except Exception:
            link_path = ''
            link_label = ''

    if not job_text:
        flash('That row does not have a job description to complete.')
        return redirect('/?sheet='+urllib.parse.quote(sheet))

    job_number, stage = record_job_completion(sheet, row, col, done_col, job_text, link_path, link_label)

    upsert_workbook_cell(
        sheet,row,done_col,
        'X',
        '#93d050',
        '',
        '',
        '',
        '',
        '1',
        '',
        session.get('username','')
    )

    if job_number:
        flash(f'Marked complete and saved to Job History for {job_number}.')
    else:
        flash('Marked complete and saved to Job History. No job number was found in the row text.')
    log('MARK_COMPLETE', f'{sheet} row {row} {stage} {job_number} {job_text}')
    return redirect('/?sheet='+urllib.parse.quote(sheet))

@app.route('/admin/job_history')
@login_required
@role_required('admin')
def admin_job_history():
    upgrade_job_history_table()
    q = request.args.get('q','').strip()
    stage = request.args.get('stage','all').strip()

    clauses = []
    params = []
    if q:
        clauses.append("(job_number LIKE ? OR description LIKE ? OR completed_by LIKE ?)")
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]
    if stage in ("Fabrication","Numbering"):
        clauses.append("stage=?")
        params.append(stage)
    where = (" WHERE " + " AND ".join(clauses)) if clauses else ""

    con = db()
    rows = con.execute(f"""SELECT * FROM job_history
                           {where}
                           ORDER BY completed_at DESC, id DESC
                           LIMIT 500""", tuple(params)).fetchall()
    try:
        job_counts = con.execute("""SELECT job_number, COUNT(*) AS n
                                    FROM job_history
                                    WHERE COALESCE(job_number,'') <> ''
                                    GROUP BY job_number
                                    ORDER BY MAX(completed_at) DESC
                                    LIMIT 50""").fetchall()
    except Exception:
        job_counts = []
    con.close()

    body = f"""
    <div class='toolbar'>
      <b>Admin Job History</b>
      <span class='small'>Permanent completion log grouped by job number.</span>
    </div>
    <form method='get' class='toolbar'>
      <input name='q' value='{html.escape(q, quote=True)}' placeholder='Search job #, description, user'>
      <select name='stage'>
        <option value='all' {'selected' if stage=='all' else ''}>All</option>
        <option value='Fabrication' {'selected' if stage=='Fabrication' else ''}>Fabrication</option>
        <option value='Numbering' {'selected' if stage=='Numbering' else ''}>Numbering</option>
      </select>
      <button>Search</button>
      <a class='btn' href='/'>Back to Schedule</a>
    </form>
    <div class='userform'>
      <b>Recent Jobs:</b>
    """
    if job_counts:
        for jc in job_counts[:20]:
            jn = jc['job_number'] or ''
            body += f" <a class='btn' href='/admin/job_history?q={urllib.parse.quote(jn)}'>{html.escape(jn)} ({jc['n']})</a>"
    else:
        body += " No completed jobs yet."
    body += "</div>"

    body += """<form method='post' action='/admin/job_history/delete_selected' onsubmit="return confirm('Delete selected Job History item(s)? This cannot be undone.')">
      <div class='toolbar'>
        <button class='red' type='submit'>Delete Selected History Items</button>
        <button type='button' onclick="document.querySelectorAll('.jobHistBox').forEach(x=>x.checked=true)">Select All Shown</button>
        <button type='button' onclick="document.querySelectorAll('.jobHistBox').forEach(x=>x.checked=false)">Clear Selection</button>
      </div>
      <table class='admin'>
      <tr>
        <th>Delete</th><th>Completed</th><th>Job #</th><th>Stage</th><th>Description</th><th>Attachment</th><th>By</th><th>Source</th><th>Whole Job</th>
      </tr>
    """
    for r in rows:
        jn = r['job_number'] or ''
        delete_job_link = ''
        if jn:
            delete_job_link = f"<form method='post' action='/admin/job_history/delete_job' style='display:inline' onsubmit=\"return confirm('Delete ALL Job History records for {html.escape(jn, quote=True)}? This cannot be undone.')\"><input type='hidden' name='job_number' value='{html.escape(jn, quote=True)}'><button class='red' type='submit'>Delete Job</button></form>"
        try:
            lp = r['link_path'] or ''
            ll = r['link_label'] or 'Open Ticket'
        except Exception:
            lp = ''
            ll = 'Open Ticket'
        hist_link = f"<a class='btn' href='{html.escape(lp, quote=True)}'>📎 {html.escape(ll)}</a>" if lp else ''
        body += f"""<tr>
          <td><input class='jobHistBox' type='checkbox' name='history_ids' value='{r['id']}'></td>
          <td>{html.escape((r['completed_at'] or '').replace('T',' '))}</td>
          <td><b>{html.escape(jn)}</b></td>
          <td>{html.escape(r['stage'] or '')}</td>
          <td>{html.escape(r['description'] or '')}</td>
          <td>{hist_link}</td>
          <td>{html.escape(r['completed_by'] or '')}</td>
          <td>{html.escape(r['source_sheet'] or '')} R{r['source_row']} C{r['source_col']}</td>
          <td>{delete_job_link}</td>
        </tr>"""
    body += "</table></form>"
    return page(body)



@app.route('/admin/job_history/delete_selected', methods=['POST'])
@login_required
@role_required('admin')
def delete_selected_job_history():
    upgrade_job_history_table()
    ids = []
    for x in request.form.getlist('history_ids'):
        try:
            ids.append(int(x))
        except Exception:
            pass
    ids = sorted(set(ids))
    if not ids:
        flash('No Job History items selected.')
        return redirect('/admin/job_history')

    con = db()
    deleted = 0
    try:
        for hid in ids:
            con.execute("DELETE FROM job_history WHERE id=?", (hid,))
            deleted += 1
        con.commit()
    finally:
        con.close()

    log('DELETE_JOB_HISTORY_ITEMS', f'{deleted} item(s)')
    flash(f'Deleted {deleted} Job History item(s).')
    return redirect('/admin/job_history')

@app.route('/admin/job_history/delete_job', methods=['POST'])
@login_required
@role_required('admin')
def delete_job_history_for_job():
    upgrade_job_history_table()
    job_number = (request.form.get('job_number') or '').strip()
    if not job_number:
        flash('No job number selected.')
        return redirect('/admin/job_history')

    con = db()
    try:
        row = con.execute("SELECT COUNT(*) AS n FROM job_history WHERE job_number=?", (job_number,)).fetchone()
        count = int((row['n'] if row else 0) or 0)
        con.execute("DELETE FROM job_history WHERE job_number=?", (job_number,))
        con.commit()
    finally:
        con.close()

    log('DELETE_JOB_HISTORY_JOB', f'{job_number}: {count} item(s)')
    flash(f'Deleted {count} Job History item(s) for job {job_number}.')
    return redirect('/admin/job_history')


@app.route('/save_command', methods=['POST'])
@login_required
@role_required('editor')
def save_command():
    sheet=request.form.get('sheet','Fabrication Schedule')
    cmd=request.form.get('cmd','save')
    save_posted_cells(sheet)
    if cmd=='sort_numbering':
        sort_side(sheet,1,2,3); log('SORT_NUMBERING',sheet); flash('Numbering sorted.')
    elif cmd=='sort_fabrication':
        sort_side(sheet,4,5,6); log('SORT_FABRICATION',sheet); flash('Fabrication sorted.')
    elif cmd=='clear_schedule':
        con=db(); con.execute("UPDATE workbook_cells SET value='', bg_color='', text_color='', link_path='', link_label='', font_size='', bold='', rich_html='' WHERE sheet_name=? AND row_num>=3 AND col_num IN (1,2,3,4,5,6)",(sheet,)); con.commit(); con.close(); log('CLEAR_SCHEDULE',sheet); flash('Schedule cleared.')
    elif cmd=='delete_comments':
        con=db(); con.execute("UPDATE workbook_cells SET value='', bg_color='', text_color='', link_path='', link_label='', font_size='', bold='', rich_html='' WHERE sheet_name=? AND row_num>=3 AND col_num IN (3,6)",(sheet,)); con.commit(); con.close(); log('DELETE_COMMENTS',sheet); flash('Comments/status cells cleared.')
    elif cmd=='print_pdf':
        try:
            pdf_path = make_schedule_pdf(sheet)
            reveal_file(pdf_path)
            log('PRINT_PDF', os.path.basename(pdf_path))
            flash('Colored schedule PDF created. The folder was opened so you can print it.')
            return redirect('/email_ready?file=' + urllib.parse.quote(os.path.basename(pdf_path)) + '&sheet=' + urllib.parse.quote(sheet))
        except Exception as e:
            log('PRINT_PDF_FAILED', str(e))
            flash('Print PDF failed: ' + str(e))
            return redirect('/?sheet='+urllib.parse.quote(sheet))
    elif cmd=='email_schedule':
        try:
            pdf_path = make_schedule_pdf(sheet)
            ok, msg = open_outlook_draft_with_attachment(pdf_path, sheet)
            log('EMAIL_SCHEDULE_PDF', os.path.basename(pdf_path) + ' | ' + msg)
            if ok:
                flash('PDF created and an Outlook email draft opened with the PDF attached.')
                return redirect('/?sheet='+urllib.parse.quote(sheet))
            else:
                reveal_file(pdf_path)
                flash('PDF created. Outlook attachment draft did not open, so the PDF folder was opened. Attach the PDF manually if needed. Reason: ' + msg)
                return redirect('/email_ready?file=' + urllib.parse.quote(os.path.basename(pdf_path)) + '&sheet=' + urllib.parse.quote(sheet))
        except Exception as e:
            log('EMAIL_SCHEDULE_PDF_FAILED', str(e))
            flash('PDF/email failed: ' + str(e))
            return redirect('/?sheet='+urllib.parse.quote(sheet))
    elif cmd=='snip_pdf':
        try:
            start_row=request.form.get('snip_start','3')
            end_row=request.form.get('snip_end','12')
            side=request.form.get('snip_side','both')
            pdf_path=make_snip_pdf(sheet, start_row, end_row, side)
            log('SNIP_PDF', os.path.basename(pdf_path))
            reveal_file(pdf_path)
            flash('Snip PDF created using the schedule numbers you typed. The folder was opened so you can print or email it.')
            return redirect('/email_ready?file=' + urllib.parse.quote(os.path.basename(pdf_path)) + '&sheet=' + urllib.parse.quote(sheet))
        except Exception as e:
            log('SNIP_PDF_FAILED', str(e))
            flash('Snip PDF failed: ' + str(e))
            return redirect('/?sheet='+urllib.parse.quote(sheet))
    return redirect('/?sheet='+urllib.parse.quote(sheet))


@app.route('/exports/<path:filename>')
@login_required
def exports(filename):
    return send_from_directory(EXPORT_FOLDER, filename, as_attachment=True)

@app.route('/email_ready')
@login_required
def email_ready():
    filename = request.args.get('file','')
    sheet = request.args.get('sheet','Fabrication Schedule')
    safe = os.path.basename(filename)
    subject = f"PMW Schedule {datetime.now().strftime('%m-%d-%y')}"
    mailto = "mailto:?subject=" + urllib.parse.quote(subject) + "&body=" + urllib.parse.quote("Please see attached PMW schedule PDF.\n\n")
    body = f"""
    <div class='workspace'><div class='login' style='max-width:650px;margin-top:35px'>
    <h2>PDF Schedule Created</h2>
    <p>The PDF was created here:</p>
    <p><b>{html.escape(safe)}</b></p>
    <p><a class='btn green' href='/exports/{urllib.parse.quote(safe)}'>Download PDF</a>
    <a class='btn' href='{mailto}'>Open Email</a>
    <a class='btn' href='/?sheet={urllib.parse.quote(sheet)}'>Back to Schedule</a></p>
    <p class='small'>Outlook could not be controlled directly on this computer. Use Download PDF, then attach it to the email window that opens.</p>
    </div></div>
    """
    return page(body)

@app.route('/import', methods=['POST'])
@login_required
@role_required('admin')
def import_route():
    f=request.files.get('workbook')
    if not f or not f.filename:
        flash('Choose an Excel workbook first.'); return redirect('/')
    name=secure_filename(f.filename); path=os.path.join(UPLOAD_FOLDER,name); f.save(path)
    try: count=import_workbook(path); flash(f'Imported {count} filled cells.')
    except Exception as e: flash(f'Import failed: {e}')
    return redirect('/')




@app.route('/preview_ticket_drop', methods=['POST'])
@login_required
@role_required('admin')
def preview_ticket_drop():
    f=request.files.get('ticketlog')
    if not f or not f.filename:
        flash('Choose the Ticket emails drop Excel file first.')
        return redirect('/')
    name=secure_filename(f.filename)
    path=os.path.join(UPLOAD_FOLDER, "ticket_preview_" + name)
    f.save(path)
    try:
        rows=ticket_rows_from_drop(path)
    except Exception as e:
        flash(f'Could not open ticket drop workbook: {e}')
        return redirect('/')

    # show newest first
    rows=list(reversed(rows))
    body="<div class='toolbar'><b>Select Ticket Emails to Import</b><span class='small'>Check the rows you want, then click Import Selected.</span></div>"
    body+=f"<form method='post' action='/import_selected_tickets'><input type='hidden' name='preview_path' value='{html.escape(path, quote=True)}'>"
    body+="<table class='admin'><tr><th>Import</th><th>Excel Row</th><th>Received</th><th>Job #</th><th>Subject</th><th>Sender</th><th>Saved Email Path</th></tr>"
    for t in rows[:300]:
        fp=t.get('file_path','') or ''
        checked = "checked" if t == rows[0] else ""
        body+=f"<tr><td><input type='checkbox' name='rows' value='{int(t.get('source_row',0))}' {checked}></td><td>{int(t.get('source_row',0))}</td><td>{html.escape(t.get('received','') or '')}</td><td>{html.escape(t.get('job_number','') or '')}</td><td>{html.escape(t.get('subject','') or '')}</td><td>{html.escape(t.get('sender','') or '')}</td><td>{html.escape(fp)}</td></tr>"
    body+="</table><div style='padding:10px'><button class='green'>Import Selected Tickets</button> <a class='btn' href='/'>Cancel</a></div></form>"
    return page(body)

@app.route('/import_selected_tickets', methods=['POST'])
@login_required
@role_required('admin')
def import_selected_tickets():
    path=request.form.get('preview_path','')
    selected=request.form.getlist('rows')
    if not path or not os.path.exists(path):
        flash('The preview file could not be found. Please select the ticket drop workbook again.')
        return redirect('/')
    try:
        imported,msg=save_selected_ticket_rows_from_drop(path, selected)
        flash(msg)
        if imported:
            return redirect('/tickets')
    except Exception as e:
        flash(f'Selected ticket import failed: {e}')
    return redirect('/')


@app.route('/import_tickets_count', methods=['POST'])
@login_required
@role_required('admin')
def import_tickets_count_route():
    f=request.files.get('ticketlog')
    if not f or not f.filename:
        flash('Choose the Ticket emails drop Excel file first.')
        return redirect('/')
    name=secure_filename(f.filename)
    path=os.path.join(UPLOAD_FOLDER,name)
    f.save(path)
    count=request.form.get('ticket_count','1')
    try:
        tickets, msg = import_last_n_tickets_from_drop(path, count)
        if tickets:
            flash(f"Imported {len(tickets)} newest ticket drop(s).")
            return redirect('/tickets')
        flash(msg)
    except Exception as e:
        flash(f'Ticket import failed: {e}')
    return redirect('/')

@app.route('/import_one_ticket', methods=['POST'])
@login_required
@role_required('admin')
def import_one_ticket_route():
    f=request.files.get('ticketlog')
    if not f or not f.filename:
        flash('Choose the Ticket emails drop Excel file first.')
        return redirect('/')
    name=secure_filename(f.filename)
    path=os.path.join(UPLOAD_FOLDER,name)
    f.save(path)
    try:
        ticket, msg = import_one_ticket_from_drop(path, "newest")
        if ticket:
            flash(f"Imported newest ticket: {ticket['job_number'] or ''} - {ticket['subject'] or ''}")
            return redirect('/tickets')
        flash(msg)
    except Exception as e:
        flash(f'Newest ticket import failed: {e}')
    return redirect('/')



def fmt_bytes(n):
    try:
        n = float(n or 0)
    except Exception:
        n = 0
    units = ['B','KB','MB','GB','TB']
    i = 0
    while n >= 1024 and i < len(units)-1:
        n /= 1024.0
        i += 1
    return f"{n:,.1f} {units[i]}"


# ===== STORAGE HEALTH v49.2 FIX =====
def pmw_fmt_bytes(n):
    try:
        n = float(n or 0)
    except Exception:
        n = 0
    units = ['B','KB','MB','GB','TB']
    i = 0
    while n >= 1024 and i < len(units)-1:
        n /= 1024.0
        i += 1
    return f"{n:,.1f} {units[i]}"

def pmw_table_count(table):
    con = db()
    try:
        row = con.execute(f"SELECT COUNT(*) AS n FROM {table}").fetchone()
        return int((row["n"] if row else 0) or 0)
    except Exception:
        return 0
    finally:
        con.close()

def pmw_blob_expr(col):
    return f"OCTET_LENGTH({col})" if USE_POSTGRES else f"LENGTH({col})"

def pmw_sum_blob(table, col):
    con = db()
    try:
        row = con.execute(f"SELECT COALESCE(SUM({pmw_blob_expr(col)}),0) AS n FROM {table}").fetchone()
        return int((row["n"] if row else 0) or 0)
    except Exception:
        return 0
    finally:
        con.close()

def pmw_storage_health_html():
    limit_bytes = 1024 * 1024 * 1024
    msg_bytes = pmw_sum_blob("ticket_links", "cloud_file_data")
    att_bytes = pmw_sum_blob("ticket_attachments", "file_data")
    ticket_bytes = msg_bytes + att_bytes

    counts = {
        "Users": pmw_table_count("users"),
        "Schedule Cells": pmw_table_count("workbook_cells"),
        "Tickets": pmw_table_count("ticket_links"),
        "Ticket Attachments": pmw_table_count("ticket_attachments"),
        "Job History": pmw_table_count("job_history"),
        "Audit Records": pmw_table_count("audit_log"),
    }

    estimated_backup_bytes = int(ticket_bytes * 1.37) + 250000
    estimated_pct = (estimated_backup_bytes / limit_bytes) * 100 if limit_bytes else 0

    if estimated_pct < 50:
        zone = "GREEN"
        zone_color = "#0f7b2f"
        advice = "Good shape. Keep weekly backups and clean up test/duplicate tickets when convenient."
    elif estimated_pct < 75:
        zone = "YELLOW"
        zone_color = "#9a6700"
        advice = "Start cleaning older ticket attachments and watch storage more often."
    elif estimated_pct < 90:
        zone = "RED"
        zone_color = "#b42318"
        advice = "Clean up old tickets soon or consider upgrading storage."
    else:
        zone = "CRITICAL"
        zone_color = "#7f1d1d"
        advice = "Take action now: delete old tickets/attachments or upgrade the database."

    html_out = f"""
    <div class='userform'>
      <h3>Database Health</h3>
      <p><b>Render Plan:</b> Free PostgreSQL — 1 GB storage</p>
      <p><b>Ticket Email Storage:</b> {pmw_fmt_bytes(msg_bytes)}</p>
      <p><b>Ticket Attachment Storage:</b> {pmw_fmt_bytes(att_bytes)}</p>
      <p><b>Total Ticket File Storage:</b> {pmw_fmt_bytes(ticket_bytes)}</p>
      <p><b>Estimated Full Backup Size:</b> {pmw_fmt_bytes(estimated_backup_bytes)}</p>
      <p><b>Estimated File Storage Load:</b> {estimated_pct:.1f}% of 1 GB</p>
      <p><b>Status:</b> <span style='color:{zone_color};font-weight:bold'>{zone}</span></p>
      <p class='small'>{html.escape(advice)}</p>
      <p>
        <a class='btn' href='/admin/ticket_cleanup'>Open Cleanup</a>
        <a class='btn' href='/admin/backup'>Download Backup</a>
      </p>
    </div>
    <div class='userform'>
      <h3>Table Counts</h3>
      <table class='admin'>
        <tr><th>Area</th><th>Records</th></tr>
    """
    for k, v in counts.items():
        html_out += f"<tr><td>{html.escape(k)}</td><td>{v}</td></tr>"
    html_out += """
      </table>
    </div>
    """
    return html_out


@app.route('/admin/storage')
@login_required
@role_required('admin')
def admin_storage():
    con=db()
    try:
        tickets = con.execute("SELECT COUNT(*) AS n FROM ticket_links").fetchone()["n"]
    except Exception:
        tickets = 0
    try:
        atts = con.execute("SELECT COUNT(*) AS n FROM ticket_attachments").fetchone()["n"]
    except Exception:
        atts = 0
    try:
        msg_bytes = con.execute("SELECT COALESCE(SUM(OCTET_LENGTH(cloud_file_data)),0) AS n FROM ticket_links").fetchone()["n"] if USE_POSTGRES else 0
    except Exception:
        msg_bytes = 0
    try:
        att_bytes = con.execute("SELECT COALESCE(SUM(OCTET_LENGTH(file_data)),0) AS n FROM ticket_attachments").fetchone()["n"] if USE_POSTGRES else 0
    except Exception:
        att_bytes = 0
    con.close()
    total = int(msg_bytes or 0) + int(att_bytes or 0)
    body = f"""
    <h2>Admin - Ticket Storage</h2><p><a class='btn green' href='/admin/export_schedule_excel'>Export Schedule to Excel</a></p>
    <div class='userform'>
      <p><b>Tickets:</b> {tickets}</p>
      <p><b>Attachments:</b> {atts}</p>
      <p><b>Original .msg storage:</b> {fmt_bytes(msg_bytes)}</p>
      <p><b>Attachment storage:</b> {fmt_bytes(att_bytes)}</p>
      <p><b>Total stored ticket files:</b> {fmt_bytes(total)}</p>
      <p class='small'>This counts files stored inside PostgreSQL. Render free/cheap databases have limits, so review this page regularly.</p>
    </div>
    """
    body += pmw_storage_health_html()
    body += """
    <h3>Cleanup</h3>
    <div class='userform'>
      <p>Use Cleanup to delete selected old/test ticket emails and attachments when storage starts getting high.</p>
      <p><a class='btn' href='/admin/ticket_cleanup'>Open Cleanup</a></p>
    </div>
    """
    return page(body)



@app.route('/api/upload_ticket_msg', methods=['POST'])
def api_upload_ticket_msg():
    supplied_key = request.headers.get('X-PMW-UPLOAD-KEY','') or request.form.get('upload_key','')
    if supplied_key != PMW_UPLOAD_KEY:
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    f = request.files.get('msgfile')
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "Missing msgfile"}), 400

    original = secure_filename(f.filename)
    if not original.lower().endswith('.msg'):
        return jsonify({"ok": False, "error": "Only .msg files are allowed"}), 400

    subject = request.form.get('subject','').strip()
    sender = request.form.get('sender','').strip()
    received = request.form.get('received','').strip()
    job_number = request.form.get('job_number','').strip()
    local_path = request.form.get('local_path','').strip()

    safe_name = f"auto_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{original}"
    save_path = os.path.join(CLOUD_TICKET_FOLDER, safe_name)
    f.save(save_path)

    preview = parse_msg_preview(save_path)
    if not subject:
        subject = preview.get('subject','') or original
    if not sender:
        sender = preview.get('sender','')
    if not received:
        received = preview.get('date','') or datetime.now().isoformat(timespec='seconds')

    msg_bytes = read_file_bytes(save_path)

    con=db()
    existing = None
    if local_path:
        existing = con.execute("SELECT id FROM ticket_links WHERE file_path=?",(local_path,)).fetchone()
    if existing:
        ticket_id = existing['id']
        con.execute("""UPDATE ticket_links
                       SET job_number=?, subject=?, sender=?, received=?, cloud_file=?, cloud_filename=?,
                           cloud_uploaded_at=?, cloud_uploaded_by=?,
                           preview_subject=?, preview_sender=?, preview_date=?, preview_body=?, preview_status=?,
                           cloud_file_data=?
                       WHERE id=?""",
                    (job_number, subject, sender, received, safe_name, original,
                     datetime.now().isoformat(timespec='seconds'), 'outlook-auto',
                     preview.get('subject',''), preview.get('sender',''), preview.get('date',''), preview.get('body',''), preview.get('status',''),
                     msg_bytes, ticket_id))
    else:
        con.execute("""INSERT INTO ticket_links(job_number,subject,sender,received,file_path,created_at,
                       cloud_file,cloud_filename,cloud_uploaded_at,cloud_uploaded_by,
                       preview_subject,preview_sender,preview_date,preview_body,preview_status,cloud_file_data)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (job_number, subject, sender, received, local_path, datetime.now().isoformat(timespec='seconds'),
                     safe_name, original, datetime.now().isoformat(timespec='seconds'), 'outlook-auto',
                     preview.get('subject',''), preview.get('sender',''), preview.get('date',''), preview.get('body',''), preview.get('status',''),
                     msg_bytes))
        try:
            ticket_id = con.execute("SELECT LASTVAL() AS id").fetchone()["id"] if USE_POSTGRES else con.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        except Exception:
            row = con.execute("SELECT id FROM ticket_links WHERE file_path=?",(local_path,)).fetchone()
            ticket_id = row['id'] if row else None
    con.commit(); con.close()

    if ticket_id:
        # replace attachment records for this ticket and extract current attachments
        con=db()
        try:
            old_atts=con.execute("SELECT stored_filename FROM ticket_attachments WHERE ticket_id=?",(ticket_id,)).fetchall()
            for a in old_atts:
                try:
                    old_path = os.path.join(CLOUD_ATTACHMENT_FOLDER, a['stored_filename'])
                    if os.path.exists(old_path):
                        os.remove(old_path)
                except Exception:
                    pass
            con.execute("DELETE FROM ticket_attachments WHERE ticket_id=?",(ticket_id,))
            con.commit()
        finally:
            con.close()
        extracted = extract_msg_attachments(ticket_id, save_path)
    else:
        extracted = []

    return jsonify({"ok": True, "ticket_id": ticket_id, "attachments": len(extracted), "subject": subject})


@app.route('/ticket_upload/<int:ticket_id>', methods=['POST'])
@login_required
@role_required('editor')
def ticket_upload(ticket_id):
    f = request.files.get('msgfile')
    if not f or not f.filename:
        flash('Choose a .msg file first.')
        return redirect('/tickets')

    original = secure_filename(f.filename)
    if not original.lower().endswith('.msg'):
        flash('Please upload the saved Outlook .msg file only.')
        return redirect('/tickets')

    con=db()
    t=con.execute("SELECT id,job_number,subject FROM ticket_links WHERE id=?",(ticket_id,)).fetchone()
    con.close()
    if not t:
        flash('Ticket not found.')
        return redirect('/tickets')

    safe_name = f"ticket_{ticket_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{original}"
    save_path = os.path.join(CLOUD_TICKET_FOLDER, safe_name)
    f.save(save_path)

    preview = parse_msg_preview(save_path)

    # Replace attachment records for this ticket when the .msg is re-uploaded
    con=db()
    old_atts=con.execute("SELECT stored_filename FROM ticket_attachments WHERE ticket_id=?",(ticket_id,)).fetchall() if True else []
    for a in old_atts:
        try:
            old_file = a['stored_filename']
            old_path = os.path.join(CLOUD_ATTACHMENT_FOLDER, old_file)
            if os.path.exists(old_path):
                os.remove(old_path)
        except Exception:
            pass
    con.execute("DELETE FROM ticket_attachments WHERE ticket_id=?",(ticket_id,))
    con.commit(); con.close()

    extracted = extract_msg_attachments(ticket_id, save_path)

    con=db()
    msg_bytes = read_file_bytes(save_path)
    con.execute("""UPDATE ticket_links
                   SET cloud_file=?, cloud_filename=?, cloud_uploaded_at=?, cloud_uploaded_by=?,
                       preview_subject=?, preview_sender=?, preview_date=?, preview_body=?, preview_status=?,
                       cloud_file_data=?
                   WHERE id=?""",
                (safe_name, original, datetime.now().isoformat(timespec='seconds'), session.get('username',''),
                 preview.get('subject',''), preview.get('sender',''), preview.get('date',''), preview.get('body',''), preview.get('status',''),
                 msg_bytes, ticket_id))
    con.commit(); con.close()

    log('UPLOAD_TICKET_MSG', f'{ticket_id} / {original}')
    flash('Cloud ticket email uploaded. Users can now download/open it from the cloud link.')
    return redirect('/tickets')



@app.route('/ticket_attachment/<int:attachment_id>')
@login_required
def ticket_attachment(attachment_id):
    con=db()
    a=con.execute("SELECT * FROM ticket_attachments WHERE id=?",(attachment_id,)).fetchone()
    con.close()
    if not a:
        flash('Attachment not found.')
        return redirect('/tickets')
    try:
        data = a['file_data']
    except Exception:
        data = None

    if data:
        return send_bytes_as_file(data, a['original_filename'], a['content_type'] or None, as_attachment=False)

    path=os.path.join(CLOUD_ATTACHMENT_FOLDER, a['stored_filename'])
    if not os.path.exists(path):
        flash('Attachment file is missing from temporary storage. Re-upload the .msg one time so v38 can save attachments permanently in the database.')
        return redirect('/tickets')
    from flask import send_file
    return send_file(path, as_attachment=False, download_name=a['original_filename'], mimetype=a['content_type'] or None)

@app.route('/ticket_attachment_download/<int:attachment_id>')
@login_required
def ticket_attachment_download(attachment_id):
    con=db()
    a=con.execute("SELECT * FROM ticket_attachments WHERE id=?",(attachment_id,)).fetchone()
    con.close()
    if not a:
        flash('Attachment not found.')
        return redirect('/tickets')
    try:
        data = a['file_data']
    except Exception:
        data = None

    if data:
        return send_bytes_as_file(data, a['original_filename'], a['content_type'] or None, as_attachment=True)

    path=os.path.join(CLOUD_ATTACHMENT_FOLDER, a['stored_filename'])
    if not os.path.exists(path):
        flash('Attachment file is missing from temporary storage. Re-upload the .msg one time so v38 can save attachments permanently in the database.')
        return redirect('/tickets')
    from flask import send_file
    return send_file(path, as_attachment=True, download_name=a['original_filename'], mimetype=a['content_type'] or None)


@app.route('/ticket_view_email/<int:ticket_id>')
@login_required
def ticket_view_email(ticket_id):
    con=db()
    t=con.execute("SELECT * FROM ticket_links WHERE id=?",(ticket_id,)).fetchone()
    atts=con.execute("SELECT * FROM ticket_attachments WHERE ticket_id=? ORDER BY id",(ticket_id,)).fetchall()
    con.close()
    if not t:
        flash('Ticket not found.')
        return redirect('/tickets')

    def getv(k):
        try:
            return t[k] or ''
        except Exception:
            return ''

    subject = getv('preview_subject') or getv('subject')
    sender = getv('preview_sender') or getv('sender')
    date = getv('preview_date') or getv('received')
    body_text = getv('preview_body')
    status = getv('preview_status')
    cloud_file = getv('cloud_file')

    if not body_text:
        body_text = "No browser preview has been created yet. Upload or re-upload the .msg file from the Tickets page."

    body_html = html.escape(body_text).replace('\\n','<br>')
    subject_html = html.escape(subject)
    sender_html = html.escape(sender)
    date_html = html.escape(date)
    status_html = html.escape(status)

    attachments_html = "<h3>Attachments</h3>"
    if not atts:
        attachments_html += "<p>No attachments were extracted from this email.</p>"
    else:
        for a in atts:
            aid=a['id']
            fname=html.escape(a['original_filename'] or 'attachment')
            ctype=(a['content_type'] or '').lower()
            view_url=f"/ticket_attachment/{aid}"
            dl_url=f"/ticket_attachment_download/{aid}"
            attachments_html += f"<div class='ticketAttachmentCard'><b>{fname}</b><a class='btn green' href='{view_url}' target='_blank'>Open Full Screen</a> <a class='btn' href='{dl_url}'>Download</a>"
            if ctype.startswith('image/'):
                attachments_html += f"<div><img src='{view_url}' class='ticketAttachmentImage'></div>"
            elif ctype == 'application/pdf':
                attachments_html += f"<div class='mobileOpenHint'>On iPhone, tap <b>Open Full Screen</b> for easier PDF viewing and pinch zoom.</div><div class='mobilePdfPortraitNotice'>PDF preview is hidden in iPhone portrait so the page can scroll normally. Tap <b>Open Full Screen</b> to view the PDF.</div><div style='margin-top:8px'><iframe src='{view_url}' class='ticketAttachmentFrame'></iframe></div>"
            else:
                attachments_html += "<p class='small'>Preview may not be available for this file type. Use Download or Open Full Screen.</p>"
            attachments_html += "</div>"

    download = f"<a class='btn green' href='/ticket_download/{ticket_id}'>Download Original .msg</a>" if cloud_file else ""
    page_body = f"""
    <div class='ticketPreviewWrap'>
      <h2>Ticket Email + Attachments</h2>
      <div class='ticketHeaderCard'>
        <p><b>Subject:</b> {subject_html}</p>
        <p><b>From:</b> {sender_html}</p>
        <p><b>Date:</b> {date_html}</p>
        <p><b>Preview Status:</b> {status_html}</p>
        <p>{download} <a class='btn' href='/tickets'>Back to Tickets</a></p>
      </div>
      <div class='ticketAttachmentArea'>
        {attachments_html}
      </div>
      <h3>Email Body</h3>
      <div class='ticketBodyCard'>
        {body_html}
      </div>
    </div>
    """
    return page(page_body)


@app.route('/ticket_download/<int:ticket_id>')
@login_required
def ticket_download(ticket_id):
    con=db()
    t=con.execute("SELECT cloud_file,cloud_filename,cloud_file_data,file_path,subject FROM ticket_links WHERE id=?",(ticket_id,)).fetchone()
    con.close()
    if not t:
        flash('Ticket not found.')
        return redirect('/tickets')
    cloud_file = (t.get('cloud_file') if hasattr(t, 'get') else t['cloud_file']) or ''
    cloud_filename = (t.get('cloud_filename') if hasattr(t, 'get') else t['cloud_filename']) or ''
    try:
        cloud_data = t['cloud_file_data']
    except Exception:
        cloud_data = None

    if cloud_data:
        return send_bytes_as_file(cloud_data, cloud_filename or cloud_file or 'ticket.msg', 'application/vnd.ms-outlook', as_attachment=True)

    if not cloud_file:
        flash('No cloud .msg file has been uploaded for this ticket yet.')
        return redirect('/tickets')
    full_path = os.path.join(CLOUD_TICKET_FOLDER, cloud_file)
    if not os.path.exists(full_path):
        flash('Cloud file is missing from temporary storage. Re-upload the .msg one time so v38 can save it permanently in the database.')
        return redirect('/tickets')
    from flask import send_file
    return send_file(full_path, as_attachment=True, download_name=cloud_filename or cloud_file)

@app.route('/ticket_link_info/<int:ticket_id>')
@login_required
def ticket_link_info(ticket_id):
    con=db()
    t=con.execute("SELECT * FROM ticket_links WHERE id=?",(ticket_id,)).fetchone()
    con.close()
    if not t:
        flash('Ticket not found.')
        return redirect('/tickets')
    subject=html.escape(t['subject'] or '')
    job=html.escape(t['job_number'] or '')
    local_path=html.escape(t['file_path'] or '')
    cloud_file=(t.get('cloud_file') if hasattr(t, 'get') else t['cloud_file']) or ''
    cloud_status = f"<p><a class='btn green' href='/ticket_view_email/{ticket_id}'>View Email in Browser</a> <a class='btn' href='/ticket_download/{ticket_id}'>Download Original .msg</a></p>" if cloud_file else "<p><b>No cloud .msg uploaded yet.</b></p>"
    body=f"""
    <h2>Ticket Email Link</h2>
    <p><b>Job:</b> {job}</p>
    <p><b>Subject:</b> {subject}</p>
    {cloud_status}
    <h3>Original Office Path</h3>
    <p class='small'>This is the old office/server path from your ticket drop workbook.</p>
    <input value="{local_path}" style="width:90%;padding:10px" onclick="this.select()">
    <p><a class='btn' href='/tickets'>Back to Tickets</a></p>
    """
    return page(body)



@app.route('/clear_cell', methods=['POST'])
@login_required
@role_required('editor')
def clear_cell():
    sheet = request.form.get('sheet') or 'Fabrication Schedule'
    try:
        r = int(request.form.get('row'))
        c = int(request.form.get('col'))
    except Exception:
        flash('Bad cell.')
        return redirect('/?sheet='+urllib.parse.quote(sheet))
    if c not in DISPLAY_COLS:
        flash('Bad cell.')
        return redirect('/?sheet='+urllib.parse.quote(sheet))
    upsert_workbook_cell(sheet,r,c,'','','','','','','','',session.get('username',''))
    log('CLEAR_CELL', f'{sheet} R{r} C{c}')
    flash('Cell cleared, including ticket/email link.')
    return redirect('/?sheet='+urllib.parse.quote(sheet))



# ===== TICKET SCHEDULED STATUS v46.1 FIX =====
def upgrade_ticket_schedule_status_columns():
    """Track when a ticket has been added to Fabrication or Numbering."""
    try:
        con = db()
        cur = con.cursor()
        columns = [
            ("scheduled_status", "TEXT DEFAULT ''"),
            ("scheduled_side", "TEXT DEFAULT ''"),
            ("scheduled_sheet", "TEXT DEFAULT ''"),
            ("scheduled_row", "INTEGER DEFAULT 0"),
            ("scheduled_by", "TEXT DEFAULT ''"),
            ("scheduled_at", "TEXT DEFAULT ''"),
        ]
        for name, ddl in columns:
            try:
                cur.execute(f"ALTER TABLE ticket_links ADD COLUMN {name} {ddl}")
                try:
                    con.commit()
                except Exception:
                    pass
            except Exception:
                try:
                    if USE_POSTGRES:
                        con.con.rollback()
                except Exception:
                    pass
        con.commit()
        con.close()
    except Exception as e:
        print("Ticket scheduled status upgrade failed:", repr(e))

def mark_ticket_scheduled(ticket_id, side, sheet, row_num):
    try:
        upgrade_ticket_schedule_status_columns()
        con = db()
        side_label = "Fabrication" if side == "fabrication" else "Numbering"
        now = datetime.now().isoformat(timespec='seconds')
        con.execute("""UPDATE ticket_links
                       SET scheduled_status=?, scheduled_side=?, scheduled_sheet=?, scheduled_row=?,
                           scheduled_by=?, scheduled_at=?
                       WHERE id=?""",
                    ("scheduled", side_label, sheet, int(row_num or 0), session.get('username',''), now, ticket_id))
        con.commit()
        con.close()
    except Exception as e:
        print("mark_ticket_scheduled failed:", repr(e))

def ticket_scheduled_badge(row):
    """Return HTML showing whether ticket has already been added to schedule."""
    try:
        status = row["scheduled_status"] or ""
    except Exception:
        status = ""
    if status == "scheduled":
        try:
            side = row["scheduled_side"] or ""
        except Exception:
            side = ""
        try:
            sched_row = row["scheduled_row"] or ""
        except Exception:
            sched_row = ""
        try:
            by = row["scheduled_by"] or ""
        except Exception:
            by = ""
        try:
            at = row["scheduled_at"] or ""
        except Exception:
            at = ""

        detail = side or "Schedule"
        if sched_row:
            detail += f" row {sched_row}"
        if by:
            detail += f" by {html.escape(str(by))}"
        if at:
            detail += f" on {html.escape(str(at).replace('T',' '))}"

        return f"<span style='color:#0f7b2f;font-weight:bold'>✓ Added to {html.escape(detail)}</span>"

    return "<span style='color:#9a6700;font-weight:bold'>Not Scheduled</span>"



# ===== TICKET CLEANUP / STORAGE MANAGEMENT v47 =====
def fmt_bytes_cleanup(n):
    try:
        n = float(n or 0)
    except Exception:
        n = 0
    units = ['B','KB','MB','GB','TB']
    i = 0
    while n >= 1024 and i < len(units)-1:
        n /= 1024.0
        i += 1
    return f"{n:,.1f} {units[i]}"

def _blob_len_expr(col):
    return f"OCTET_LENGTH({col})" if USE_POSTGRES else f"LENGTH({col})"

def ticket_storage_totals():
    con = db()
    try:
        tickets = con.execute("SELECT COUNT(*) AS n FROM ticket_links").fetchone()["n"]
    except Exception:
        tickets = 0
    try:
        attachments = con.execute("SELECT COUNT(*) AS n FROM ticket_attachments").fetchone()["n"]
    except Exception:
        attachments = 0
    try:
        msg_bytes = con.execute(f"SELECT COALESCE(SUM({_blob_len_expr('cloud_file_data')}),0) AS n FROM ticket_links").fetchone()["n"]
    except Exception:
        msg_bytes = 0
    try:
        att_bytes = con.execute(f"SELECT COALESCE(SUM({_blob_len_expr('file_data')}),0) AS n FROM ticket_attachments").fetchone()["n"]
    except Exception:
        att_bytes = 0
    con.close()
    msg_bytes = int(msg_bytes or 0)
    att_bytes = int(att_bytes or 0)
    return {"tickets": tickets, "attachments": attachments, "msg_bytes": msg_bytes, "att_bytes": att_bytes, "total_bytes": msg_bytes + att_bytes}

def ticket_file_size(ticket_id):
    con = db()
    try:
        q = f"SELECT COALESCE({_blob_len_expr('cloud_file_data')},0) AS n FROM ticket_links WHERE id=?"
        row = con.execute(q, (ticket_id,)).fetchone()
        msg_size = int((row["n"] if row else 0) or 0)
    except Exception:
        msg_size = 0
    try:
        q = f"SELECT COALESCE(SUM({_blob_len_expr('file_data')}),0) AS n FROM ticket_attachments WHERE ticket_id=?"
        row = con.execute(q, (ticket_id,)).fetchone()
        att_size = int((row["n"] if row else 0) or 0)
    except Exception:
        att_size = 0
    try:
        att_count = con.execute("SELECT COUNT(*) AS n FROM ticket_attachments WHERE ticket_id=?", (ticket_id,)).fetchone()["n"]
    except Exception:
        att_count = 0
    con.close()
    return msg_size, att_size, att_count


# ===== BACKUP / RESTORE SYSTEM v49 =====
BACKUP_TABLES = [
    "users",
    "workbook_cells",
    "ticket_links",
    "ticket_attachments",
    "audit_log",
    "schedule_settings",
    "job_history"
]

BINARY_COLUMNS = {
    "ticket_links": ["cloud_file_data"],
    "ticket_attachments": ["file_data"]
}

def table_exists(table):
    con = db()
    try:
        if USE_POSTGRES:
            row = con.execute("SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_name=?) AS ok", (table,)).fetchone()
            ok = bool(row["ok"]) if row else False
        else:
            row = con.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone()
            ok = bool(row)
    except Exception:
        ok = False
    finally:
        con.close()
    return ok

def get_table_columns(con, table):
    if USE_POSTGRES:
        rows = con.execute("""SELECT column_name FROM information_schema.columns
                              WHERE table_name=?
                              ORDER BY ordinal_position""", (table,)).fetchall()
        return [r["column_name"] for r in rows]
    rows = con.execute(f"PRAGMA table_info({table})").fetchall()
    return [r["name"] for r in rows]

def _backup_value(table, col, val):
    if val is None:
        return None
    if col in BINARY_COLUMNS.get(table, []):
        try:
            if isinstance(val, memoryview):
                val = val.tobytes()
            if isinstance(val, str):
                # Already text, keep as text backup.
                raw = val.encode("utf-8")
            else:
                raw = bytes(val)
            return {"__pmw_blob_b64__": base64.b64encode(raw).decode("ascii")}
        except Exception:
            return None
    return val

def _restore_value(table, col, val):
    if isinstance(val, dict) and "__pmw_blob_b64__" in val:
        try:
            return base64.b64decode(val["__pmw_blob_b64__"])
        except Exception:
            return None
    return val

def build_backup_payload():
    con = db()
    payload = {
        "app": "PMW Ticket + Fabrication",
        "backup_version": 1,
        "exported_at": datetime.now().isoformat(timespec='seconds'),
        "tables": {}
    }
    try:
        for table in BACKUP_TABLES:
            if not table_exists(table):
                payload["tables"][table] = {"columns": [], "rows": [], "missing": True}
                continue
            cols = get_table_columns(con, table)
            rows = con.execute(f"SELECT * FROM {table}").fetchall()
            out_rows = []
            for row in rows:
                obj = {}
                for col in cols:
                    try:
                        obj[col] = _backup_value(table, col, row[col])
                    except Exception:
                        obj[col] = None
                out_rows.append(obj)
            payload["tables"][table] = {"columns": cols, "rows": out_rows, "missing": False}
    finally:
        con.close()
    return payload

def write_server_backup_copy():
    """Writes a server-side backup file. On Render Free, filesystem may be temporary, so download backup is still the real safety copy."""
    try:
        os.makedirs(EXPORT_FOLDER, exist_ok=True)
        payload = build_backup_payload()
        fn = "PMW_BACKUP_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".json"
        path = os.path.join(EXPORT_FOLDER, fn)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f)
        return path
    except Exception as e:
        print("Server backup copy failed:", repr(e))
        return ""

def maybe_daily_backup():
    """Disabled automatic server-side backup.

    The previous automatic backup ran before normal page requests and could timeout
    on Render when the database contained large ticket email/attachment blobs.
    Manual Admin > Backup > Download Full Backup remains available.
    """
    return ""

@app.before_request
def pmw_auto_daily_backup_hook():
    # Disabled to prevent Gunicorn worker timeouts during normal app use.
    # Manual full backups are still handled by /admin/download_backup.
    return None


# ===== EXCEL SCHEDULE EXPORT v50.4 =====
def safe_excel_color(hex_color):
    h = (hex_color or '').strip().replace('#','')
    if len(h) == 6:
        return 'FF' + h.upper()
    return None

def export_schedule_workbook():
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    sheets = sheet_names()
    if not sheets:
        sheets = ['Fabrication Schedule']

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet in sheets:
        ws = wb.create_sheet(title=(sheet[:31] or "Schedule"))
        d = cells_for(sheet)
        meta = cell_meta_for(sheet)

        # Header row
        headers = ["#", "NUMBERING", "DONE", "#", "FABRICATION", "DONE"]
        for idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=idx, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFD9EAD3")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for r in range(3, 51):
            excel_r = r - 1
            for c in DISPLAY_COLS:
                v = d.get((r,c), '')
                m = meta.get((r,c), {})
                cell = ws.cell(row=excel_r, column=c, value=v)

                bg = safe_excel_color(m.get('bg_color',''))
                txt = safe_excel_color(m.get('text_color',''))

                if bg:
                    cell.fill = PatternFill("solid", fgColor=bg)
                if txt:
                    cell.font = Font(
                        color=txt,
                        bold=bool(m.get('bold')),
                        size=float(m.get('font_size')) if str(m.get('font_size','')).strip().replace('.','',1).isdigit() else None
                    )
                else:
                    cell.font = Font(
                        bold=bool(m.get('bold')),
                        size=float(m.get('font_size')) if str(m.get('font_size','')).strip().replace('.','',1).isdigit() else None
                    )

                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border

                link = (m.get('link_path') or '').strip()
                if link:
                    # Internal app links are exported as text in adjacent note/comment style.
                    # If it looks like an http link, make it clickable.
                    if link.startswith('http://') or link.startswith('https://'):
                        cell.hyperlink = link
                        cell.style = "Hyperlink"
                    comment_label = (m.get('link_label') or 'LINK').strip()
                    try:
                        cell.value = (str(v) + "  [Attachment: " + comment_label + "]").strip()
                    except Exception:
                        pass

        # Widths to match schedule.
        widths = {1:8, 2:42, 3:14, 4:8, 5:42, 6:14}
        for c, w in widths.items():
            ws.column_dimensions[get_column_letter(c)].width = w

        for rr in range(1, 50):
            ws.row_dimensions[rr].height = 24

        ws.freeze_panes = "A2"

    return wb

@app.route('/admin/export_schedule_excel')
@login_required
@role_required('admin')
def export_schedule_excel():
    try:
        wb = export_schedule_workbook()
        fn = "PMW_Schedule_Export_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
        export_folder = globals().get('EXPORT_FOLDER', '/tmp')
        path = os.path.join(export_folder, fn)
        os.makedirs(export_folder, exist_ok=True)
        wb.save(path)
        log('EXPORT_SCHEDULE_EXCEL', fn)
        try:
            return send_file(path, as_attachment=True, download_name=fn)
        except TypeError:
            return send_file(path, as_attachment=True, attachment_filename=fn)
    except Exception as e:
        flash('Excel export failed: ' + str(e))
        return redirect('/admin/backup')


@app.route('/admin/backup')
@login_required
@role_required('admin')
def admin_backup():
    counts = {}
    con = db()
    try:
        for table in BACKUP_TABLES:
            try:
                if table_exists(table):
                    counts[table] = con.execute(f"SELECT COUNT(*) AS n FROM {table}").fetchone()["n"]
                else:
                    counts[table] = "missing"
            except Exception:
                counts[table] = "error"
    finally:
        con.close()

    body = """
    <div class='toolbar'>
      <b>Admin Backup / Restore</b>
      <span class='small'>Download a full PMW data backup and restore it if needed.</span>
    </div>
    <div class='userform'>
      <h3>Download Backup</h3>
      <p><a class='btn green' href='/admin/export_schedule_excel'>Export Schedule to Excel</a></p>
      <p>This exports users, schedule cells, ticket emails, ticket attachments, job history, settings, and audit records into one JSON file.</p>
      <p><a class='btn green' href='/admin/download_backup'>Download Full Backup</a></p>
      <p class='small'><b>Important:</b> Render's app filesystem can be temporary. Keep downloaded backups on your PC, company drive, or cloud storage.</p>
    </div>
    <div class='userform'>
      <h3>Restore Backup</h3>
      <p>Restoring will replace the selected PMW database tables with the contents of the uploaded backup file.</p>
      <form method='post' action='/admin/restore_backup' enctype='multipart/form-data' onsubmit="return confirm('RESTORE BACKUP? This will replace current PMW data tables with the uploaded backup. Continue?')">
        <input type='file' name='backup_file' accept='.json,application/json' required>
        <p><label>Type RESTORE to confirm: <input name='confirm_text' required></label></p>
        <button class='red' type='submit'>Restore Uploaded Backup</button>
      </form>
    </div>
    <div class='userform'>
      <h3>Current Table Counts</h3>
      <table class='admin'><tr><th>Table</th><th>Rows</th></tr>
    """
    for table in BACKUP_TABLES:
        body += f"<tr><td>{html.escape(table)}</td><td>{html.escape(str(counts.get(table,'')))}</td></tr>"
    body += "</table></div>"
    return page(body)

@app.route('/admin/download_backup')
@login_required
@role_required('admin')
def download_backup():
    payload = build_backup_payload()
    data = json.dumps(payload).encode("utf-8")
    fn = "PMW_FULL_BACKUP_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".json"
    try:
        from flask import Response
        resp = Response(data, mimetype="application/json")
        resp.headers["Content-Disposition"] = f"attachment; filename={fn}"
        log('DOWNLOAD_BACKUP', fn)
        return resp
    except Exception:
        path = os.path.join(EXPORT_FOLDER, fn)
        with open(path, "wb") as f:
            f.write(data)
        return send_file(path, as_attachment=True, download_name=fn)

def reset_pg_sequence_if_needed(con, table):
    if not USE_POSTGRES:
        return
    try:
        # Only common id tables need this. Safe if sequence exists.
        con.execute(f"SELECT setval(pg_get_serial_sequence('{table}', 'id'), COALESCE((SELECT MAX(id) FROM {table}), 1), true)")
        try:
            con.commit()
        except Exception:
            pass
    except Exception:
        try:
            con.con.rollback()
        except Exception:
            pass

@app.route('/admin/restore_backup', methods=['POST'])
@login_required
@role_required('admin')
def restore_backup():
    if (request.form.get('confirm_text') or '').strip() != 'RESTORE':
        flash('Restore cancelled. You must type RESTORE exactly.')
        return redirect('/admin/backup')
    f = request.files.get('backup_file')
    if not f:
        flash('No backup file uploaded.')
        return redirect('/admin/backup')
    try:
        payload = json.load(f.stream)
    except Exception as e:
        flash('Could not read backup JSON: ' + str(e))
        return redirect('/admin/backup')

    if not isinstance(payload, dict) or "tables" not in payload:
        flash('Invalid PMW backup file.')
        return redirect('/admin/backup')

    # Ensure current tables exist before restore.
    try:
        init_db()
    except Exception:
        pass
    try:
        upgrade_schedule_settings_table()
    except Exception:
        pass
    try:
        upgrade_ticket_schedule_status_columns()
    except Exception:
        pass
    try:
        upgrade_job_history_table()
    except Exception:
        pass

    con = db()
    restored = {}
    try:
        for table in BACKUP_TABLES:
            table_payload = payload.get("tables", {}).get(table)
            if not table_payload or table_payload.get("missing"):
                continue
            if not table_exists(table):
                continue
            cols_existing = get_table_columns(con, table)
            rows = table_payload.get("rows", [])
            # Use only columns that exist in current schema.
            backup_cols = table_payload.get("columns") or []
            cols = [c for c in backup_cols if c in cols_existing]
            if not cols:
                continue

            con.execute(f"DELETE FROM {table}")
            placeholders = ",".join(["?"] * len(cols))
            col_list = ",".join(cols)
            for obj in rows:
                vals = [_restore_value(table, c, obj.get(c)) for c in cols]
                con.execute(f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})", tuple(vals))
            restored[table] = len(rows)
        con.commit()
    except Exception as e:
        try:
            con.con.rollback()
        except Exception:
            pass
        con.close()
        flash('Restore failed: ' + str(e))
        return redirect('/admin/backup')
    finally:
        try:
            con.close()
        except Exception:
            pass

    # Reset PostgreSQL sequences after explicit id inserts.
    con = db()
    try:
        for table in ["users", "ticket_links", "ticket_attachments", "audit_log", "job_history"]:
            reset_pg_sequence_if_needed(con, table)
    finally:
        con.close()

    log('RESTORE_BACKUP', ', '.join([f"{k}:{v}" for k,v in restored.items()]))
    flash('Backup restored: ' + ', '.join([f"{k}={v}" for k,v in restored.items()]))
    return redirect('/admin/backup')


@app.route('/admin/ticket_cleanup')
@login_required
@role_required('admin')
def ticket_cleanup():
    q = request.args.get('q','').strip()
    status = request.args.get('status','all').strip()
    totals = ticket_storage_totals()

    con = db()
    clauses = []
    params = []
    if q:
        clauses.append("(job_number LIKE ? OR subject LIKE ? OR sender LIKE ?)")
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]
    if status == "scheduled":
        clauses.append("COALESCE(scheduled_status,'')='scheduled'")
    elif status == "unscheduled":
        clauses.append("COALESCE(scheduled_status,'')<>'scheduled'")
    where = (" WHERE " + " AND ".join(clauses)) if clauses else ""
    rows = con.execute(f"SELECT * FROM ticket_links {where} ORDER BY id DESC LIMIT 300", tuple(params)).fetchall()
    con.close()

    body = f"""
    <div class='toolbar'>
      <b>Ticket Cleanup / Storage</b>
      <span class='small'>Admin only. Delete selected old/test ticket emails to free database space.</span>
    </div>
    <div class='userform'>
      <p><b>Total Tickets:</b> {totals['tickets']} &nbsp; <b>Attachments:</b> {totals['attachments']} &nbsp; <b>Storage:</b> {fmt_bytes_cleanup(totals['total_bytes'])}</p>
      <p><b>.msg files:</b> {fmt_bytes_cleanup(totals['msg_bytes'])} &nbsp; <b>Attachments:</b> {fmt_bytes_cleanup(totals['att_bytes'])}</p>
      <p class='small'>Deleting here removes the ticket email record, stored .msg file, and stored attachments from PostgreSQL. Existing schedule cells are not removed automatically.</p>
    </div>
    <form method='get' class='toolbar'>
      <input name='q' value='{html.escape(q, quote=True)}' placeholder='Search job, subject, sender'>
      <select name='status'>
        <option value='all' {'selected' if status=='all' else ''}>All</option>
        <option value='scheduled' {'selected' if status=='scheduled' else ''}>Scheduled</option>
        <option value='unscheduled' {'selected' if status=='unscheduled' else ''}>Not Scheduled</option>
      </select>
      <button>Filter</button>
      <a class='btn' href='/tickets'>Back to Tickets</a>
    </form>
    <form method='post' action='/admin/delete_selected_tickets' onsubmit="return confirm('Delete selected ticket emails and attachments from the database? This cannot be undone.')">
      <div class='toolbar'>
        <button class='red' type='submit'>Delete Selected</button>
        <button type='button' onclick="document.querySelectorAll('.ticketDeleteBox').forEach(x=>x.checked=true)">Select All Shown</button>
        <button type='button' onclick="document.querySelectorAll('.ticketDeleteBox').forEach(x=>x.checked=false)">Clear Selection</button>
      </div>
      <table class='admin'>
        <tr><th>Delete</th><th>ID</th><th>Status</th><th>Received</th><th>Job</th><th>Subject</th><th>Sender</th><th>Files</th><th>Size</th><th>Open</th></tr>
    """
    for r in rows:
        tid = r["id"]
        msg_size, att_size, att_count = ticket_file_size(tid)
        try:
            sched = ticket_scheduled_badge(r)
        except Exception:
            sched = ''
        try:
            cloud = r['cloud_file'] or ''
        except Exception:
            cloud = ''
        openlink = f"<a class='btn' href='/ticket_view_email/{tid}'>View</a>" if cloud else f"<a class='btn' href='/ticket_link_info/{tid}'>Info</a>"
        body += f"""<tr>
          <td><input class='ticketDeleteBox' type='checkbox' name='ticket_ids' value='{tid}'></td>
          <td>{tid}</td>
          <td>{sched}</td>
          <td>{html.escape(r['received'] or '')}</td>
          <td>{html.escape(r['job_number'] or '')}</td>
          <td>{html.escape(r['subject'] or '')}</td>
          <td>{html.escape(r['sender'] or '')}</td>
          <td>{att_count} attachment(s)</td>
          <td>{fmt_bytes_cleanup(msg_size + att_size)}</td>
          <td>{openlink}</td>
        </tr>"""
    body += "</table></form>"
    return page(body)

@app.route('/admin/delete_selected_tickets', methods=['POST'])
@login_required
@role_required('admin')
def delete_selected_tickets():
    ids = []
    for x in request.form.getlist('ticket_ids'):
        try:
            ids.append(int(x))
        except Exception:
            pass
    ids = sorted(set(ids))
    if not ids:
        flash('No tickets selected.')
        return redirect('/admin/ticket_cleanup')

    con = db()
    deleted = 0
    freed = 0
    try:
        for tid in ids:
            try:
                msg_size, att_size, att_count = ticket_file_size(tid)
                freed += msg_size + att_size
            except Exception:
                pass

            try:
                atts = con.execute("SELECT stored_filename FROM ticket_attachments WHERE ticket_id=?", (tid,)).fetchall()
                for a in atts:
                    fn = a["stored_filename"] or ""
                    if fn:
                        p = os.path.join(CLOUD_ATTACHMENT_FOLDER, fn)
                        if os.path.exists(p):
                            try: os.remove(p)
                            except Exception: pass
            except Exception:
                pass

            try:
                t = con.execute("SELECT cloud_file FROM ticket_links WHERE id=?", (tid,)).fetchone()
                if t and (t["cloud_file"] or ""):
                    p = os.path.join(CLOUD_TICKET_FOLDER, t["cloud_file"])
                    if os.path.exists(p):
                        try: os.remove(p)
                        except Exception: pass
            except Exception:
                pass

            con.execute("DELETE FROM ticket_attachments WHERE ticket_id=?", (tid,))
            con.execute("DELETE FROM ticket_links WHERE id=?", (tid,))
            deleted += 1
        con.commit()
    finally:
        con.close()

    log('DELETE_TICKETS', f'{deleted} tickets, approx {fmt_bytes_cleanup(freed)} freed')
    flash(f'Deleted {deleted} ticket(s). Approx storage freed: {fmt_bytes_cleanup(freed)}.')
    return redirect('/admin/ticket_cleanup')


@app.route('/tickets')
@login_required
def tickets():
    q=request.args.get('q','').strip()
    con=db()
    if q:
        rows=con.execute("""SELECT * FROM ticket_links WHERE job_number LIKE ? OR subject LIKE ? OR sender LIKE ?
                            ORDER BY id DESC LIMIT 200""",(f'%{q}%',f'%{q}%',f'%{q}%')).fetchall()
    else:
        status_filter=request.args.get('status','').strip(); rows=con.execute("SELECT * FROM ticket_links WHERE COALESCE(scheduled_status,'') <> 'scheduled' ORDER BY id DESC LIMIT 200").fetchall() if status_filter=='unscheduled' else con.execute("SELECT * FROM ticket_links ORDER BY id DESC LIMIT 200").fetchall()
    con.close()
    body="<div class='toolbar'><b>Imported Tickets</b><form style='display:inline-flex;gap:5px;margin-left:10px'><input name='q' value='"+html.escape(q, quote=True)+"' placeholder='Search job, subject, sender'><button>Search</button></form><span class='small'>Use Add to Numbering/Fabrication to place a linked ticket on the schedule.</span></div>"
    body+="<table class='admin'><tr><th>Status</th><th>Received</th><th>Job</th><th>Subject</th><th>Sender</th><th>Email Link</th><th>Cloud Upload</th><th>Add to Schedule</th></tr>"
    for r in rows:
        cloud_file = ''
        try:
            cloud_file = r['cloud_file'] or ''
        except Exception:
            cloud_file = ''
        if cloud_file:
            openlink = "<a class='btn green' href='/ticket_view_email/"+str(r['id'])+"'>View Email</a> <a class='btn' href='/ticket_download/"+str(r['id'])+"'>Download .msg</a> <a class='btn' href='/ticket_link_info/"+str(r['id'])+"'>Info</a>"
        else:
            openlink = "<a class='btn' href='/ticket_link_info/"+str(r['id'])+"'>Office Path</a>"
        uploadform = "<form method='post' action='/ticket_upload/"+str(r['id'])+"' enctype='multipart/form-data' style='display:flex;gap:4px;align-items:center'><input type='file' name='msgfile' accept='.msg' style='max-width:190px'><button>Upload .msg</button></form>"
        scheduled_html = ticket_scheduled_badge(r)
        try:
            already_scheduled = (r['scheduled_status'] or '') == 'scheduled'
        except Exception:
            already_scheduled = False
        if already_scheduled:
            addlinks = "<span class='small'>Already added</span> <a class='btn' href='/add_ticket_to_schedule?id="+str(r['id'])+"&side=numbering'>Add Again Numbering</a> <a class='btn' href='/add_ticket_to_schedule?id="+str(r['id'])+"&side=fabrication'>Add Again Fabrication</a>"
        else:
            addlinks = "<a class='btn green' href='/add_ticket_to_schedule?id="+str(r['id'])+"&side=numbering'>Numbering</a> <a class='btn green' href='/add_ticket_to_schedule?id="+str(r['id'])+"&side=fabrication'>Fabrication</a>"
        body += f"<tr><td>{scheduled_html}</td><td>{html.escape(r['received'] or '')}</td><td>{html.escape(r['job_number'] or '')}</td><td>{html.escape(r['subject'] or '')}</td><td>{html.escape(r['sender'] or '')}</td><td>{openlink}</td><td>{uploadform}</td><td>{addlinks}</td></tr>"
    body+="</table>"
    return page(body)

@app.route('/open_ticket_id')
@login_required
def open_ticket_id():
    tid=request.args.get('id','')
    con=db(); r=con.execute("SELECT * FROM ticket_links WHERE id=?",(tid,)).fetchone(); con.close()
    if not r:
        flash('Ticket not found.')
        return redirect('/tickets')
    ok,msg=open_path_on_this_pc(r['file_path'] or '')
    flash('Opened ticket/email.' if ok else 'Could not open ticket/email. Reason: '+msg)
    return redirect('/tickets')

@app.route('/add_ticket_to_schedule')
@login_required
@role_required('editor')
def add_ticket_to_schedule():
    ticket_id = request.args.get('id')
    side = request.args.get('side','fabrication')
    sheet = request.args.get('sheet','Fabrication Schedule')

    con=db()
    t=con.execute("SELECT * FROM ticket_links WHERE id=?",(ticket_id,)).fetchone()
    if not t:
        con.close()
        flash('Ticket not found.')
        return redirect('/tickets')

    if side == 'numbering':
        job_col = 2
    else:
        job_col = 5

    target_row = None
    for r in range(3,51):
        row=con.execute("SELECT value,link_path FROM workbook_cells WHERE sheet_name=? AND row_num=? AND col_num=?",
                        (sheet,r,job_col)).fetchone()
        if not row or ((row['value'] or '').strip()=='' and (row['link_path'] or '').strip()==''):
            target_row = r
            break

    if target_row is None:
        con.close()
        flash('No open row available on schedule.')
        return redirect('/tickets')

    subject = t['subject'] or ''
    job = t['job_number'] or ''

    cloud_file = ''
    try:
        cloud_file = t['cloud_file'] or ''
    except Exception:
        cloud_file = ''

    if cloud_file.strip():
        link_path = '/ticket_view_email/' + str(ticket_id)
        label = 'View Email + Attachments'
    else:
        local_path = t['file_path'] or ''
        link_path = local_path if local_path else ('/ticket_link_info/' + str(ticket_id))
        label = 'Office Path'

    display_text = (job + ' - ' + subject).strip(' -') if job else subject
    con.close()

    upsert_workbook_cell(
        sheet,target_row,job_col,
        display_text,
        '#fff066',
        '',
        link_path,
        label,
        '',
        '',
        '',
        session.get('username','')
    )

    mark_ticket_scheduled(ticket_id, side, sheet, target_row)
    log('ADD_TICKET_TO_SCHEDULE', f'ticket {ticket_id} -> {side} row {target_row}')
    flash('Ticket added to schedule with its email/attachment preview link.')
    return redirect('/?sheet=' + urllib.parse.quote(sheet))


@app.route('/open_link')
@login_required
def open_link():
    sheet=request.args.get('sheet','Fabrication Schedule')
    row=int(request.args.get('row','0') or 0)
    col=int(request.args.get('col','0') or 0)
    con=db()
    rec=con.execute("SELECT link_path FROM workbook_cells WHERE sheet_name=? AND row_num=? AND col_num=?",(sheet,row,col)).fetchone()
    con.close()
    if not rec or not (rec['link_path'] or '').strip():
        flash('No ticket/email link is attached to that cell.')
        return redirect('/?sheet='+urllib.parse.quote(sheet))
    link = rec['link_path'] or ''
    if link.startswith('/'):
        return redirect(link)
    ok,msg=open_path_on_this_pc(link)
    flash('Opened linked ticket/email.' if ok else 'Could not open linked ticket/email. Reason: '+msg)
    return redirect('/?sheet='+urllib.parse.quote(sheet))


@app.route('/autosave_cell', methods=['POST'])
@login_required
@role_required('editor')
def autosave_cell():
    data = request.get_json(silent=True) or {}
    sheet = data.get('sheet') or 'Fabrication Schedule'
    try:
        r = int(data.get('row'))
        c = int(data.get('col'))
    except Exception:
        return jsonify({"ok": False, "error": "Bad row/col"}), 400
    if r < 1 or r > 200 or c not in DISPLAY_COLS:
        return jsonify({"ok": False, "error": "Out of range"}), 400

    val = (data.get('value') or '').replace('\r',' ').replace('\n',' ').strip()
    bg = (data.get('bg_color') or '').strip()
    txt = (data.get('text_color') or '').strip()
    link = (data.get('link_path') or '').strip()
    label = (data.get('link_label') or '').strip()
    fsize = (data.get('font_size') or '').strip()
    bold = (data.get('bold') or '').strip()
    rich = (data.get('rich_html') or '').strip()
    loaded_at = (data.get('loaded_at') or '').strip()
    now = datetime.now().isoformat(timespec='seconds')

    con = db()
    current = con.execute("SELECT updated_at, updated_by, value FROM workbook_cells WHERE sheet_name=? AND row_num=? AND col_num=?", (sheet,r,c)).fetchone()
    current_at = ''
    current_by = ''
    if current:
        try:
            current_at = current['updated_at'] or ''
            current_by = current['updated_by'] or ''
        except Exception:
            current_at = ''
            current_by = ''

    # Conflict protection:
    # If the browser loaded an older copy of this cell, do not let it overwrite a newer database value.
    if current_at and loaded_at and current_at != loaded_at:
        con.close()
        return jsonify({
            "ok": False,
            "conflict": True,
            "error": "Cell changed by another user. Refresh required.",
            "current_updated_at": current_at,
            "current_updated_by": current_by
        }), 409

    # v46.4: If loaded_at is blank, allow the save.
    # Some existing cells/pages may not have timestamp hidden fields yet.
    # Only block when BOTH loaded_at and current_at exist and are different.

    con.execute("""INSERT OR REPLACE INTO workbook_cells(
        sheet_name,row_num,col_num,value,bg_color,text_color,link_path,link_label,font_size,bold,rich_html,updated_by,updated_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (sheet,r,c,val,bg,txt,link,label,fsize,bold,rich,session.get('username',''),now))
    con.commit()
    con.close()
    return jsonify({"ok": True, "saved_at": now})


@app.route('/users')
@login_required
@role_required('admin')
def users():
    con=db()
    rows=con.execute("SELECT id,username,role,active,created_at FROM users ORDER BY username").fetchall()
    con.close()

    role_options = "<option value='viewer'>viewer - can only view</option><option value='editor'>editor - can view and edit schedule</option><option value='admin'>admin - full access</option>"

    body = """
    <h2>Admin - Users & Permissions</h2>
    <div class='userform'>
      <h3>Create New User</h3>
      <form method='post' action='/users/create'>
        <input name='username' placeholder='username' required>
        <input name='password' placeholder='password' required>
        <select name='role'>""" + role_options + """</select>
        <button class='green'>Create User</button>
      </form>
      <p class='small'><b>viewer</b> can only view. <b>editor</b> can edit schedules/tickets. <b>admin</b> can manage users.</p>
    </div>
    <table class='admin'>
      <tr><th>Username</th><th>Role</th><th>Active</th><th>Created</th><th>Change Role</th><th>Password</th><th>Status</th></tr>
    """

    for u in rows:
        uid=u['id']
        username=html.escape(u['username'] or '')
        role=u['role'] or 'viewer'
        active='Yes' if int(u['active'] or 0)==1 else 'No'
        role_class='role-' + html.escape(role)
        selected_viewer='selected' if role=='viewer' else ''
        selected_editor='selected' if role=='editor' else ''
        selected_admin='selected' if role=='admin' else ''
        disable_self = "disabled" if u['username']==session.get('username') else ""

        body += f"""
        <tr>
          <td><b>{username}</b></td>
          <td><span class='rolebadge {role_class}'>{html.escape(role)}</span></td>
          <td>{active}</td>
          <td>{html.escape(u['created_at'] or '')}</td>
          <td>
            <form method='post' action='/users/update_role' style='display:inline-flex;gap:5px'>
              <input type='hidden' name='user_id' value='{uid}'>
              <select name='role'>
                <option value='viewer' {selected_viewer}>viewer</option>
                <option value='editor' {selected_editor}>editor</option>
                <option value='admin' {selected_admin}>admin</option>
              </select>
              <button>Update</button>
            </form>
          </td>
          <td>
            <form method='post' action='/users/reset_password' style='display:inline-flex;gap:5px'>
              <input type='hidden' name='user_id' value='{uid}'>
              <input name='password' placeholder='new password'>
              <button>Reset</button>
            </form>
          </td>
          <td>
            <form method='post' action='/users/toggle_active' style='display:inline'>
              <input type='hidden' name='user_id' value='{uid}'>
              <button {disable_self}>{'Disable' if int(u['active'] or 0)==1 else 'Enable'}</button>
            </form>
          </td>
        </tr>
        """

    body += "</table>"
    return page(body)

@app.route('/users/create', methods=['POST'])
@login_required
@role_required('admin')
def users_create():
    username=(request.form.get('username') or '').strip()
    password=(request.form.get('password') or '').strip()
    role=(request.form.get('role') or 'viewer').strip()
    if role not in ('viewer','editor','admin'):
        role='viewer'
    if not username or not password:
        flash('Username and password are required.')
        return redirect('/users')
    con=db()
    try:
        con.execute("INSERT INTO users(username,password_hash,role,active,created_at) VALUES(?,?,?,?,?)",
                    (username,generate_password_hash(password),role,1,datetime.now().isoformat(timespec='seconds')))
        con.commit()
        log('CREATE_USER', f'{username} / {role}')
        flash(f'Created user {username} as {role}.')
    except Exception as e:
        flash('Could not create user: ' + str(e))
    con.close()
    return redirect('/users')

@app.route('/users/update_role', methods=['POST'])
@login_required
@role_required('admin')
def users_update_role():
    uid=request.form.get('user_id')
    role=(request.form.get('role') or 'viewer').strip()
    if role not in ('viewer','editor','admin'):
        role='viewer'
    con=db()
    row=con.execute("SELECT username FROM users WHERE id=?",(uid,)).fetchone()
    if row and row['username']==session.get('username') and role!='admin':
        flash('You cannot remove admin from your own logged-in account.')
        con.close()
        return redirect('/users')
    con.execute("UPDATE users SET role=? WHERE id=?",(role,uid))
    con.commit()
    con.close()
    log('UPDATE_USER_ROLE', f'id {uid} -> {role}')
    flash('User role updated.')
    return redirect('/users')

@app.route('/users/reset_password', methods=['POST'])
@login_required
@role_required('admin')
def users_reset_password():
    uid=request.form.get('user_id')
    password=(request.form.get('password') or '').strip()
    if not password:
        flash('Enter a new password first.')
        return redirect('/users')
    con=db()
    con.execute("UPDATE users SET password_hash=? WHERE id=?",(generate_password_hash(password),uid))
    con.commit()
    con.close()
    log('RESET_USER_PASSWORD', f'id {uid}')
    flash('Password reset.')
    return redirect('/users')

@app.route('/users/toggle_active', methods=['POST'])
@login_required
@role_required('admin')
def users_toggle_active():
    uid=request.form.get('user_id')
    con=db()
    row=con.execute("SELECT username,active FROM users WHERE id=?",(uid,)).fetchone()
    if not row:
        flash('User not found.')
        con.close()
        return redirect('/users')
    if row['username']==session.get('username'):
        flash('You cannot disable your own logged-in account.')
        con.close()
        return redirect('/users')
    new_active=0 if int(row['active'] or 0)==1 else 1
    con.execute("UPDATE users SET active=? WHERE id=?",(new_active,uid))
    con.commit()
    con.close()
    log('TOGGLE_USER_ACTIVE', f'id {uid} -> {new_active}')
    flash('User status updated.')
    return redirect('/users')


@app.route('/audit')
@login_required
@role_required('admin')
def audit():
    con=db(); rows=con.execute('SELECT * FROM audit_log ORDER BY id DESC LIMIT 200').fetchall(); con.close()
    body='<table class="admin"><tr><th>When</th><th>User</th><th>Action</th><th>Details</th></tr>'
    for r in rows: body += f"<tr><td>{r['created_at']}</td><td>{html.escape(r['username'] or '')}</td><td>{r['action']}</td><td>{html.escape(r['details'] or '')}</td></tr>"
    return page(body+'</table>')






def upgrade_ticket_db_file_storage():
    """Store uploaded .msg and extracted attachments in DB so Render restarts do not lose files."""
    try:
        con=db(); cur=con.cursor()
        if USE_POSTGRES:
            try:
                cur.execute("ALTER TABLE ticket_links ADD COLUMN cloud_file_data BYTEA")
            except Exception:
                try: con.con.rollback()
                except Exception: pass
            try:
                cur.execute("ALTER TABLE ticket_attachments ADD COLUMN file_data BYTEA")
            except Exception:
                try: con.con.rollback()
                except Exception: pass
        else:
            try:
                cur.execute("ALTER TABLE ticket_links ADD COLUMN cloud_file_data BLOB")
            except Exception:
                pass
            try:
                cur.execute("ALTER TABLE ticket_attachments ADD COLUMN file_data BLOB")
            except Exception:
                pass
        con.commit(); con.close()
    except Exception as e:
        print("DB file storage upgrade failed:", repr(e))

def read_file_bytes(path):
    with open(path, 'rb') as f:
        return f.read()

def send_bytes_as_file(data, filename, mimetype=None, as_attachment=False):
    from flask import send_file
    if data is None:
        flash('Stored file data is missing.')
        return redirect('/tickets')
    if isinstance(data, memoryview):
        data = data.tobytes()
    return send_file(BytesIO(data), as_attachment=as_attachment, download_name=filename, mimetype=mimetype or 'application/octet-stream')


def upgrade_ticket_attachment_tables():
    try:
        con=db(); cur=con.cursor()
        if USE_POSTGRES:
            cur.execute("""CREATE TABLE IF NOT EXISTS ticket_attachments(
                id SERIAL PRIMARY KEY,
                ticket_id INTEGER,
                original_filename TEXT,
                stored_filename TEXT,
                content_type TEXT,
                size_bytes INTEGER DEFAULT 0,
                uploaded_at TEXT,
                uploaded_by TEXT
            )""")
        else:
            cur.execute("""CREATE TABLE IF NOT EXISTS ticket_attachments(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_id INTEGER,
                original_filename TEXT,
                stored_filename TEXT,
                content_type TEXT,
                size_bytes INTEGER DEFAULT 0,
                uploaded_at TEXT,
                uploaded_by TEXT
            )""")
        con.commit(); con.close()
    except Exception as e:
        print("Ticket attachment table upgrade failed:", repr(e))

def save_ticket_attachment_record(ticket_id, original, stored, content_type, size_bytes, file_data=None):
    con=db()
    con.execute("""INSERT INTO ticket_attachments(ticket_id,original_filename,stored_filename,content_type,size_bytes,uploaded_at,uploaded_by,file_data)
                   VALUES(?,?,?,?,?,?,?,?)""",
                (ticket_id, original, stored, content_type, int(size_bytes or 0), datetime.now().isoformat(timespec='seconds'), session.get('username',''), file_data))
    con.commit(); con.close()

def extract_msg_attachments(ticket_id, msg_path):
    """Extract attachments from .msg into cloud_ticket_attachments folder."""
    saved = []
    if extract_msg is None:
        return saved
    try:
        msg = extract_msg.Message(msg_path)
        attachments = getattr(msg, 'attachments', []) or []
        for idx, att in enumerate(attachments, start=1):
            try:
                raw_name = getattr(att, 'longFilename', None) or getattr(att, 'shortFilename', None) or f"attachment_{idx}"
                original = secure_filename(raw_name) or f"attachment_{idx}"
                stored = f"ticket_{ticket_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{idx}_{original}"
                path = os.path.join(CLOUD_ATTACHMENT_FOLDER, stored)

                data = None
                if hasattr(att, 'data'):
                    data = att.data
                if data:
                    with open(path, 'wb') as f:
                        f.write(data)
                else:
                    # fallback: extract-msg attachment object can save itself
                    try:
                        att.save(customPath=CLOUD_ATTACHMENT_FOLDER, customFilename=stored)
                    except TypeError:
                        att.save(customPath=CLOUD_ATTACHMENT_FOLDER)
                        # If saved with original name, rename if possible
                        possible = os.path.join(CLOUD_ATTACHMENT_FOLDER, original)
                        if os.path.exists(possible):
                            os.replace(possible, path)

                if os.path.exists(path):
                    ctype = mimetypes.guess_type(original)[0] or 'application/octet-stream'
                    size = os.path.getsize(path)
                    save_ticket_attachment_record(ticket_id, original, stored, ctype, size, read_file_bytes(path))
                    saved.append({"filename": original, "stored": stored, "content_type": ctype, "size": size})
            except Exception as e:
                print("Attachment extract failed:", repr(e))
        try:
            msg.close()
        except Exception:
            pass
    except Exception as e:
        print("Extract attachments failed:", repr(e))
    return saved


def upgrade_ticket_preview_columns():
    try:
        con=db(); cur=con.cursor()
        for coldef in [
            "preview_subject TEXT DEFAULT ''",
            "preview_sender TEXT DEFAULT ''",
            "preview_date TEXT DEFAULT ''",
            "preview_body TEXT DEFAULT ''",
            "preview_status TEXT DEFAULT ''"
        ]:
            try:
                cur.execute("ALTER TABLE ticket_links ADD COLUMN " + coldef)
            except Exception:
                try:
                    con.con.rollback()
                except Exception:
                    pass
        con.commit(); con.close()
    except Exception as e:
        print("Ticket preview column upgrade failed:", repr(e))

def parse_msg_preview(msg_path):
    if extract_msg is None:
        return {
            "subject": "",
            "sender": "",
            "date": "",
            "body": "Email preview library is not available on this server yet. You can still download the original .msg file.",
            "status": "extract_msg not installed"
        }
    try:
        msg = extract_msg.Message(msg_path)
        subject = msg.subject or ""
        sender = msg.sender or ""
        date = str(msg.date or "")
        body = msg.body or ""
        if not body.strip():
            body = "No readable body text was found in this .msg file. Download the original .msg file to open in Outlook."
        try:
            msg.close()
        except Exception:
            pass
        return {"subject": subject, "sender": sender, "date": date, "body": body, "status": "ok"}
    except Exception as e:
        return {
            "subject": "",
            "sender": "",
            "date": "",
            "body": "Could not parse this .msg file for browser preview. Download the original .msg file to open in Outlook. Error: " + str(e),
            "status": "parse failed"
        }


def upgrade_ticket_cloud_columns():
    try:
        con=db(); cur=con.cursor()
        for coldef in [
            "cloud_file TEXT DEFAULT ''",
            "cloud_filename TEXT DEFAULT ''",
            "cloud_uploaded_at TEXT DEFAULT ''",
            "cloud_uploaded_by TEXT DEFAULT ''"
        ]:
            try:
                cur.execute("ALTER TABLE ticket_links ADD COLUMN " + coldef)
            except Exception:
                try:
                    con.con.rollback()
                except Exception:
                    pass
        con.commit(); con.close()
    except Exception as e:
        print("Ticket cloud column upgrade failed:", repr(e))



# ===== SCHEDULE DATE CONTROL v45.1 =====
def upgrade_schedule_settings_table():
    try:
        con = db()
        cur = con.cursor()
        cur.execute("""CREATE TABLE IF NOT EXISTS schedule_settings(
            sheet_name TEXT PRIMARY KEY,
            schedule_date TEXT DEFAULT '',
            auto_today INTEGER DEFAULT 1,
            updated_by TEXT DEFAULT '',
            updated_at TEXT DEFAULT ''
        )""")
        con.commit()
        con.close()
    except Exception as e:
        print("Schedule settings upgrade failed:", repr(e))

def pretty_schedule_date(raw):
    try:
        dt = datetime.strptime(raw, "%Y-%m-%d")
        return dt.strftime("%A, %B %d, %Y").replace(" 0", " ")
    except Exception:
        return raw or datetime.now().strftime("%A, %B %d, %Y").replace(" 0", " ")

def get_schedule_date_settings(sheet):
    try:
        upgrade_schedule_settings_table()
        con = db()
        row = con.execute("SELECT schedule_date, auto_today FROM schedule_settings WHERE sheet_name=?", (sheet,)).fetchone()
        con.close()
    except Exception:
        row = None

    today = datetime.now().strftime("%Y-%m-%d")

    if not row:
        schedule_date = today
        auto_today = 1
    else:
        try:
            auto_today = int(row["auto_today"] or 0)
        except Exception:
            auto_today = 1
        schedule_date = (row["schedule_date"] or "").strip()
        if auto_today or not schedule_date:
            schedule_date = today

    return {
        "schedule_date": schedule_date,
        "auto_today": auto_today,
        "display_date": pretty_schedule_date(schedule_date)
    }


def startup_init_for_cloud():
    try:
        print("PMW DB MODE:", "PostgreSQL" if USE_POSTGRES else "SQLite", "DATABASE_URL set:", bool(DATABASE_URL), "psycopg:", bool(psycopg))
        init_db()
        upgrade_job_history_table()
        upgrade_ticket_schedule_status_columns()
        upgrade_schedule_settings_table()
        upgrade_ticket_cloud_columns()
        upgrade_ticket_preview_columns()
        upgrade_ticket_attachment_tables()
        upgrade_ticket_db_file_storage()

        con=db()
        try:
            row = con.execute("SELECT COUNT(*) AS n FROM workbook_cells").fetchone()
            n = row["n"] if row else 0
        except Exception:
            n = 0
        con.close()

        starter = os.path.join(APP_DIR, 'Ticket +Fabrication-ACTIVE(1).xlsm')
        if n == 0 and os.path.exists(starter):
            with app.test_request_context('/'):
                session['username'] = 'system'
                try:
                    import_workbook(starter)
                    print("Starter workbook imported.")
                except Exception as e:
                    print("Starter workbook import skipped:", repr(e))
    except Exception as e:
        print("Startup database initialization failed:", repr(e))

startup_init_for_cloud()

if __name__ == '__main__':
    init_db()
    con=db(); n=con.execute('SELECT COUNT(*) FROM workbook_cells').fetchone()[0]; con.close()
    starter=os.path.join(APP_DIR,'Ticket +Fabrication-ACTIVE(1).xlsm')
    if n == 0 and os.path.exists(starter):
        with app.test_request_context('/'):
            session['username']='system'
            try: import_workbook(starter)
            except Exception as e: print('Starter import skipped:',e)
    print('====================================================')
    print('PMW Ticket + Fabrication APP v51 Plain Cell Edit Arrows')
    print('Open http://127.0.0.1:5050')
    print('====================================================')
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5050)), debug=False)
